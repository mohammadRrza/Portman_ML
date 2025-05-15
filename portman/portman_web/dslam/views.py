import random

from django.shortcuts import render
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.db import connection
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from datetime import date, timedelta, datetime
from django.db.models import Count
from django.core.serializers import serialize
from django.db.models import Value
from django.db.models.functions import Concat
from django.core.serializers.json import DjangoJSONEncoder
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.postgres.search import SearchVector
from django.http import StreamingHttpResponse
from rest_framework import pagination
from classes.mailer import Mail
from classes.mailer import Ticket
from users.models import User
from rest_framework.parsers import FileUploadParser

from classes.portman_logging import PortmanLogging

from .models import Rented_port, PortmanZabbixHosts
from .services.get_snmp_port_trrafic import GetPortTrafficService
from .services.get_data_from_IBS import GetDataFromIBS
from .services.get_user_info_from_partak import GetDataFromPartak

"""from rtkit.resource import RTResource
from rtkit.resource import RTResource
from rtkit.authenticators import BasicAuthenticator, CookieAuthenticator
from rtkit.errors import RTResourceError"""

import re
import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import urllib.request, urllib.parse, urllib.error
from requests.auth import HTTPDigestAuth
from rest_framework.response import Response
from rest_framework import status, views, mixins, viewsets, permissions
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from requests.auth import HTTPBasicAuth
from users.permissions.dslam_permission import DSLAMView, DSLAMEdit
from users.permissions.dslamport_permission import DSLAMPortView, DSLAMPortEdit
from users.permissions.command_permission import CommandView, CommandEdit
from users.permissions.telecom_center_permission import TelecomCenterView, TelecomCenterEdit
import requests, base64
from khayyam import JalaliDatetime
import csv, time
import simplejson as json
import io as StringIO
from collections import defaultdict
from dslam import utility
import logging
import sys, os
import jwt
from classes.utils import PortmanUtils
from classes.base_permissions import ADMIN, SUPPORT, DIRECTRESELLER, RESELLER, SUPPORT_ADMIN, BASIC

from dslam.models import DSLAM, TelecomCenter, DSLAMPort, DSLAMPortSnapshot, LineProfile, TelecomCenterLocation, Vlan, \
    DSLAMBulkCommandResult, \
    TelecomCenterMDF, DSLAMEvent, DSLAMPortEvent, Reseller, ResellerPort, CustomerPort, DSLAMLocation, \
    DSLAMICMPSnapshot, DSLAMICMP, \
    PortCommand, ResellerPort, City, Command, DSLAMType, Terminal, DSLAMStatusSnapshot, DSLAMStatus, \
    DSLAMCommand, CityLocation, MDFDSLAM, DSLAMPortVlan, \
    DSLAMPortMac, DSLAMBoard, DSLAMFaultyConfig, DSLAMPortFaulty, DSLAMTypeCommand, DSLAMCart
from dslam.permissions import HasAccessToDslam, IsAdminUser, HasAccessToDslamPort, \
    HasAccessToDslamPortSnapshot
from dslam.serializers import *
from users.helpers import add_audit_log
from .command_recognise import command_recognise
import openpyxl
from .services.fiber_home_get_card import FiberHomeGetCardStatusService
from .services.add_to_pishgaman import AddToPishgamanService
from .services.check_status import check_status
from .services.get_snmp_port_status import GetPortStatusParamsService
from .services.mini_services import retrieve_allowed_dslams
from dslam.dslam_mail import Mail
from contact.models import Order
from config.settings import DSLAM_BACKUP_PATH


class LargeResultsSetPagination(PageNumberPagination):
    page_size = 1000
    page_size_query_param = 'page_size'
    max_page_size = max


class VlanViewSet(mixins.ListModelMixin,
                  mixins.CreateModelMixin,
                  mixins.RetrieveModelMixin,
                  mixins.UpdateModelMixin,
                  mixins.DestroyModelMixin,
                  viewsets.GenericViewSet):
    permission_classes = (IsAuthenticated,)
    serializer_class = VlanSerializer
    queryset = Vlan.objects.all()

    def get_queryset(self):
        queryset = self.queryset
        reseller_id = self.request.query_params.get('reseller_id', None)
        reseller_name = self.request.query_params.get('reseller_name', None)
        dslam_id = self.request.query_params.get('dslam_id', None)
        vlan_name = self.request.query_params.get('vlan_name', None)
        vlan_id = self.request.query_params.get('vlan_id', None)
        sort_field = self.request.query_params.get('sort_field', None)

        if reseller_id:
            queryset = queryset.filter(reseller__id=reseller_id)

        if reseller_name:
            resellers = Reseller.objects.filter(name__istartswith=reseller_name)
            queryset = queryset.filter(reseller__in=resellers)

        if vlan_name:
            queryset = queryset.filter(vlan_name__istartswith=vlan_name)

        if vlan_id:
            queryset = queryset.filter(vlan_id=vlan_id)

        if dslam_id:
            port_ids = DSLAMPort.objects.filter(dslam__id=dslam_id).values_list('id', flat=True)
            vlan_obj_ids = DSLAMPortVlan.objects.filter(port__id__in=port_ids).values_list('vlan__vlan_id', flat=True)
            queryset = queryset.filter(vlan_id__in=vlan_obj_ids)

        if sort_field:
            queryset = queryset.order_by(sort_field)

        return queryset

    def create(self, request, *args, **kwargs):
        self.user = request.user
        data = request.data
        vlan_name = data.get('vlan_name')
        if not vlan_name:
            data['vlan_name'] = data.get('vlan_id')
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        description = 'ADD Vlan {0}'.format(serializer.data['vlan_id'])
        add_audit_log(request, 'Vlan', serializer.data['id'], 'Create Vlan', description)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(methods=['POST'], detail=False)
    def assign_to_reseller(self, request):
        user = request.user
        data = request.data
        reseller_id = data.get('reseller_id')
        vlan_id = data.get('vlan_id')
        if reseller_id:
            identifier_keys = ResellerPort.objects.filter(reseller__id=reseller_id).values_list('identifier_key',
                                                                                                flat=True)
            dslam_ids = MDFDSLAM.objects.filter(identifier_key__in=identifier_keys).values_list('dslam_id',
                                                                                                flat=True).distinct()

        params = {
            "is_queue": False,
            "type": "dslam",
            "vlan_id": vlan_id,
            "vlan_name": data.get('vlan_name', vlan_id),
            "username": user.username
        }

        for dslam_id in dslam_ids:
            if params.get('vlan_id'):
                result = utility.dslam_port_run_command(mdf_dslam_obj.dslam_id, 'create vlan', params)

        reseller = Reseller.objects.get(id=reseller_id)
        vlan = Vlan.objects.get(vlan_id=vlan_id)
        vlan.reseller = reseller
        vlan.save()
        serializer = self.get_serializer(vlan)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
        # return Response(data={
        #    'id': vlan.id,
        #    'vlan_id': vlan.vlan_id,
        #    'vlan_name': vlan.vlan_name,
        #    'reseller': reseller.id
        #    }, status=status.HTTP_201_CREATED, headers=headers)

    @action(methods=['POST'], detail=False)
    def assign_to_subscibers(self, request):
        data = request.data
        user = request.user
        new_vlan_id = data.get('new_vlan_id')
        vlan_obj = Vlan.objects.get(vlan_id=new_vlan_id)
        reseller_vlan = data.get('reseller_vlan')
        identifier_key = data.get('identifier_key')
        card_ports = data.get('card_port')

        dslam_ports = defaultdict(list)
        print('*************************************************')
        print(data)
        if reseller_vlan:
            flag = reseller_vlan.get('flag')
            reseller_id = reseller_vlan.get('reseller_id')
            mdf_ports = None

            if not flag:
                return Response({'result': 'please send flag item.'}, status=status.HTTP_400_BAD_REQUEST)

            identifier_keys = ResellerPort.objects.filter(reseller__id=reseller_id).values_list('identifier_key',
                                                                                                flat=True)

            if flag == 'all':
                mdf_ports = MDFDSLAM.objects.filter(identifier_key__in=identifier_key).values('dslam_id', 'slot_number',
                                                                                              'port_number')
                for mdf_port in mdf_ports:
                    try:
                        port = DSLAMPort.objects.get(
                            dslam__id=mdf_port.get('dslam_id'),
                            slot_number=mdf_port.get('slot_number'),
                            port_number=mdf_port.get('port_number')).values('id', 'slot_number', 'port_number',
                                                                            'port_index')
                        dslam_ports[mdf_port.get('dslam_id')].append(port)
                    except:
                        pass

            elif flag == 'vlan':
                vlan_id = Vlan.objects.get(vlan_id=reseller_vlan.get('vlan_id')).vlan_id
                ports = DSLAMPortVlan.objects.filter(vlan__vlan_id=vlan_id).values('port__id', 'port__dslam_id',
                                                                                   'port__slot_number',
                                                                                   'port__port_number',
                                                                                   'port__port_index')
                for port in ports:
                    dslam_ports[port.get('port__dslam_id')].append({
                        'port_number': port.get('port__port_number'),
                        'slot_number': port.get('port__slot_number'),
                        'port_index': port.get('port__port_index'),
                        'id': port.get('port__id')
                    })

            elif flag == 'no-vlan':
                identifier_keys = ResellerPort.objects.filter(reseller__id=reseller_id).values_list('identifier_key',
                                                                                                    flat=True)
                mdf_ports = set(
                    MDFDSLAM.objects.filter(identifier_key__in=identifier_key).values_list('dslam_id', 'slot_number',
                                                                                           'port_number'))
                vlan_obj_ids = Vlan.objects.filter(reseller__id=reseller_id).values_list('vlan__id', flat=True)
                ports = set(DSLAMPortVlan.objects.filter(vlan__id__in=vlan_obj__ids).values_list('port__dslam_id',
                                                                                                 'port__slot_number',
                                                                                                 'port__port_number'))
                reseller_ports_no_vlan = ports - mdf_ports

                for dslam_id, slot_number, port_index in reseller_ports_no_vlan:
                    try:
                        port = DSLAMPort.objects.get(
                            dslam__id=dslam_id,
                            slot_number=slot_number,
                            port_number=port_number).values('id', 'slot_number', 'port_number', 'port_index')
                        dslam_ports[dslam_id].append(port)
                    except:
                        pass


        elif identifier_key:
            mdf_ports = MDFDSLAM.objects.filter(identifier_key=identifier_key).values('dslam_id', 'slot_number',
                                                                                      'port_number')
            for mdf_port in mdf_ports:
                try:
                    port = DSLAMPort.objects.get(
                        dslam__id=mdf_port.get('dslam_id'),
                        slot_number=mdf_port.get('slot_number'),
                        port_number=mdf_port.get('port_number')).values('id', 'slot_number', 'port_number',
                                                                        'port_index')
                    dslam_ports[mdf_port.get('dslam_id')].append(port)
                except:
                    pass

        elif card_ports:
            print()
            '-------++====++++------------------------------------'
            print()
            card_ports
            print()
            '-------------------------------------------'
            port = DSLAMPort.objects.get(
                dslam__id=card_ports.get('dslam_id'),
                slot_number=card_ports.get('slot_number'),
                port_number=card_ports.get('port_number'))
            dslam_ports[card_ports.get('dslam_id')].append({
                'id': port.id,
                'slot_number': port.slot_number,
                'port_number': port.port_number})

        if len(dslam_ports) > 0:
            for dslam_id, port_indexes in list(dslam_ports.items()):
                params = {
                    "is_queue": False,
                    "type": "dslam",
                    "dslam_id": dslam_id,
                    "vlan_id": vlan_obj.vlan_id,
                    "vlan_name": vlan_obj.vlan_name,
                    "username": user.username
                }
                result = utility.dslam_port_run_command(dslam_id, 'create vlan', params)
                print('////////////////////////////////////////////////////////')
                params = {
                    "type": "dslamport",
                    "is_queue": False,
                    "vlan_id": vlan_obj.vlan_id,
                    "vlan_name": vlan_obj.vlan_name,
                    "dslam_id": dslam_id,
                    "port_indexes": port_indexes,
                    "username": user.username
                }
                result = utility.dslam_port_run_command(dslam_id, 'add to vlan', params)
                for port in port_indexes:
                    print()
                    port.get('id')
                    port_vlan = DSLAMPortVlan.objects.get(port__id=port.get('id'))
                    port_vlan.vlan = vlan_obj
                    port_vlan.save()
            print('-=-=-=-=-=-=-=-=-=')

            return Response({'result': 'vlans assign to subscribers'}, status=status.HTTP_201_CREATED)
        else:
            print('afsfsfafsdfgawsfgsadgfasdgfadgfbadfkgjadpidbgjiadfgoiuo')
            return Response({'error': 'port does not existed!'}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        data = request.data
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        if data.get('params') and data.get('dslam_id'):
            params['username'] = request.user.username
            result = utility.dslam_port_run_command(dslam_obj.pk, 'create vlan', params)
        self.perform_update(serializer)
        description = 'Update Vlan {0}: {1}'.format(instance.vlan_name, request.data)
        add_audit_log(request, 'Vlan', instance.id, 'Update Vlan', description)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        data = request.data
        instance = self.get_object()
        self.perform_destroy(instance)
        description = 'Delete vlan {0}'.format(instance.vlan_id)
        add_audit_log(request, 'Vlan', instance.id, 'Delete Vlan', description)
        return Response(status=status.HTTP_204_NO_CONTENT)


class DSLAMPortVlanViewSet(mixins.ListModelMixin,
                           mixins.CreateModelMixin,
                           mixins.RetrieveModelMixin,
                           mixins.UpdateModelMixin,
                           mixins.DestroyModelMixin,
                           viewsets.GenericViewSet):
    permission_classes = (IsAuthenticated,)
    serializer_class = DSLAMPortVlanSerializer
    queryset = DSLAMPortVlan.objects.all()

    def get_queryset(self):
        dslamport_id = self.request.query_params.get('dslamport_id', None)
        sort_field = self.request.query_params.get('sort_field', None)
        queryset = self.queryset

        if dslamport_id:
            queryset = queryset.filter(port__id=dslamport_id)

        if sort_field:
            queryset = queryset.order_by(sort_field)

        return queryset

    def create(self, request, *args, **kwargs):
        data = request.data
        if data.get('params') and data.get('dslam_id'):
            result = utility.dslam_port_run_command(data.get('dslam_id'), 'add to vlan', data.get('params'))

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        description = 'ADD port {0} to Vlan {1}'.format(serializer.data['vlan_id'], data.get('params'))
        add_audit_log(request, 'DSLAMPortVlan', serializer.data['id'], 'Add port to vlan', description)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def destroy(self, request, *args, **kwargs):
        data = request.data
        if data.get('dslam_id') and data.get('params'):
            result = utility.dslam_port_run_command(dslam_obj.pk, 'delete from vlan', params)
        instance = self.get_object()
        self.perform_destroy(instance)
        description = 'Delete DSLAMPortVlan {0}: {1}'.format(instance.name, request.data)
        add_audit_log(request, 'DSLAMPortVlan', instance.id, 'Delete DSLAMPortVlan', description)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        data = request.data
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        if data.get('params') and data.get('dslam_id'):
            result = utility.dslam_port_run_command(dslam_obj.pk, 'add to vlan', params)
        self.perform_update(serializer)
        description = 'Update DSLAMPortVlan {0}: {1}'.format(instance.name, request.data)
        add_audit_log(request, 'DSLAMPortVlan', instance.id, 'Update DSLAMPort Vlan', description)
        return Response(serializer.data)


class TerminalViewSet(mixins.ListModelMixin,
                      viewsets.GenericViewSet):
    permission_classes = (IsAuthenticated, HasAccessToDslam)
    serializer_class = TerminalSerializer
    queryset = Terminal.objects.all()
    paginate_by = None
    paginate_by_param = None
    paginator = None


class LineProfileViewSet(mixins.ListModelMixin,
                         mixins.CreateModelMixin,
                         mixins.DestroyModelMixin,
                         viewsets.GenericViewSet):
    permission_classes = (IsAuthenticated, HasAccessToDslam)
    serializer_class = LineProfileSerializer

    @action(methods=['GET'], detail=False)
    def assign_to_port(self, request):
        user = request.user
        data = request.data
        params = data.get('params')
        result = utility.dslam_port_run_command(dslam_id, 'change lineprofile port', params)
        if 'error' not in result['result']:
            dslamport = DSLAMPort.objects.get(dslam__id=dslam_id, slot_number=slot_number, port_number=port_number)
            serializer = DSLAMPortSerializer(dslamport)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(result, status=status.HTTP_400_BAD_REQUEST)

    def get_queryset(self):
        queryset = LineProfile.objects.all()
        profile_name = self.request.query_params.get('profile_name', None)
        sort_field = self.request.query_params.get('sort_field', None)

        if profile_name:
            queryset = queryset.filter(name__istartswith=profile_name)

        if sort_field:
            queryset = queryset.order_by(sort_field)

        return queryset

    def create(self, request, *args, **kwargs):
        data = request.data
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        extra_settings = data.get('extra_settings_info')
        lst_profile_settings = []

        for setting in extra_settings:
            if setting.get('value'):
                lst_profile_settings.append(setting)

        if bool(lst_profile_settings):
            line_profile = LineProfile.objects.get(id=serializer.data.get('id'))
            for setting in lst_profile_settings:
                lineprofile_settings = LineProfileExtraSettings()
                lineprofile_settings.line_profile = line_profile
                lineprofile_settings.attr_name = setting.get('key')
                lineprofile_settings.attr_value = setting.get('value')
                lineprofile_settings.save()
        description = 'Create LineProfile : {0}'.format(serializer.data['name'])
        add_audit_log(request, 'LineProfile', serializer.data['id'], 'Create LineProfile', description)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        description = 'Delete LineProfile : {0}'.format(instance.name)
        add_audit_log(request, 'LineProfile', instance.id, 'Delete LineProfile', description)
        return Response(status=status.HTTP_204_NO_CONTENT)

    """
    Update a model instance.
    """

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        description = 'Delete LineProfile : {0}'.format(instance.name)
        add_audit_log(request, 'LineProfile', serializer.data['id'], 'Update LineProfile', description)
        return Response(serializer.data)


class TelecomCenterMDFViewSet(mixins.ListModelMixin,
                              mixins.CreateModelMixin,
                              mixins.RetrieveModelMixin,
                              mixins.UpdateModelMixin,
                              mixins.DestroyModelMixin,
                              viewsets.GenericViewSet):
    paginate_by = None
    paginate_by_param = None
    paginator = None
    permission_classes = (IsAuthenticated,)
    serializer_class = TelecomCenterMDFSerializer

    def get_queryset(self):
        queryset = TelecomCenterMDF.objects.all().order_by('priority')
        telecom_id = self.request.query_params.get('telecom_id', None)
        sort_field = self.request.query_params.get('sort_field', None)

        if telecom_id:
            queryset = queryset.filter(telecom_center__id=telecom_id)

        if sort_field:
            queryset = queryset.order_by(sort_field)

        return queryset

    """
    Create a model instance.
    """

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        description = 'Create Telecom Center MDF: {0}'.format(serializer.data.get('name'))
        add_audit_log(request, 'TelecomCenterMDF', serializer.data['id'], 'Create Telecome Center MDF', description)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    """
    Update a model instance.
    """

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        description = 'Update Telecom Center MDF {0}: {1}'.format(instance.id, request.data)
        add_audit_log(request, 'TelecomCenterMDF', instance.id, 'Update Telecom Center MDF', description)
        return Response(serializer.data)

    """
    Destroy a model instance.
    """

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        description = 'Delete Telecom Center MDF: {0}'.format(instance.id)
        add_audit_log(request, 'TelecomCenterMDF', instance.id, 'Delete Telecom Center MDF', description)
        return Response(status=status.HTTP_204_NO_CONTENT)


class DSLAMViewSet(mixins.ListModelMixin,
                   mixins.CreateModelMixin,
                   mixins.RetrieveModelMixin,
                   mixins.UpdateModelMixin,
                   mixins.DestroyModelMixin,
                   viewsets.GenericViewSet):
    queryset = DSLAM.objects.all()
    queryset = queryset.exclude(name__icontains='collected').order_by('-id')
    # permission_classes = (IsAuthenticated, DSLAMView, DSLAMEdit)
    serializer_class = DSLAMSerializer

    # pagination_class = LargeResultsSetPagination

    def get_serializer(self, *args, **kwargs):
        # if self.request.user.is_superuser or self.request.user.has_permission('edit_dslam'):
        if self.request.user.is_superuser or self.request.user.type_contains([ADMIN, SUPPORT_ADMIN]):
            return DSLAMSerializer(request=self.request, *args, **kwargs)
        elif self.request.user.type_contains(SUPPORT):
            _fields = ['set_snmp_community', 'get_snmp_community', 'snmp_port'] # 'telnet_password', 'telnet_username'
            return DSLAMSerializer(request=self.request, remove_fields=_fields, *args, **kwargs)
        else:
            _fields = ['telnet_password', 'telnet_username', 'set_snmp_community', 'get_snmp_community', 'snmp_port',
                       'ip', 'total_ports_count', 'down_ports_count', 'up_ports_count']
            return DSLAMSerializer(request=self.request, remove_fields=_fields, *args, **kwargs)

    @action(methods=['GET'], detail=True)
    def current(self, request):
        serializer = DSLAMSerializer(request.user, request=request)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(methods=['GET'], detail=False)
    def get_queryset(self):
        queryset = self.queryset
        user = self.request.user
        payload_token = PortmanUtils.decode_token(self.request)
        username = payload_token['username']
        sort_field = self.request.query_params.get('sort_field', None)
        dslam_name = self.request.query_params.get('search_dslam', None)
        ip = self.request.query_params.get('search_ip', None)
        ip_list = self.request.query_params.get('search_ip_list', None)
        city_id = self.request.query_params.get('search_city', None)
        city_name = self.request.query_params.get('city_name', None)
        province_name = self.request.query_params.get('province_name', None)
        telecom = self.request.query_params.get('search_telecom', None)
        telecom_name = self.request.query_params.get('telecom_name', None)
        active = self.request.query_params.get('search_active', None)
        status = self.request.query_params.get('search_status', None)
        dslam_type_id = self.request.query_params.get('search_type', None)
        ldap_login = self.request.query_params.get('ldap_login', None)

        page_size = int(self.request.query_params.get('page_size', 10))
        if page_size < 1 or page_size > 30:
            page_size = 10
        pagination.PageNumberPagination.page_size = page_size


        '''
        if username and ldap_login == 'false':
            user_type = User.objects.get(username=username).type
        if username and user_type != 'ADMIN' and ldap_login == 'false':
            user_type = User.objects.get(username=username).type
            user_id = User.objects.get(username=username).id
            model_user = User()
            model_user.type = user_type
            model_user.set_user_id(user_id)
            allowed_dslams = model_user.get_allowed_dslams()
            if user_type == 'SUPPORT':
                queryset = queryset.filter(id__in=allowed_dslams)
            if user_type == 'DIRECTRESELLER':
                queryset = queryset.filter(id__in=allowed_dslams) '''

        if username and ldap_login == 'false':
            user = User.objects.filter(username=username).first()
            if user and user.type_contains(ADMIN) == False:
                if user.type_contains([DIRECTRESELLER, RESELLER]):
                    # model_user = User(type=user.type)
                    # model_user.set_user_id(user.id)
                    allowed_dslams = user.get_allowed_dslams()
                    queryset = queryset.filter(id__in=allowed_dslams)

        if dslam_type_id:
            queryset = queryset.filter(dslam_type__id=dslam_type_id)

        if dslam_name:
            queryset = queryset.filter(name__icontains=dslam_name)

        if ip:
            ip = ip.strip()
            if len(ip.split('.')) != 4:
                queryset = queryset.filter(ip__icontains=ip)
            else:
                queryset = queryset.filter(ip=ip)

        if ip_list:
            for ip in ip_list.split(','):
                queryset = queryset.filter(ip__istartswith=ip)

        if status:
            queryset = queryset.filter(status=status)

        if active:
            queryset = queryset.filter(active=bool(active))

        if province_name:
            province_objs = City.objects.filter(name__icontains=province_name, parent=None).values_list('id', flat=True)
            city_ids = City.objects.filter(parent__id__in=province_objs).values_list('id', flat=True)
            telecom_ids = TelecomCenter.objects.filter(city__id__in=city_ids).values_list('id', flat=True)
            queryset = queryset.filter(telecom_center__id__in=telecom_ids)

        if city_name:
            city_ids = City.objects.filter(name__icontains=city_name).values_list('id', flat=True)
            telecom_ids = TelecomCenter.objects.filter(city__id__in=city_ids).values_list('id', flat=True)
            queryset = queryset.filter(telecom_center__id__in=telecom_ids)

        if city_id:
            city = City.objects.get(id=city_id)
            if city.parent is None:
                city_ids = City.objects.filter(parent=city).values_list('id', flat=True)
                telecom_ids = TelecomCenter.objects.filter(city__id__in=city_ids).values_list('id', flat=True)
            else:
                telecom_ids = TelecomCenter.objects.filter(city=city).values_list('id', flat=True)
            queryset = queryset.filter(telecom_center__id__in=telecom_ids)

        if telecom:
            telecom_obj = TelecomCenter.objects.get(id=telecom)
            queryset = queryset.filter(telecom_center=telecom_obj)

        if telecom_name:
            telecom_ids = TelecomCenter.objects.filter(name__icontains=telecom_name).values_list('id', flat=True)
            queryset = queryset.filter(telecom_center__id__in=telecom_ids)

        if sort_field:
            if sort_field.replace('-', '') in ('telecom_center',):
                sort_field += '__name'
            elif sort_field.replace('-', '') in ('city',):
                sort_field = sort_field.replace('city', 'telecom_center__city__name')

            queryset = queryset.order_by(sort_field)

        return queryset

    """
    Create a model instance.
    """

    def create(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data, request=request)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        description = 'Add DSLAM {0}'.format(serializer.data['name'])
        add_audit_log(request, 'DSLAM', serializer.data['id'], 'Add DSLAM', description)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    """
    Update a model instance.
    """

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = self.serializer_class(instance, data=request.data, request=self.request, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        description = 'Update DSLAM {0}: {1}'.format(instance.name, request.data)
        add_audit_log(request, 'DSLAM', instance.id, 'Update DSLAM', description)
        return Response(serializer.data)

    """
    Destroy a model instance.
    """

    def destroy(self, request, *args, **kwargs):
        if self.request.user.is_superuser:
            instance = self.get_object()
            self.perform_destroy(instance)
            description = 'Delete DSLAM {0}'.format(instance.name)
            add_audit_log(request, 'DSLAM', instance.id, 'Delete DSLAM', description)
            return Response(status=status.HTTP_204_NO_CONTENT)
        return Response(status=status.HTTP_403_FORBIDDEN)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        user = request.user
        if user.type_contains([DIRECTRESELLER, RESELLER]):
            allowed_dslams = retrieve_allowed_dslams(user)
            if instance.id not in allowed_dslams:
                return Response("Permission Denied", status=status.HTTP_403_FORBIDDEN)
        return Response(serializer.data)



    @action(methods=['GET'], detail=False)
    def get_dslam_names(self, request):
        data = request.query_params
        exclude_dslam_ids = data.get('exclude_dslam_id')
        dslam_name = data.get('dslam_name')
        dslams = DSLAM.objects.all()
        if dslam_name:
            dslams = dslams.filter(name__istartswith=dslam_name)

        if exclude_dslam_ids:
            exclude_dslam_ids = eval(exclude_dslam_ids)
            dslams = dslams.exclude(id__in=exclude_dslam_ids)

        return Response({'result': dslams.annotate(search_name=Concat('ip', Value(' - '), 'name')).values('id',
                                                                                                          'search_name',
                                                                                                          'name')})

    @action(methods=['GET'], detail=False)
    def get_all_dslams(self, request):
        dslams = DSLAM.objects.all().values()
        return Response({'result': dslams})

    @action(methods=['GET'], detail=True)
    def getvlan(self, request, pk=None):
        user = request.user
        data = self.request.query_params
        dslam_id = self.get_object().id
        if not dslam_id:
            return Response({'result': 'dslam_id does not exists'}, status=status.HTTP_400_BAD_REQUEST)
        vlan_obj_ids = DSLAMPortVlan.objects.filter(port__dslam__id=dslam_id).values_list('vlan', flat=True).distinct()
        vlans_json = json.dumps([dict(item) for item in
                                 Vlan.objects.filter(id__in=vlan_obj_ids).values('id', 'vlan_id', 'vlan_name',
                                                                                 'reseller')])
        return JsonResponse(
            {'vlans': vlans_json})

    @action(methods=['GET'], detail=True)
    def vlan_usage_percentage(self, request, pk=None):
        user = request.user
        data = self.request.query_params
        dslam = self.get_object()
        if not dslam:
            return Response({'result': 'dslam_id does not exists'}, status=status.HTTP_400_BAD_REQUEST)
        result = utility.dslam_port_run_command(dslam.id, 'vlan show', {"is_queue": False, "type": "dslam"})
        if bool(result):
            all_vlan_count = {key: {'vlan_name': value, 'total': 0, 'percentage': 0.0} for key, value in
                              list(result['result'].items())}
            ports = DSLAMPort.objects.filter(dslam=dslam)
            vlans_count = DSLAMPortVlan.objects.filter(port__in=ports).values('vlan__vlan_id', 'vlan__vlan_name',
                                                                              'vlan').annotate(
                total=Count('vlan')).order_by('total')
            sum_vlan_count = sum([int(item['total']) for item in vlans_count])
            for vlan_count in vlans_count:
                all_vlan_count[str(vlan_count['vlan__vlan_id'])]['total'] = vlan_count['total']
                all_vlan_count[str(vlan_count['vlan__vlan_id'])]['id'] = vlan_count['vlan']
                all_vlan_count[str(vlan_count['vlan__vlan_id'])]['percentage'] = float(
                    vlan_count['total']) / sum_vlan_count * 100
            return JsonResponse({'values': all_vlan_count})
        return JsonResponse(
            {'values': []})

    @action(methods=['GET'], detail=True)
    def scan(self, request, pk=None):
        dslam_id = self.get_object().id
        scan_type = self.request.query_params.get('type', None)
        if scan_type == 'port':
            result = utility.scan_dslamport(dslam_id)
        elif scan_type == 'general':
            result = utility.scan_dslam_general_info(dslam_id)
        if 'error' not in result:
            return JsonResponse({'values': 'dslam ready to updating'})
        return JsonResponse({'values': 'error when dslam updating'})

    @action(methods=['GET'], detail=False)
    def count(self, request):
        dslam_count = DSLAM.objects.all().count()
        return Response({'dslam_count': dslam_count})

    @action(methods=['GET'], detail=True)
    def dslamreport(self, request, pk=None):
        dslam_obj = self.get_object()
        dslam_info = dslam_obj.get_info()
        queryset = DSLAMPort.objects.filter(dslam=dslam_obj)
        total_ports = queryset.count()
        up_ports = queryset.filter(admin_status='UNLOCK').count()
        down_ports = queryset.filter(admin_status='LOCK').count()
        sync_ports = queryset.filter(oper_status='SYNC').count()
        nosync_ports = queryset.filter(oper_status='NO-SYNC').count()

        line_profile_usage = DSLAMPort.objects.filter(dslam=dslam_obj).values(
            'line_profile'
        ).annotate(
            usage_count=Count('line_profile')
        )

        line_profile_usage = [{'name': item['line_profile'], 'y': item['usage_count']} for item in line_profile_usage]

        dslam_info['updated_at'] = str(JalaliDatetime(dslam_info['updated_at']).strftime("%Y-%m-%d %H:%M:%S"))
        dslam_info['created_at'] = str(JalaliDatetime(dslam_info['created_at']).strftime("%Y-%m-%d %H:%M:%S"))
        # dslam_info['last_sync'] = str(JalaliDatetime(dslam_info['last_sync']).strftime("%Y-%m-%d %H:%M:%S"))

        return JsonResponse({ \
            'dslam': dslam_info, 'total_ports': total_ports, \
            'up_ports': up_ports, 'down_ports': down_ports, 'sync_ports': sync_ports, \
            'dslam_type': dslam_obj.dslam_type.name, 'nosync_ports': nosync_ports, \
            'line_profile_usage': json.dumps(line_profile_usage), \
            'dslam_availability': dslam_obj.get_dslam_availability}
        )

    @action(methods=['GET'], detail=True)
    def dslam_curr_temperature_report(self, request, pk=None):
        user = request.user
        dslam = self.get_object()
        if not dslam:
            return Response({'result': 'Invalid DSLAM'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            dslamstatus = DSLAMStatus.objects.get(dslam=dslam)
        except Exception as e:
            print()
            e
            return Response({'result': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        lst_line_cards = []
        lst_line_card_values = []
        for obj_card_temp in dslamstatus.line_card_temp:
            lst_line_cards.append(obj_card_temp['name'])
            lst_line_card_values.append(int(obj_card_temp['CurValue']))
        return JsonResponse({'names': lst_line_cards, 'values': lst_line_card_values})

    @action(methods=['GET'], detail=True)
    def dslam_range_temperature_report(self, request, pk=None):
        user = request.user
        dslam = self.get_object()
        lc_name = self.request.query_params.get('lc_name', None)
        if not lc_name:
            # LC1-TEMP1
            card_number = self.request.query_params.get('card_number', None)
            temp_name = self.request.query_params.get('temp_name', None)
            lc_name = 'LC{0}-{1}'.format(card_number, temp_name.upper())
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)
        if not dslam:
            return Response({'result': 'Invalid DSLAM'}, status=status.HTTP_400_BAD_REQUEST)
        # try:
        if start_date:
            # for example start_date = 20120505 and end_date = 20120606
            start_date = datetime.strptime(start_date[0:4] + '-' + start_date[4:6] + '-' + start_date[6:8],
                                           '%Y-%m-%d').date()
            if end_date:
                end_date = datetime.strptime(end_date[0:4] + '-' + end_date[4:6] + '-' + end_date[6:8],
                                             '%Y-%m-%d').date()
            else:
                end_date = date.today()

            dslamstatus_snapshots = DSLAMStatusSnapshot.objects.filter(
                dslam_id=dslam.id
            ).filter(created_at__gte=start_date, created_at__lt=end_date).order_by('created_at')

        else:
            now = datetime.now()
            DD = timedelta(days=7)
            date_from = now - DD
            dslamstatus_snapshots = DSLAMStatusSnapshot.objects.filter(
                dslam_id=dslam.id
            ).filter(created_at__gte=date_from).order_by('created_at')

        lst_datetimes = []
        lst_line_card_values = []
        for obj in dslamstatus_snapshots:
            for card_temp in obj.line_card_temp:
                if isinstance(card_temp, dict):
                    if card_temp['name'] == lc_name:
                        lst_datetimes.append(JalaliDatetime(obj.created_at).strftime("%Y-%m-%d %H:%M:%S"))
                        lst_line_card_values.append(int(card_temp['CurValue']))
        return JsonResponse({'names': lst_datetimes, 'values': lst_line_card_values})

    @action(methods=['GET'], detail=True)
    def dslam_range_ping_report(self, request):
        user = request.user
        dslam_id = self.request.query_params.get('dslam_id', None)
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)
        identifier_key = self.request.query_params.get('identifier_key', None)
        username = self.request.query_params.get('username', None)

        if username:
            identifier_key = CustomerPort.objects.get(username=username).identifier_key

        if identifier_key:
            dslam_id = MDFDSLAM.objects.filter(identifier_key=identifier_key)[0].dslam_id

        if dslam_id:
            dslam = DSLAM.objects.get(id=dslam_id)
        else:
            return Response({'result': 'please send dslam_id '}, status=status.HTTP_400_BAD_REQUEST)

        if start_date:
            pass
        else:
            now = datetime.now()
            DD = timedelta(days=14)
            date_from = now - DD
            try:
                dslam_icmp_snapshots = DSLAMICMPSnapshot.objects.filter(
                    dslam_id=dslam.id
                ).filter(created_at__gte=date_from).order_by('created_at')


            except Exception as e:
                print()
                e
                return Response({'result': 'Dont Still Add This DSLAM For Getting ICMP'},
                                status=status.HTTP_400_BAD_REQUEST)

        lst_datetimes = []
        lst_avgping_values = []
        dslamicmp_values = {}
        try:
            for obj in dslam_icmp_snapshots:
                if obj.avgping == 'NaN' or obj.avgping is None:
                    continue
                persian_date = JalaliDatetime(obj.created_at).strftime("%Y-%m-%d %H:%M:%S")
                lst_datetimes.append(persian_date)
                lst_avgping_values.append(float(obj.avgping))
                icmp_values = {
                    'avgping': obj.avgping,
                    'jitter': obj.jitter,
                    'maxping': obj.maxping,
                    'minping': obj.minping,
                    'packet_loss': obj.packet_loss,
                    'received': obj.received,
                    'sent': obj.sent,
                }
                dslamicmp_values[persian_date] = icmp_values
        except:
            print()
            'give error dslam icmp : \n{0} - type is {1}'.format(obj.avgping, type(obj.avgping))

        return JsonResponse(
            {'names': lst_datetimes, 'values': lst_avgping_values, 'dslamicmp_values': dslamicmp_values}, safe=False)

    @action(methods=['GET'], detail=True)
    def dslam_icmp_info(self, request, pk=None):
        user = request.user
        dslam = self.get_object()
        date = self.request.query_params.get('date', None)
        if date:
            date_start = JalaliDatetime.strptime(date, '%Y-%m-%d %H:%M:%S').todatetime()
            DD = timedelta(seconds=1)
            date_end = date_start + DD
            dslam_snapshots = DSLAMICMPSnapshot.objects.filter(
                dslam_id=dslam.id
            ).get(created_at__gte=date_start, created_at__lt=date_end)
            return Response({'result': serialize('json', [dslam_snapshots, ])})

    @action(methods=['GET'], detail=True)
    def dslam_current_icmp_result(self, request, pk=None):
        user = request.user
        dslam = self.get_object()
        try:
            dslamicmp_obj = DSLAMICMP.objects.get(dslam=dslam)
        except:
            return Response({'result': 'Dont Still Add This DSLAM For Getting ICMP Protocol Result'},
                            status=status.HTTP_400_BAD_REQUEST)

        return Response({'result': serialize('json', [dslamicmp_obj])})

    @action(methods=['GET'], detail=True)
    def dslamslot(self, request, pk=None):
        user = request.user
        dslam = self.get_object()
        if dslam:
            query = '''select distinct(slot_number) from dslam_dslamport where dslam_id={0} order by slot_number'''.format(
                dslam.id)
            cursor = connection.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            results = {'slot_numbers': [row[0] for row in rows]}
            json_data = json.dumps(results)
            return HttpResponse(json_data, content_type="application/json")

    @action(methods=['GET'], detail=True)
    def board(self, request, pk=None):
        user = request.user
        dslam = self.get_object()
        if dslam:
            dslamboard_objs = DSLAMBoard.objects.filter(dslam=dslam)

        boards = [board.as_json() for board in dslamboard_objs]
        return HttpResponse(json.dumps(boards), content_type="application/json")


class DSLAMCartViewSet(mixins.ListModelMixin,
                       mixins.CreateModelMixin,
                       mixins.RetrieveModelMixin,
                       mixins.UpdateModelMixin,
                       mixins.DestroyModelMixin,
                       viewsets.GenericViewSet):
    queryset = DSLAMCart.objects.all()
    permission_classes = (IsAuthenticated,)
    serializer_class = DSLAMCartSerializer
    pagination_class = LargeResultsSetPagination

    def get_queryset(self):
        queryset = self.queryset
        telecom_id = self.request.query_params.get('telecom_center_id', None)
        if telecom_id:
            queryset = queryset.filter(telecom_center__id=telecom_id)
        return queryset

    """
    Create a model instance.
    """

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class DSLAMLocationViewSet(mixins.ListModelMixin,
                           mixins.CreateModelMixin,
                           mixins.RetrieveModelMixin,
                           mixins.UpdateModelMixin,
                           mixins.DestroyModelMixin,
                           viewsets.GenericViewSet):
    permission_classes = (IsAuthenticated, HasAccessToDslam)
    serializer_class = DSLAMLocationSerializer
    paginate_by = None
    paginate_by_param = None
    paginator = None

    def get_queryset(self):
        queryset = DSLAMLocation.objects.all()

        telecom_id = self.request.query_params.get('telecom-center', None)
        city_id = self.request.query_params.get('city_id', None)
        dslam_id = self.request.query_params.get('dslam_id', None)

        if dslam_id:
            dslam_obj = DSLAM.objects.get(id=dslam_id)
            queryset = queryset.filter(dslam=dslam_obj)

        if city_id is not None and city_id != '' and not telecom_id:
            city = City.objects.get(id=city_id)
            telecom_objs = TelecomCenter.objects.filter(city=city)
            queryset = queryset.filter(dslam__telecom_center__in=telecom_objs)

        if queryset.count() > 0:
            if telecom_id:
                telecom_obj = TelecomCenter.objects.get(id=telecom_id)
                queryset = queryset.filter(dslam__telecom_center__in=[telecom_obj, ])

        return queryset

    """
    Create a model instance.
    """

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        dslam_name = DSLAM.objects.get(id=request.data['dslam']).name
        description = 'Add DSLAM Location , DSLAM: {0} - lat: {1} - long: {2}'.format(
            dslam_name,
            request.data['dslam_lat'],
            request.data['dslam_long'])
        add_audit_log(request, 'DSLAMLocation', serializer.data['id'], 'Add DSLAM Location', description)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    """
    Update a model instance.
    """

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        description = 'Update DSLAM Location, DSLAM: {0} - params: {1}'.format(instance.dslam.name, request.data)
        add_audit_log(request, 'DSLAM', instance.id, 'Update DSLAM Location', description)
        return Response(serializer.data)

    """
    Destroy a model instance.
    """

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        description = 'Delete DSLAM Location, DSLAM: {0}'.format(instance.dslam.name)
        add_audit_log(request, 'DSLAM', instance.id, 'Delete DSLAM', description)
        return Response(status=status.HTTP_204_NO_CONTENT)


class DSLAMEventViewsSet(mixins.ListModelMixin,
                         mixins.RetrieveModelMixin,
                         mixins.UpdateModelMixin,
                         mixins.DestroyModelMixin,
                         viewsets.GenericViewSet):
    permission_classes = (IsAuthenticated, HasAccessToDslam)
    serializer_class = DSLAMEventSerializer
    queryset = DSLAMEvent.objects.all()

    def get_queryset(self):
        queryset = DSLAMEvent.objects.all()
        status = self.request.query_params.get('search_status', None)
        event_type = self.request.query_params.get('search_event_type', None)
        dslam_name = self.request.query_params.get('search_dslam_name', None)
        sort_field = self.request.query_params.get('sort_field', None)

        if status:
            queryset = queryset.filter(status=status)

        if dslam_name:
            dslam = DSLAM.objects.get(name=dslam_name)
            queryset = queryset.filter(dslam=dslam)

        if event_type:
            queryset = queryset.filter(type=event_type)

        if sort_field:
            queryset = queryset.order_by(sort_field)
        return queryset

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = DSLAMEventSerializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        new_instance = serializer.save()
        description = 'Update DSLAM Event, DSLAM: {0} - params: {1}'.format(new_instance.dslam.name, request.data)
        add_audit_log(request, 'DSLAM', new_instance.id, 'Update DSLAM Event', description)
        return Response(serializer.data)

    """
    Destroy a model instance.
    """

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        description = 'Delete DSLAM Event, DSLAM: {0}'.format(instance.dslam.name)
        add_audit_log(request, 'DSLAM', instance.id, 'Delete DSLAM Event', description)
        return Response(status=status.HTTP_204_NO_CONTENT)


class DSLAMPortEventViewsSet(mixins.ListModelMixin,
                             mixins.RetrieveModelMixin,
                             mixins.UpdateModelMixin,
                             mixins.DestroyModelMixin,
                             viewsets.GenericViewSet):
    permission_classes = (IsAuthenticated, HasAccessToDslam)
    serializer_class = DSLAMPortEventSerializer
    queryset = DSLAMPortEvent.objects.all()

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = DSLAMEventSerializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        new_instance = serializer.save()
        description = 'Update DSLAM Port Event,  DSLAM: {0} - Card: {1} - Slot: {2}, params: {3}'.format(
            new_instance.dslam.name,
            new_instance.slot_number,
            new_instance.port_number,
            request.data)
        add_audit_log(request, 'DSLAM', new_instance.id, 'Update DSLAM Event', description)
        return Response(serializer.data)

    """
    Destroy a model instance.
    """

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        description = 'Delete DSLAM Port Event,  DSLAM: {0} - Card: {1} - Slot: {2}, params: {3}'.format(
            instance.dslam.name,
            instance.slot_number,
            instance.port_number)
        add_audit_log(request, 'DSLAM', instance.id, 'Delete DSLAM Port Event', description)
        return Response(status=status.HTTP_204_NO_CONTENT)


class DSLAMRunICMPCommandView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        self.user = request.user
        data = request.data
        icmp_type = data.get('icmp_type')
        params = data.get('params')
        dslam_id = data.get('dslam_id')
        result = utility.run_icmp_command(dslam_id, icmp_type, params)

        dslam_name = DSLAM.objects.get(id=dslam_id).name

        description = 'Run {0} Command on DSLAM {1} with params: {2}'.format(icmp_type, dslam_name, params)
        add_audit_log(request, 'DSLAMICMP', None, 'Run ICMP Command on DSLAM', description)

        return Response({'result': result}, status=status.HTTP_201_CREATED)


class DSLAMPortRunCommandView(views.APIView):

    def post(self, request, format=None):
        user = request.user
        data = request.data
        command = data.get('command', None)
        dslam_id = data.get('dslam_id', None)
        params = data.get('params', None)
        try:
            dslam_obj = DSLAM.objects.get(id=dslam_id)

        except Exception as ex:
            return JsonResponse({'result': str(ex)}, status=status.HTTP_400_BAD_REQUEST)
        logging.info(command)
        if not command:
            return Response({'result': 'Command does not exits'}, status=status.HTTP_400_BAD_REQUEST)
        params["username"] = user.username
        logging.info(params)
        result = utility.dslam_port_run_command(dslam_obj.pk, command, params)
        print(result)
        logging.info(user.username)

        if result:
            logging.info(result)
            if 'Busy' in result:
                return Response({'result': result}, status=status.HTTP_400_BAD_REQUEST)
            else:
                description = 'Run Command {0} on DSLAM {1}'.format(
                    command,
                    dslam_obj.name)

                add_audit_log(request, 'DSLAMCommand', None, 'Run Command On DSLAM Port', description)
        logging.info(user.username)
        return Response({'result': result}, status=status.HTTP_202_ACCEPTED)


class GetPortInfoView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        user = request.user
        data = request.data
        try:
            dslam_obj = DSLAM.objects.get(id=int(data['dslam_id']))
        except:
            return Response({'result': 'Invalid dslam_id'}, status=status.HTTP_400_BAD_REQUEST)

        port_id = data.get('port_id', None)
        if not port_id:
            return Response({'result': 'Invalid port_id'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            port_obj = DSLAMPort.objects.get(dslam=dslam_obj, id=port_id)
        except:
            return Response({'result': 'Port does not exists'}, status=status.HTTP_404_NOT_FOUND)

        if user.type_contains(RESELLER):
            if not user.reseller.resellerport_set.filter(
                    port_name=port_obj.port_name,
                    dslam=dslam_obj,
            ).exists():
                return Response({}, status=status.HTTP_401_UNAUTHORIZED)

        utility.get_port_info(dslam_obj.pk, port_obj.pk)
        return Response({'messege': 'params posted'})


class ResetAdminStatusView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        user = request.user
        data = request.data
        try:
            dslam_obj = DSLAM.objects.get(id=int(data['dslam_id']))
        except:
            return Response({'result': 'Invalid dslam_id'}, status=status.HTTP_400_BAD_REQUEST)

        port_id = data.get('port_id', None)
        if not port_id:
            return Response({'result': 'Invalid port_id'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            port_obj = DSLAMPort.objects.get(dslam=dslam_obj, id=port_id)
        except:
            return Response({'result': 'Port does not exists'}, status=status.HTTP_404_NOT_FOUND)

        if user.type_contains(RESELLER):
            if not user.reseller.resellerport_set.filter(
                    port_name=port_obj.port_name,
                    dslam=dslam_obj,
            ).exists():
                return Response({}, status=status.HTTP_401_UNAUTHORIZED)
        result = utility.reset_admin_status(dslam_obj.pk, port_obj.pk)
        description = 'Reset Admin Status on DSLAM {0} Card {1} port {2}'.format(
            dslam_obj.name,
            port_obj.slot_number,
            port_obj.port_number)
        add_audit_log(request, 'DSLAMCommand', None, 'Reset Admin '
                                                     'Status Port', description)
        return Response({'messege': result})


class ChangePortAdminStatusView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        user = request.user
        data = request.data
        try:
            dslam_obj = DSLAM.objects.get(id=int(data['dslam_id']))
        except:
            return Response({'result': 'Invalid dslam_id'}, status=status.HTTP_400_BAD_REQUEST)

        port_id = data.get('port_id', None)
        if not port_id:
            return Response({'result': 'Invalid port_id'}, status=status.HTTP_400_BAD_REQUEST)

        new_status = data.get('new_status', None)
        if not new_status:
            return Response({'result': 'new_status does not exists'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            port_obj = DSLAMPort.objects.get(dslam=dslam_obj, id=port_id)
        except:
            return Response({'result': 'Port does not exists'}, status=status.HTTP_404_NOT_FOUND)

        if user.type_contains(RESELLER):
            if not user.reseller.resellerport_set.filter(
                    port_name=port_obj.port_name,
                    dslam=dslam_obj,
            ).exists():
                return Response({}, status=status.HTTP_401_UNAUTHORIZED)

        result = utility.change_port_admin_status(dslam_obj.pk, port_obj.pk, new_status)
        description = 'Change Admin Status on DSLAM {0} Card {1} port {2} to new status {3}'.format(
            dslam_obj.name,
            port_obj.slot_number,
            port_obj.port_number,
            new_status)
        add_audit_log(request, 'DSLAMCommand', None, 'Change Admin Status Port', description)
        return Response({'messege': result})


class ChangePortLineProfileView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        user = request.user
        data = request.data
        try:
            dslam_obj = DSLAM.objects.get(id=int(data['dslam_id']))
        except:
            return Response({'result': 'Invalid dslam_id'}, status=status.HTTP_400_BAD_REQUEST)

        port_id = data.get('port_id', None)
        if not port_id:
            return Response({'result': 'Invalid port_id'}, status=status.HTTP_400_BAD_REQUEST)

        new_line_profile = data.get('new_line_profile', None)
        if not new_line_profile:
            return Response({'result': 'new_line_profile does not exists'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            port_obj = DSLAMPort.objects.get(dslam=dslam_obj, id=port_id)
        except:
            return Response({'result': 'Port does not exists'}, status=status.HTTP_404_NOT_FOUND)

        if user.type_contains(RESELLER):
            if not user.reseller.resellerport_set.filter(
                    port_name=port_obj.port_name,
                    dslam=dslam_obj,
            ).exists():
                return Response({}, status=status.HTTP_401_UNAUTHORIZED)
        new_line_profile_name = LineProfile.objects.get(id=new_line_profile).name
        result = utility.change_port_line_profile(dslam_obj.pk, port_obj.pk, new_line_profile_name)
        description = 'Change Line Profile on DSLAM {0} Card {1} port {2} to new line profile {3}'.format(
            dslam_obj.name,
            port_obj.slot_number,
            port_obj.port_number,
            new_line_profile)
        add_audit_log(request, 'DSLAMCommand', None, 'Change Line Profile Port', description)
        return Response({'messege': result})


class DSLAMTypeViewSet(mixins.ListModelMixin,
                       mixins.CreateModelMixin,
                       mixins.RetrieveModelMixin,
                       mixins.UpdateModelMixin,
                       mixins.DestroyModelMixin,
                       viewsets.GenericViewSet):
    serializer_class = DSLAMTypeSerializer
    permission_classes = (IsAuthenticated,)
    queryset = DSLAMType.objects.all()
    paginate_by = None
    paginate_by_param = None
    paginator = None

    """
    Destroy a model instance.
    """

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        description = 'Delete DSLAM type {0} '.format(instance.name)
        add_audit_log(request, 'DSLAMType', instance.id, 'Delete DSLAM Type', description)
        return Response(status=status.HTTP_204_NO_CONTENT)

    """
    Create a model instance.
    """

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        city_name = City.objects.get(id=serializer.data['city']).name
        description = 'Add DSLAM Type {0}'.format(serializer.data['name'])
        add_audit_log(request, 'DSLAMType', serializer.data['id'], 'Add DSLAM type', description)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    """
    Update a model instance.
    """

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        description = 'Update DSLAM type {0}, params: {1}'.format(instance.name, request.data)
        add_audit_log(request, 'DSLAMType', instance.id, 'Update DSLAM Type', description)
        return Response(serializer.data)


class CommandViewSet(mixins.ListModelMixin,
                     mixins.CreateModelMixin,
                     mixins.RetrieveModelMixin,
                     mixins.UpdateModelMixin,
                     mixins.DestroyModelMixin,
                     viewsets.GenericViewSet):
    serializer_class = CommandSerializer
    permission_classes = (IsAuthenticated,)
    queryset = Command.objects.all()
    paginate_by = None
    paginate_by_param = None
    paginator = None

    def get_queryset(self):
        payload_token = PortmanUtils.decode_token(self.request)
        queryset = self.queryset
        username = payload_token['username'] #self.request.query_params.get('username', None)
        ldap_email = payload_token['email'] #self.request.query_params.get('ldap_email', None)
        command_type = self.request.query_params.get('type', None)
        dslam_id = self.request.query_params.get('dslam_id', None)
        command_type = self.request.query_params.get('type', None)
        exclude_type = self.request.query_params.get('exclude_type', None)
        command_name = self.request.query_params.get('command_name', None)
        exclude_command_ids = self.request.query_params.get('exclude_command_id', None)
        ldap_login = self.request.query_params.get('ldap_login', None)

        if dslam_id:
            dslam_type = DSLAM.objects.get(id=dslam_id).dslam_type_id
            user = self.request.user
            allowed_commands = user.get_allowed_commands()
            allowed_commands_dslam_type = DSLAMTypeCommand.objects.filter(dslam_type=dslam_type).values_list(
                'command_id', flat=True)
            if user.type_contains([SUPPORT, RESELLER, DIRECTRESELLER, BASIC]):
                queryset = queryset.filter(id__in=allowed_commands_dslam_type)
                queryset = queryset.filter(id__in=allowed_commands)
            elif user.type_contains([ADMIN]):
                queryset = queryset.filter(id__in=allowed_commands_dslam_type)
            else:
                queryset = queryset.filter(id=0)

        if exclude_command_ids:
            exclude_command_ids = eval(exclude_command_ids)
            if exclude_command_ids or len(exclude_command_ids) > 0:
                queryset = queryset.exclude(id__in=exclude_command_ids)
        show_command = self.request.query_params.get('show_command', None)
        if show_command:
            queryset = queryset.filter(show_command=show_command)
        if command_type:
            queryset = queryset.filter(type=command_type)
        if exclude_type:
            queryset = queryset.exclude(type=exclude_type)
        if command_name:
            queryset = queryset.filter(text__icontains=command_name)
        return queryset

    """
    Destroy a model instance.
    """

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        description = 'Delete Command {0} '.format(instance.name)
        add_audit_log(request, 'Command', instance.id, 'Delete Command', description)
        return Response(status=status.HTTP_204_NO_CONTENT)

    """
    Create a model instance.
    """

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        city_name = City.objects.get(id=serializer.data['city']).name
        description = 'Add Command {0}'.format(serializer.data['name'])
        add_audit_log(request, 'Command', serializer.data['id'], 'Add Command', description)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    """
    Update a model instance.
    """

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        description = 'Update Command {0}, params: {1}'.format(instance.name, request.data)
        add_audit_log(request, 'Command', instance.id, 'Update Command', description)
        return Response(serializer.data)


class CityViewSet(mixins.ListModelMixin,
                  mixins.CreateModelMixin,
                  mixins.RetrieveModelMixin,
                  mixins.UpdateModelMixin,
                  mixins.DestroyModelMixin,
                  viewsets.GenericViewSet):
    serializer_class = CitySerializer
    permission_classes = (IsAuthenticated,)

    @action(methods=['GET'], detail=False)
    def tree_view(self, request):
        q = '''select id,name as text, abbr, english_name ,parent_id from dslam_city order by id desc'''
        cursor = connection.cursor()
        cursor.execute(q)
        rows = cursor.fetchall()
        result = []
        parent = '#'

        keys = ('id', 'text', 'abbr', 'english_name', 'parent')
        for row in rows:
            if row[4] == None:
                city_row = (row[0], row[1], row[2], row[3], '#')
            else:
                city_row = row
            result.append(dict(list(zip(keys, city_row))))
        json_data = json.dumps(sorted(result, key=lambda k: k['id']))
        return HttpResponse(json_data, content_type="application/json")

    @action(methods=['GET'], detail=False)
    def table_view(self, request):
        q = '''select a.id,a.name,b.name from dslam_city a left join dslam_city b on a.parent_id = b.id order by a.id desc'''
        cursor = connection.cursor()
        cursor.execute(q)
        rows = cursor.fetchall()
        result = []
        parent = '#'

        keys = ('id', 'name', 'parent',)
        for row in rows:
            if row[2] == None:
                city_row = (row[0], row[1], '#')
            else:
                city_row = row
            result.append(dict(list(zip(keys, city_row))))
        json_data = json.dumps(sorted(result, key=lambda k: k['id']))
        return HttpResponse(json_data, content_type="application/json")

    def get_queryset(self):
        queryset = City.objects.all()
        parent = self.request.query_params.get('parent', None)
        city_name = self.request.query_params.get('city_name', None)
        has_location = self.request.query_params.get('has_location', None)
        page_size = self.request.query_params.get('page_size', 150)
        if parent:
            if parent == 'all':
                queryset = queryset.filter(parent=None)
            elif parent == 'never':
                queryset = queryset.exclude(parent=None)
            else:
                try:
                    queryset = queryset.filter(parent=parent)
                except:
                    pass
        if city_name:
            queryset = queryset.filter(Q(name__icontains=city_name) | Q(english_name__icontains=city_name))
        if has_location:
            if has_location == 'true':
                queryset = queryset.filter(lat__isnull=False)
        pagination.PageNumberPagination.page_size = page_size
        return queryset.order_by('name')

    """
    Destroy a model instance.
    """

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        description = 'Delete City {0} '.format(instance.name)
        add_audit_log(request, 'City', instance.id, 'Delete City', description)
        return Response(status=status.HTTP_204_NO_CONTENT)

    """
    Create a model instance.
    """

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        city_name = serializer.data['name']
        description = 'Add City {0}'.format(city_name)
        add_audit_log(request, 'City', serializer.data['id'], 'Add City', description)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    """
    Update a model instance.
    """

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        description = 'Update City {0}, params: {1}'.format(instance.name, request.data)
        add_audit_log(request, 'City', instance.id, 'Update City', description)
        return Response(serializer.data)


class CityLocationViewSet(mixins.ListModelMixin,
                          mixins.CreateModelMixin,
                          mixins.RetrieveModelMixin,
                          mixins.UpdateModelMixin,
                          mixins.DestroyModelMixin,
                          viewsets.GenericViewSet):
    serializer_class = CityLocationSerializer
    permission_classes = (IsAuthenticated, IsAdminUser)
    paginate_by = None
    paginate_by_param = None
    paginator = None

    def get_queryset(self):
        user = self.request.user
        city_id = self.request.query_params.get('city_id', None)
        queryset = CityLocation.objects.all()
        if city_id:
            city = City.objects.get(id=city_id)
            queryset = queryset.filter(city=city)
        return queryset

    """
    Destroy a model instance.
    """

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        description = 'Delete City {0} Location'.format(instance.city.name)
        add_audit_log(request, 'CityLocation', instance.id, 'Delete City Location', description)
        return Response(status=status.HTTP_204_NO_CONTENT)

    """
    Create a model instance.
    """

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        city_name = City.objects.get(id=serializer.data['city']).name
        description = 'Add City Location , City: {0} - lat: {1} - long: {2}'.format(
            city_name,
            serializer.data['city_lat'],
            serializer.data['city_long'])
        add_audit_log(request, 'CityLocation', serializer.data['id'], 'Add City Location', description)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    """
    Update a model instance.
    """

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        description = 'Update City Location , City: {0} - lat: {1} - long: {2}'.format(
            instance.city.name,
            request.data['city_lat'],
            request.data['city_long'])
        add_audit_log(request, 'CityLocation', instance.id, 'Update City Location', description)
        return Response(serializer.data)


class TelecomCenterViewSet(mixins.ListModelMixin,
                           mixins.CreateModelMixin,
                           mixins.RetrieveModelMixin,
                           mixins.UpdateModelMixin,
                           mixins.DestroyModelMixin,
                           viewsets.GenericViewSet):
    serializer_class = TelecomCenterSerializer
    queryset = TelecomCenter.objects.all()

    def destroy(self, request, *args, **kwargs):
        telecom_center = self.get_object()
        dslam_count = DSLAM.objects.filter(telecom_center=telecom_center).count()
        if dslam_count > 0:
            return Response({'msg': 'this telecom used in dslams'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)
        else:
            self.perform_destroy(telecom_center)
            description = 'Delete Telecom Center {0}'.format(telecom_center.name)
            add_audit_log(request, 'TelecomCenter', telecom_center.id, 'Delete Telecom Center', description)
            return Response(status=status.HTTP_204_NO_CONTENT)

    def get_queryset(self):
        queryset = self.queryset
        name = self.request.query_params.get('search_name', None)
        prefix_bukht_name = self.request.query_params.get('search_prefix_bukht_name', None)
        city_id = self.request.query_params.get('search_city_id', None)
        city_name = self.request.query_params.get('city_name', None)

        orderby = self.request.query_params.get('sort_field', None)

        if city_name:
            city_ids = City.objects.filter(name__icontains=city_name).values_list('id', flat=True)
            queryset = queryset.filter(city__id__in=city_ids)

        if city_id:
            city_obj = City.objects.get(id=city_id)
            if city_obj.parent == None:
                city_ids = City.objects.filter(parent=city_obj).values_list('id', flat=True)
                queryset = queryset.filter(city__id__in=city_ids)
            else:
                queryset = queryset.filter(city=city_obj)

        if name:
            queryset = queryset.filter(name__istartswith=name)

        if prefix_bukht_name:
            queryset = queryset.filter(prefix_bukht_name__icontains=prefix_bukht_name)

        if orderby:
            queryset = queryset.order_by(orderby)

        return queryset

    @action(methods=['GET'], detail=False)
    def get_without_paging(self, request):
        telecom_objs = TelecomCenter.objects.all()
        data = request.query_params
        city_id = data.get('city_id')
        if city_id:
            try:
                city_obj = City.objects.get(id=city_id)
                if city_obj.parent == None:
                    city_ids = City.objects.filter(parent=city_obj).values_list('id', flat=True)
                    telecom_objs = telecom_objs.filter(city__id__in=city_ids)
                else:
                    telecom_objs = telecom_objs.filter(city=city_obj)

            except:
                return Response({'result': 'city_id does not exists'}, status=status.HTTP_400_BAD_REQUEST)
        data = [item.as_json() for item in telecom_objs]
        return HttpResponse(json.dumps(data), content_type='application/json; charset=UTF-8')

    """
    Create a model instance.
    """

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        description = 'Create Telecom Center {0}'.format(serializer.data['name'])
        add_audit_log(request, 'TelecomCenter', serializer.data['id'], 'Create Telecome Center', description)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    """
    Update a model instance.
    """

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        description = 'Update Telecom Center {0}: {1}'.format(instance.name, request.data)
        add_audit_log(request, 'TelecomCenter', instance.id, 'Update Telecom Center', description)
        return Response(serializer.data)


class TelecomCenterLocationViewSet(mixins.ListModelMixin,
                                   mixins.CreateModelMixin,
                                   mixins.RetrieveModelMixin,
                                   mixins.UpdateModelMixin,
                                   mixins.DestroyModelMixin,
                                   viewsets.GenericViewSet):
    serializer_class = TelecomCenterLocationSerializer
    permission_classes = (IsAuthenticated, IsAdminUser)
    paginate_by = None
    paginate_by_param = None
    paginator = None

    def get_queryset(self):
        user = self.request.user
        queryset = TelecomCenterLocation.objects.all()
        telecom_id = self.request.query_params.get('telecom_id', None)
        if telecom_id:
            queryset = queryset.filter(telecom_center__id=telecom_id)
        return queryset

    """
    Create a model instance.
    """

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    """
    Update a model instance.
    """

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)


class DSLAMPortViewSet(mixins.ListModelMixin,
                       mixins.RetrieveModelMixin,
                       viewsets.GenericViewSet):
    serializer_class = DSLAMPortSerializer
    queryset = DSLAMPort.objects.all().order_by('port_index')
    # permission_classes = (IsAuthenticated, DSLAMPortView, DSLAMPortEdit)
    pagination_class = LargeResultsSetPagination

    def get_queryset(self):

        user = self.request.user
        queryset = self.queryset

        dslam_name = self.request.query_params.get('search_dslam_name', None)
        dslam_ip = self.request.query_params.get('search_dslam_ip', None)
        dslam_id = self.request.query_params.get('search_dslam_id', None)
        port_name = self.request.query_params.get('search_port_name', None)
        port_index = self.request.query_params.get('search_port_index', None)
        telecom_id = self.request.query_params.get('search_telecom', None)
        oper_status = self.request.query_params.get('search_oper_status', None)
        admin_status = self.request.query_params.get('search_admin_status', None)
        sort_field = self.request.query_params.get('sort_field', None)
        city_id = self.request.query_params.get('search_city', None)
        vlan_obj_id = self.request.query_params.get('search_vlan_id', None)
        line_profile = self.request.query_params.get('search_line_profile', None)
        down_attainable_rate_range = self.request.query_params.get('search_down_attainable_rate_range', None)
        up_attainable_rate_range = self.request.query_params.get('search_up_attainable_rate_range', None)
        down_attenuation_range = self.request.query_params.get('search_down_attenuation_range', None)
        up_attenuation_range = self.request.query_params.get('search_up_attenuation_range', None)
        down_snr_range = self.request.query_params.get('search_down_snr_range', None)
        up_snr_range = self.request.query_params.get('search_up_snr_range', None)
        down_tx_rate_range = self.request.query_params.get('search_down_tx_rate_range', None)
        up_tx_rate_range = self.request.query_params.get('search_up_tx_rate_range', None)
        down_snr_flag = self.request.query_params.get('search_down_snr_flag', None)
        up_snr_flag = self.request.query_params.get('search_up_snr_flag', None)
        down_atten_flag = self.request.query_params.get('search_down_atten_flag', None)
        up_atten_flag = self.request.query_params.get('search_up_atten_flag', None)
        slot_number = self.request.query_params.get('search_slot_number', None)
        port_number = self.request.query_params.get('search_port_number', None)
        slot_number_from = self.request.query_params.get('search_slot_from', None)
        slot_number_to = self.request.query_params.get('search_slot_to', None)
        port_number_from = self.request.query_params.get('search_port_from', None)
        port_number_to = self.request.query_params.get('search_port_to', None)
        mac_address = self.request.query_params.get('search_mac_address', None)
        username = self.request.query_params.get('search_username', None)
        identifier_keys = [self.request.query_params.get('search_identifier_key', None)]
        vpi = [self.request.query_params.get('vpi', None)]

        if slot_number:
            queryset = queryset.filter(slot_number=slot_number)

        if port_number:
            queryset = queryset.filter(port_number=port_number)

        if slot_number_from:
            if slot_number_to:
                queryset = queryset.filter(
                    slot_number__gte=slot_number_from,
                    slot_number__lt=slot_number_to)
            else:
                queryset = queryset.filter(slot_number__gte=slot_number_from)

        if port_number_from:
            if port_number_to:
                queryset = queryset.filter(
                    port_number__gte=port_number_from,
                    port_number__lt=port_number_to)
            else:
                queryset = queryset.filter(port_number__gte=port_number_from)

        if down_snr_flag:
            queryset = queryset.filter(downstream_snr_flag=down_snr_flag)

        if up_snr_flag:
            queryset = queryset.filter(upstream_snr_flag=up_snr_flag)

        if down_atten_flag:
            queryset = queryset.filter(downstream_attenuation_flag=down_atten_flag)

        if up_atten_flag:
            queryset = queryset.filter(upstream_attenuation_flag=up_atten_flag)

        if down_attainable_rate_range:
            down_attainable_rate_range = eval(down_attainable_rate_range)
            queryset = queryset.filter(downstream_attainable_rate__gte=down_attainable_rate_range['from'], \
                                       downstream_attainable_rate__lt=down_attainable_rate_range['to'])

        if up_attainable_rate_range:
            up_attainable_rate_range = eval(up_attainable_rate_range)
            queryset = queryset.filter(upstream_attainable_rate__gte=up_attainable_rate_range['from'], \
                                       upstream_attainable_rate__lt=up_attainable_rate_range['to'])

        if down_attenuation_range:
            down_attenuation_range = eval(down_attenuation_range)
            queryset = queryset.filter(downstream_attenuation__gte=down_attenuation_range['from'], \
                                       downstream_attenuation__lt=down_attenuation_range['to'])

        if up_attenuation_range:
            up_attenuation_range = eval(up_attenuation_range)
            queryset = queryset.filter(upstream_attenuation__gte=up_attenuation_range['from'], \
                                       upstream_attenuation__lt=up_attenuation_range['to'])

        if down_snr_range:
            down_snr_range = eval(down_snr_range)
            queryset = queryset.filter(downstream_snr__gte=down_snr_range['from'], \
                                       downstream_snr__lt=down_snr_range['to'])

        if up_snr_range:
            up_snr_range = eval(up_snr_range)
            queryset = queryset.filter(upstream_snr__gte=up_snr_range['from'], \
                                       upstream_snr__lt=up_snr_range['to'])

        if down_tx_rate_range:
            down_tx_rate_range = eval(down_tx_rate_range)
            queryset = queryset.filter(downstream_tx_rate__gte=down_tx_rate_range['from'], \
                                       downstream_tx_rate__lt=down_tx_rate_range['to'])

        if up_tx_rate_range:
            up_tx_rate_range = eval(up_tx_rate_range)
            queryset = queryset.filter(upstream_tx_rate__gte=up_tx_rate_range['from'], \
                                       upstream_tx_rate__lt=up_tx_rate_range['to'])

        if city_id:
            telecom_ids = TelecomCenter.objects.filter(city__id=city_id).values_list('id', flat=True)
            dslam_ids = DSLAM.objects.filter(telecom_center__id__in=telecom_ids).values_list('id', flat=True)
            queryset = queryset.filter(dslam__id__in=dslam_ids)

        if telecom_id:
            dslam_ids = DSLAM.objects.filter(telecom_center__id=telecom_id).values_list('id', flat=True)
            queryset = queryset.filter(dslam__id__in=dslam_ids)

        if dslam_id:
            queryset = queryset.filter(dslam__id=dslam_id)

        if dslam_name:
            queryset = queryset.filter(dslam__name__istartswith=dslam_name)

        if dslam_ip:
            if len(dslam_ip.split('.')) != 4:
                queryset = queryset.filter(dslam__ip__istartswith=dslam_ip)
            else:
                queryset = queryset.filter(dslam__ip=dslam_ip)

        if port_name:
            queryset = queryset.filter(port_name=port_name)

        if port_index:
            queryset = queryset.filter(port_index=port_index)

        if line_profile:
            queryset = queryset.filter(line_profile=line_profile)

        if admin_status:
            queryset = queryset.filter(admin_status=admin_status)

        if oper_status:
            queryset = queryset.filter(oper_status=oper_status)

        if vlan_obj_id:
            ports_id = DSLAMPortVlan.objects.filter(vlan__id=vlan_obj_id).values_list('port', flat=True)
            queryset = queryset.filter(id__in=ports_id)

        if sort_field:
            queryset = queryset.order_by(sort_field)

        if mac_address:
            port_ids = DSLAMPortMac.objects.filter(mac_address=mac_address).values_list('port__id', flat=True)
            queryset = queryset.filter(id__in=port_ids)

        if username:
            identifier_keys = CustomerPort.objects.filter(username=username).values_list('identifier_key', flat=True)

        if identifier_keys[0]:
            mdf_dslam_objs = MDFDSLAM.objects.filter(identifier_key__in=identifier_keys).values('dslam_id',
                                                                                                'slot_number',
                                                                                                'port_number')
            port_ids = []
            for mdf_dslam_obj in mdf_dslam_objs:
                port_ids.append(DSLAMPort.objects.get(
                    dslam_id=mdf_dslam_obj['dslam_id'],
                    slot_number=mdf_dslam_obj['slot_number'],
                    port_number=mdf_dslam_obj['port_number']).pk)
            queryset = queryset.filter(id__in=port_ids)

        return queryset

    @action(methods=['GET'], detail=True)
    def mac_address(self, request, pk=None):
        user = request.user
        port_obj = self.get_object()
        port_macs = DSLAMPortMac.objects.filter(port=port_obj).values_list('mac_address', flat=True)
        return Response({'mac': port_macs})

    @action(methods=['GET'], detail=True)
    def report(self, request, pk=None):
        user = request.user
        port_obj = self.get_object()
        dslam_id = port_obj.dslam.id
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)
        if start_date:
            # for example start_date = 20120505 and end_date = 20120606
            start_date = datetime.strptime(start_date[0:4] + '-' + start_date[4:6] + '-' + start_date[6:8],
                                           '%Y-%m-%d').date()
            if end_date:
                end_date = datetime.strptime(end_date[0:4] + '-' + end_date[4:6] + '-' + end_date[6:8],
                                             '%Y-%m-%d').date()
            else:
                end_date = date.today()

            port_snapshots = DSLAMPortSnapshot.objects.filter(
                dslam_id=dslam_id, port_index=port_obj.port_index
            ).filter(snp_date__gte=start_date, snp_date__lt=end_date).order_by('snp_date')

        else:
            now = datetime.now()
            DD = timedelta(weeks=1)
            date_from = now - DD
            port_snapshots = DSLAMPortSnapshot.objects.filter(
                dslam_id=dslam_id, port_index=port_obj.port_index
            ).filter(snp_date__gte=date_from).order_by('snp_date')

        up_snr_data, down_snr_data = [], []
        up_tx_rate, down_tx_rate = [], []
        up_attenuation, down_attenuation = [], []
        up_attainable_rate, down_attainable_rate = [], []

        oper_status = {'SYNC': 0, 'NO-SYNC': 0, 'OTHER': 0}  # up_count, down_count
        dates = []
        for snp in port_snapshots:
            dates.append(snp.snp_date.strftime('%Y-%m-%d %H:%M'))
            if snp.oper_status in ['SYNC', 'NO-SYNC']:
                oper_status[snp.oper_status] += 1
            else:
                oper_status['OTHER'] += 1

            up_snr_data.append(snp.upstream_snr)
            down_snr_data.append(snp.downstream_snr)
            up_tx_rate.append(snp.upstream_tx_rate)
            down_tx_rate.append(snp.downstream_tx_rate)
            up_attenuation.append(snp.upstream_attenuation)
            down_attenuation.append(snp.downstream_attenuation)
            up_attainable_rate.append(snp.upstream_attainable_rate)
            down_attainable_rate.append(snp.downstream_attainable_rate)

        snr_data = [{'name': 'UP Stream SNR', 'data': up_snr_data}, {'name': 'DOWN Stream SNR', 'data': down_snr_data}]
        attenuation_data = [{'name': 'UP Stream Attenuation', 'data': up_attenuation},
                            {'name': 'DOWN Stream Attenuation', 'data': down_attenuation}]
        tx_data = [{'name': 'Down Stream TX Rate', 'data': down_tx_rate},
                   {'name': 'UP Stream TX Rate', 'data': up_tx_rate}]
        attainable_rate_data = [{'name': 'UP Stream Attainable Rate', 'data': up_attainable_rate},
                                {'name': 'DOWN Stream Attainable Rate', 'data': down_attainable_rate}]
        oper_status_data = {'data': [{'name': name, 'y': value} for name, value in list(oper_status.items())]}

        return JsonResponse(
            {'dates': json.dumps(dates),
             'oper_status': json.dumps(oper_status_data),
             'snr_data': json.dumps(snr_data),
             'attenuation_data': json.dumps(attenuation_data),
             'tx_data': json.dumps(tx_data),
             'attainable_rate_data': json.dumps(attainable_rate_data),
             })


class PortStatusReportView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def get(self, request, format=None):
        user = request.user

        slot_number = request.GET.get('slot_number', None)
        port_number = request.GET.get('port_number', None)
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)

        try:
            dslam_obj = DSLAM.objects.get(id=int(request.GET['dslam_id']))
        except:
            return Response({'result': 'Invalid dslam_id'}, status=status.HTTP_400_BAD_REQUEST)

        if not port_number:
            return Response({'result': 'Invalid port number'}, status=status.HTTP_400_BAD_REQUEST)

        if not slot_number:
            return Response({'result': 'Invalid slot number'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            port_obj = DSLAMPort.objects.get(dslam=dslam_obj, port_number=port_number, slot_number=slot_number)
        except:
            return Response({'result': 'Port does not exists'}, status=status.HTTP_404_NOT_FOUND)

        # try:
        if start_date:
            # for example start_date = 20120505 and end_date = 20120606
            start_date = datetime.strptime(start_date[0:4] + '-' + start_date[4:6] + '-' + start_date[6:8],
                                           '%Y-%m-%d').date()
            if end_date:
                end_date = datetime.strptime(end_date[0:4] + '-' + end_date[4:6] + '-' + end_date[6:8],
                                             '%Y-%m-%d').date()
            else:
                end_date = date.today()

            port_snapshots = DSLAMPortSnapshot.objects.filter(
                dslam_id=dslam_obj.pk
            ).filter(
                port_index=port_obj.port_index
            ).filter(snp_date__gte=start_date, snp_date__lt=end_date).order_by('snp_date')

        else:
            now = datetime.now()
            DD = timedelta(weeks=1)
            date_from = now - DD
            port_snapshots = DSLAMPortSnapshot.objects.filter(
                dslam_id=dslam_obj.pk
            ).filter(
                port_index=port_obj.port_index
            ).filter(snp_date__gte=date_from).order_by('snp_date')

        up_snr_data, down_snr_data = [], []
        up_tx_rate, down_tx_rate = [], []
        up_attenuation, down_attenuation = [], []
        up_attainable_rate, down_attainable_rate = [], []

        oper_status = {'SYNC': 0, 'NO-SYNC': 0, 'OTHER': 0}  # up_count, down_count
        dates = []
        for snp in port_snapshots:
            dates.append(snp.snp_date.strftime('%Y-%m-%d %H:%M'))
            if snp.oper_status in ['SYNC', 'NO-SYNC']:
                oper_status[snp.oper_status] += 1
            else:
                oper_status['OTHER'] += 1

            up_snr_data.append(snp.upstream_snr)
            down_snr_data.append(snp.downstream_snr)
            up_tx_rate.append(snp.upstream_tx_rate)
            down_tx_rate.append(snp.downstream_tx_rate)
            up_attenuation.append(snp.upstream_attenuation)
            down_attenuation.append(snp.downstream_attenuation)
            up_attainable_rate.append(snp.upstream_attainable_rate)
            down_attainable_rate.append(snp.downstream_attainable_rate)

        snr_data = [{'name': 'UP Stream SNR', 'data': up_snr_data}, {'name': 'DOWN Stream SNR', 'data': down_snr_data}]
        attenuation_data = [{'name': 'UP Stream Attenuation', 'data': up_attenuation},
                            {'name': 'DOWN Stream Attenuation', 'data': down_attenuation}]
        tx_data = [{'name': 'Down Stream TX Rate', 'data': down_tx_rate},
                   {'name': 'UP Stream TX Rate', 'data': up_tx_rate}]
        attainable_rate_data = [{'name': 'UP Stream Attainable Rate', 'data': up_attainable_rate},
                                {'name': 'DOWN Stream Attainable Rate', 'data': down_attainable_rate}]
        oper_status_data = [{'name': name, 'y': value} for name, value in list(oper_status.items())]
        reseller_obj = None
        customer_obj = None
        identifier_key = None
        try:
            identifier_key = MDFDSLAM.objects.get(dslam_id=port_obj.dslam.id, slot_number=port_obj.slot_number,
                                                  port_number=port_obj.port_number).identifier_key

        except Exception as ex:
            print()
            ex

        if identifier_key:
            try:
                customer = CustomerPort.objects.get(identifier_key=identifier_key)
                customer_obj = serialize('json', [customer])
            except Exception as ex:
                print()
                ex

            try:
                reseller = ResellerPort.objects.get(identifier_key=identifier_key).reseller
                reseller_obj = serialize('json', [reseller])
            except Exception as ex:
                print()
                ex

        dslamport_vlan_obj = DSLAMPortVlan.objects.filter(port=port_obj)
        vlan_obj_serializer = None
        if dslamport_vlan_obj.count() > 0:
            vlan_obj_serializer = serialize('json', [dslamport_vlan_obj.order_by('-created_at')[0].vlan])

        dslam_seializer_obj = serialize('json', [dslam_obj])
        dslam_json_obj = json.loads(dslam_seializer_obj)
        dslam_json_obj[0]['fields']['updated_at'] = JalaliDatetime(datetime.strptime( \
            ' '.join(dslam_json_obj[0]['fields']['updated_at'].split('T')).split('.')[0] \
            , "%Y-%m-%d %H:%M:%S")).strftime("%Y-%m-%d %H:%M:%S")
        dslam_json_obj[0]['fields']['created_at'] = JalaliDatetime(datetime.strptime( \
            ' '.join(dslam_json_obj[0]['fields']['created_at'].split('T')).split('.')[0] \
            , "%Y-%m-%d %H:%M:%S")).strftime("%Y-%m-%d %H:%M:%S")
        dslam_json_obj[0]['fields']['last_sync'] = JalaliDatetime(datetime.strptime( \
            ' '.join(dslam_json_obj[0]['fields']['last_sync'].split('T')).split('.')[0] \
            , "%Y-%m-%d %H:%M:%S")).strftime("%Y-%m-%d %H:%M:%S")

        dslamport_seializer_obj = serialize('json', [port_obj])

        dslamport_json_obj = json.loads(dslamport_seializer_obj)
        dslamport_json_obj[0]['fields']['updated_at'] = JalaliDatetime(datetime.strptime( \
            ' '.join(dslamport_json_obj[0]['fields']['updated_at'].split('T')).split('.')[0] \
            , "%Y-%m-%d %H:%M:%S")).strftime("%Y-%m-%d %H:%M:%S")
        dslamport_json_obj[0]['fields']['created_at'] = JalaliDatetime(datetime.strptime( \
            ' '.join(dslamport_json_obj[0]['fields']['created_at'].split('T')).split('.')[0] \
            , "%Y-%m-%d %H:%M:%S")).strftime("%Y-%m-%d %H:%M:%S")

        return JsonResponse(
            {'dslam': dslam_json_obj,
             'port': dslamport_json_obj,
             'dates': json.dumps(dates),
             'oper_status': json.dumps(oper_status_data),
             'snr_data': json.dumps(snr_data),
             'attenuation_data': json.dumps(attenuation_data),
             'tx_data': json.dumps(tx_data),
             'customer': customer_obj,
             'reseller': reseller_obj,
             'attainable_rate_data': json.dumps(attainable_rate_data),
             'port_vlan': vlan_obj_serializer
             })
        # except:
        #    return JsonResponse({'result':'error'})


class ResellerViewSet(mixins.ListModelMixin,
                      mixins.CreateModelMixin,
                      mixins.RetrieveModelMixin,
                      mixins.UpdateModelMixin,
                      mixins.DestroyModelMixin,
                      viewsets.GenericViewSet):
    serializer_class = ResellerSerializer
    permission_classes = (IsAuthenticated, IsAdminUser)

    def get_queryset(self):
        reseller_queryset = Reseller.objects.all()
        name = self.request.query_params.get('name', None)
        tel = self.request.query_params.get('tel', None)
        fax = self.request.query_params.get('fax', None)
        address = self.request.query_params.get('address', None)
        city_id = self.request.query_params.get('city_id', None)
        vpi = self.request.query_params.get('vpi', None)
        vci = self.request.query_params.get('vci', None)
        orderby = self.request.query_params.get('sort_field', None)

        if name:
            try:
                reseller_queryset = reseller_queryset.filter(name__icontains=name)
            except:
                pass
        if name:
            try:
                reseller_queryset = reseller_queryset.filter(name__icontains=name)
            except:
                pass
        if address:
            try:
                reseller_queryset = reseller_queryset.filter(address__icontains=address)
            except:
                pass
        if tel:
            try:
                reseller_queryset = reseller_queryset.filter(tel__contains=tel)
            except:
                pass
        if fax:
            try:
                reseller_queryset = reseller_queryset.filter(fax__contains=fax)
            except:
                pass
        if city_id:
            try:
                city = City.objects.get(id=city_id)
                reseller_queryset = reseller_queryset.filter(city=city)
            except:
                pass

        if orderby:
            reseller_queryset = reseller_queryset.order_by(orderby)

            if vpi:
                try:
                    reseller_queryset = reseller_queryset.filter(vpi=vpi)
                except:
                    pass

        return reseller_queryset

    @action(methods=['GET'], detail=True)
    def report(self, request, pk=None):
        user = request.user
        reseller_obj = self.get_object()
        telecom_center_ports = defaultdict(list)
        for reseller_port in ResellerPort.objects.filter(reseller=reseller_obj):
            telecom_center_ports[reseller_port.telecom_center_id].append(reseller_port.identifier_key)

        data = []

        total_ports_count = 0
        up_ports_count = 0
        down_ports_count = 0
        for telecom_center_id, identifier_keys in list(telecom_center_ports.items()):
            tc_object = TelecomCenter.objects.get(id=telecom_center_id)
            customer_port = set(
                CustomerPort.objects.filter(telecom_center_id=telecom_center_id).values_list('identifier_key',
                                                                                             flat=True))

            intersection_port = set(identifier_keys).intersection(customer_port)

            total_ports_count += len(identifier_keys)
            up_ports_count += len(intersection_port)
            down_ports_count += len(identifier_keys) - len(intersection_port)

            data.append({
                'telecom_center': {'id': tc_object.id, 'name': tc_object.name},
                'total_ports_count': len(identifier_keys),
                'up_ports_count': len(intersection_port),
                'down_ports_count': len(identifier_keys) - len(intersection_port),
            })

        return JsonResponse({
            'total_ports_count': total_ports_count,
            'up_ports_count': up_ports_count,
            'down_ports_count': down_ports_count,
            'data': data,
        })


class ResellerPortViewSet(mixins.ListModelMixin,
                          mixins.CreateModelMixin,
                          mixins.RetrieveModelMixin,
                          mixins.UpdateModelMixin,
                          mixins.DestroyModelMixin,
                          viewsets.GenericViewSet):
    serializer_class = ResellerPortSerializer
    permission_classes = (IsAuthenticated, HasAccessToDslamPort)

    def get_queryset(self):
        queryset = ResellerPort.objects.all()
        identifier_key = self.request.query_params.get('identifire_key', None)
        dslam_name = self.request.query_params.get('search_dslam_name', None)
        reseller_name = self.request.query_params.get('search_reseller_name', None)
        dslam_port_id = self.request.query_params.get('dslamport_id', None)
        orderby = self.request.query_params.get('sort_field', None)

        if dslam_port_id:
            port_obj = DSLAMPort.objects.get(id=dslam_port_id)
            try:
                identifier_key = MDFDSLAM.objects.get(dslam_id=port_obj.dslam_id, slot_number=port_obj.slot_number,
                                                      port_number=port_obj.port_number).identifier_key
            except Exception as ex:
                print()
                ex
                return []

        if identifier_key:
            print()
            identifier_key
            queryset = queryset.filter(identifier_key=identifier_key)

        if dslam_name:
            dslam_ids = DSLAM.objects.filter(name__istartswith=dslam_name).values_list('id', flat=True)
            identifier_keys = MDFDSLAM.objects.filter(dslam_id__in=dslam_ids).values_list('identifier_key', flat=True)
            queryset = queryset.filter(identifier_key__in=identifier_keys)
        if reseller_name:
            reseller_ids = Reseller.objects.filter(name__istartswith=reseller_name).values_list('id', flat=True)
            queryset = queryset.filter(reseller__id__in=reseller_ids)

        if orderby:
            queryset = queryset.order_by(orderby)
        return queryset

    def create(self, request, *args, **kwargs):
        print()
        request.data
        self.user = request.user
        dslam_id = request.data.get('dslam')
        slot_number = request.data.get('slot_number')
        slot_from = request.data.get('slot_number_from')
        slot_to = request.data.get('slot_number_to')

        port_number = request.data.get('port_number')
        port_to = request.data.get('port_number_to')
        port_from = request.data.get('port_number_from')

        reseller_id = request.data.get('reseller')
        identifier_key = request.data.get('identifier_key')
        vlan_id = request.data.get('vlan_id')

        dslam_obj = DSLAM.objects.get(id=dslam_id)
        print()
        dslam_obj
        reseller_obj = Reseller.objects.get(id=reseller_id)

        if not reseller_id:
            return Response({'result': 'reseller_id does not exists'}, status=status.HTTP_400_BAD_REQUEST)
        reseller = Reseller.objects.get(id=reseller_id)

        if identifier_key:
            mdf_dslam_obj = MDFDSLAM.objects.get(identifier_key=identifier_key)
            mdf_dslams = ((mdf_dslam_obj.telecom_center_id, mdf_dslam_obj.slot_number, mdf_dslam_obj.port_number,
                           mdf_dslam_obj.identifier_key),)

        elif port_number and slot_number and dslam_id and reseller_id:
            print()
            'here'
            t = MDFDSLAM.objects.filter(dslam_id=dslam_id)
            print()
            t
            print()
            '====='
            mdf_dslams = MDFDSLAM.objects.filter(
                dslam_id=dslam_id,
                slot_number=slot_number,
                port_number=port_number).values_list(
                'telecom_center_id',
                'slot_number',
                'port_number',
                'identifier_key')

        elif slot_from and slot_to and port_from and port_to and port_from:
            mdf_dslams = MDFDSLAM.objects.filter(
                dslam_id=dslam_id,
                slot_number__gte=slot_from,
                port_number__gte=port_from).filter(
                slot_number__lte=slot_to,
                port_number__lte=port_to).values_list('telecom_center_id', 'slot_number', 'port_number',
                                                      'identifier_key')
        else:
            return Response({'result': 'bad parameter send'}, status=status.HTTP_400_BAD_REQUEST)

        print()
        mdf_dslams
        added_item = []
        exist_item = []
        port_indexes = []
        for telecom_center_id, slot_number, port_number, identifier_key in mdf_dslams:
            try:
                rp = ResellerPort()
                rp.reseller = reseller
                rp.telecom_center_id = telecom_center_id
                rp.identifier_key = identifier_key
                rp.save()
                port_indexes.append({
                    'slot_number': slot_number,
                    'port_number': port_number,
                    'port_index': DSLAMPort.objects.get(dslam__id=dslam_id, port_number=port_number,
                                                        slot_number=slot_number).port_index
                })
                added_item.append(
                    {'reseller': reseller.name, 'identifier_key': identifier_key, 'slot_number': slot_number,
                     'port_number': port_number})
            except Exception as ex:
                exist_item.append(
                    {'reseller': reseller.name, 'identifier_key': identifier_key, 'slot_number': slot_number,
                     'port_number': port_number})

        if len(added_item) > 0:
            params = dict(vlan_id=vlan_id, port_indexes=port_indexes, username=self.user.username)

            result = utility.dslam_port_run_command(dslam_id, 'add to vlan', params)

        return Response({'exist_item': exist_item, 'added_item': added_item}, status=status.HTTP_201_CREATED)


class CustomerPortViewSet(mixins.ListModelMixin,
                          mixins.CreateModelMixin,
                          mixins.RetrieveModelMixin,
                          mixins.UpdateModelMixin,
                          mixins.DestroyModelMixin,
                          viewsets.GenericViewSet):
    serializer_class = CustomerPortSerializer
    permission_classes = (IsAuthenticated, HasAccessToDslamPort)
    queryset = CustomerPort.objects.all()

    def get_queryset(self):
        user = self.request.user
        customerports = self.queryset

        identifire_key = self.request.query_params.get('identifier_key', None)
        dslam_port_id = self.request.query_params.get('dslamport_id', None)
        firstname = self.request.query_params.get('firstname', None)
        lastname = self.request.query_params.get('lastname', None)
        username = self.request.query_params.get('username', None)
        tel = self.request.query_params.get('tel', None)
        mobile = self.request.query_params.get('mobile', None)
        email = self.request.query_params.get('email', None)
        national_code = self.request.query_params.get('national_code', None)
        sort_field = self.request.query_params.get('sort_field', None)

        if dslam_port_id:
            port_obj = DSLAMPort.objects.get(id=dslam_port_id)
            try:
                mdfdslam_obj = MDFDSLAM.objects.get(dslam_id=port_obj.dslam_id, slot_number=port_obj.slot_number,
                                                    port_number=port_obj.port_number)
                customerports = customerports.filter(identifier_key=mdfdslam_obj.identifier_key)
            except Exception as ex:
                return []

        if identifire_key:
            try:
                customerports = customerports.filter(identifier_key__istartswith=identifire_key)
            except Exception as ex:
                print()
                '=>', ex
                pass

        if firstname:
            try:
                customerports = customerports.filter(firstname__istartswith=firstname)
            except:
                pass
        if lastname:
            try:
                customerports = customerports.filter(lastname__istartswith=lastname)
            except:
                pass
        if username:
            try:
                customerports = customerports.filter(username__istartswith=username)
            except:
                pass
        if tel:
            try:
                customerports = customerports.filter(tel__istartswith=tel)
            except:
                pass
        if mobile:
            try:
                customerports = customerports.filter(mobile__istartswith=mobile)
            except:
                pass
        if email:
            try:
                customerports = customerports.filter(email=email)
            except Exception as ex:
                pass

        if national_code:
            try:
                customerports = customerports.filter(national_code__contains=national_code)
            except:
                pass
        if sort_field:
            customerports = customerports.order_by(sort_field)
        return customerports

    def create(self, request, *args, **kwargs):
        data = request.data

        identifier_key = data.get('identifier_key')
        username = data.get('username')
        params = data.get('params')
        vlan_obj_id = data.get('v_id')

        if identifier_key:
            try:
                mdf_dslam_obj = MDFDSLAM.objects.filter(identifier_key=identifier_key)[0]
            except Exception as ex:
                return Response({'result': \
                                     'identifier_key does not exist'}, \
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'result': \
                                 'username or identifier_key does not exists'}, \
                            status=status.HTTP_400_BAD_REQUEST)

        if params.get('vlan_id') and int(params.get('vlan_id')) == 1:
            vlan_obj = Vlan.objects.get(vlan_id='1', reseller=None)
            params['vlan_id'] = vlan_obj.vlan_id
        elif vlan_obj_id:
            vlan_obj = Vlan.objects.get(id=vlan_obj_id)
            params['vlan_id'] = vlan_obj.vlan_id

        if 'port_conditions' not in list(params.keys()):
            params['port_conditions'] = {'slot_number': mdf_dslam_obj.slot_number,
                                         'port_number': mdf_dslam_obj.port_number}
            params['dslam_id'] = mdf_dslam_obj.dslam_id

        if params.get('vlan_id'):
            params["username"] = request.user.username
            result = utility.dslam_port_run_command(mdf_dslam_obj.dslam_id, 'port pvc set', params)
            if 'error' in result:
                return Response({'result': 'service is not available'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        # return Response({'result': identifier_key}, status=status.HTTP_201_CREATED)

        port_obj = DSLAMPort.objects.get(
            dslam__id=mdf_dslam_obj.dslam_id,
            slot_number=mdf_dslam_obj.slot_number,
            port_number=mdf_dslam_obj.port_number)

        if params.get('vlan_id'):
            dslamport_vlan_obj, created_port_vlan_obj = DSLAMPortVlan.objects.get_or_create(
                port=port_obj,
                vlan=vlan_obj
            )

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        # End Port PVC Set

        headers = self.get_success_headers(serializer.data)
        description = 'Create Customer {0}'.format(serializer.data['username'])
        add_audit_log(request, 'Customer', serializer.data['id'], 'Create Customer', description)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(methods=['GET'], detail=True)
    def activeport(self, request):
        user = request.user
        data = request.data
        identifier_key = data.get('identifier_key')
        username = data.get('username')
        params = data.get('params')
        vlan_obj_id = data.get('vlan_id')

        if username:
            identifier_key = CustomerPort.objects.get(username=username).identifier_key
        if identifier_key:
            try:
                mdf_dslam_obj = MDFDSLAM.objects.filter(identifier_key=identifier_key)[0]
            except Exception as ex:
                return Response({'result': \
                                     'identifier_key does not exist'}, \
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'result': \
                                 'username or identifier_key does not exists'}, \
                            status=status.HTTP_400_BAD_REQUEST)

        if int(params.get('vlan_id')) == 1:
            vlan_obj = Vlan.objects.get(vlan_id='1', reseller=None)
            params['vlan_id'] = vlan_obj.vlan_id
        elif vlan_obj_id:
            vlan_obj = Vlan.objects.get(id=vlan_obj_id)
            params['vlan_id'] = vlan_obj.vlan_id
        else:
            return Response({'result': \
                                 'vlan_obj_id does not exist'}, \
                            status=status.HTTP_400_BAD_REQUEST)

        if 'card_ports' not in list(params.keys()):
            params['port_indexes'] = [
                {'slot_number': mdf_dslam_obj.slot_number, 'port_number': mdf_dslam_obj.port_number}]
        if params.get('vlan_id'):
            params["username"] = user.username
            result = utility.dslam_port_run_command(mdf_dslam_obj.dslam_id, params.get('command'), params)
            if 'error' in result:
                return Response({'result': 'service is not available'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        port_obj = DSLAMPort.objects.get(
            dslam__id=mdf_dslam_obj.dslam_id,
            slot_number=mdf_dslam_obj.slot_number,
            port_number=mdf_dslam_obj.port_number)

        if params.get('vlan_id'):
            dslamport_vlan_obj, created_port_vlan_obj = DSLAMPortVlan.objects.get_or_create(
                port=port_obj,
                vlan=vlan_obj
            )

        return Response({'result': 'Port Acitived'}, status=status.HTTP_201_CREATED)

    """
    Update a model instance.
    """

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        description = 'Update Customer Port {0}: {1}'.format(instance.name, request.data)
        add_audit_log(request, 'CustomerPort', instance.id, 'Update Customer Port', description)
        return Response(serializer.data)


class PortCommandViewSet(mixins.ListModelMixin,
                         mixins.RetrieveModelMixin,
                         viewsets.GenericViewSet):
    serializer_class = PortCommandSerializer
    permission_classes = (IsAuthenticated,)
    paginate_by = None
    paginate_by_param = None
    paginator = None

    def get_queryset(self):
        user = self.request.user
        print()
        '--------------------------'
        print()
        self.request.query_params
        print()
        '--------------------------'
        portcommands = PortCommand.objects.all().order_by('-created_at')

        dslam_id = self.request.query_params.get('dslam', None)
        dslamport_id = self.request.query_params.get('dslamport_id', None)
        slot_number = self.request.query_params.get('slot_number', None)
        port_number = self.request.query_params.get('port_number', None)
        port_index = self.request.query_params.get('port_index', None)
        limit_row = self.request.query_params.get('limit_row', None)
        command_type_id = self.request.query_params.get('command_type_id', None)
        command_type_name = self.request.query_params.get('command_type_name', None)

        port_obj = None
        if dslamport_id:
            try:
                port_obj = DSLAMPort.objects.get(id=dslamport_id)
            except Exception as ex:
                print()
                ex
                return []

        elif slot_number and port_number and dslam_id:
            port_obj = DSLAMPort.objects.get(dslam__id=dslam_id, port_number=port_number, slot_number=slot_number)
        else:
            print()
            '=>>>>>>>>> please send dslamport_id'
            return []

        if command_type_name:
            portcommands = portcommands.filter(command__text=command_type_name)
        elif command_type_id:
            portcommands = portcommands.filter(command__id=command_type_id)
        else:
            query = """SELECT id FROM
            (SELECT *,row_number() OVER(PARTITION BY card_ports->'card', card_ports->'port',command_id ORDER BY created_at DESC) as rnk FROM dslam_portcommand )
            as d WHERE rnk = 1"""
            portcommands = portcommands.extra(where=["id in (%s)" % query])

        if limit_row:
            portcommands = portcommands.filter(dslam__id=dslam_id, card_ports__contains=[
                {"slot_number": port_obj.slot_number, "port_number": port_obj.port_number}])[:int(limit_row)]
        else:
            portcommands = portcommands.filter(dslam__id=dslam_id, card_ports__contains=[
                {"slot_number": port_obj.slot_number, "port_number": port_obj.port_number}])
        return portcommands


class DSLAMCommandViewSet(mixins.ListModelMixin,
                          mixins.RetrieveModelMixin,
                          viewsets.GenericViewSet):
    serializer_class = DSLAMCommandSerializer
    permission_classes = (IsAuthenticated,)
    queryset = DSLAMCommand.objects.all().order_by('-created_at')
    paginate_by = None
    paginate_by_param = None
    paginator = None

    def get_queryset(self):
        user = self.request.user
        dslamcommands = self.queryset

        dslam_id = self.request.query_params.get('dslam', None)
        limit_row = self.request.query_params.get('limit_row', None)
        command_type_id = self.request.query_params.get('command_type_id', None)
        command_type_name = self.request.query_params.get('command_type', None)
        if command_type_id:
            dslamcommands = dslamcommands.filter(command__id=command_type_id)
        if command_type_name:
            dslamcommands = dslamcommands.filter(command__text=command_type_name)
        else:
            query = """SELECT id FROM
            (SELECT *,row_number() OVER(PARTITION BY dslam_id,command_id ORDER BY created_at DESC) as rnk FROM dslam_dslamcommand )
            as d WHERE rnk = 1"""
            dslamcommands = dslamcommands.extra(where=["id in (%s)" % query])

        try:
            dslam = DSLAM.objects.get(id=dslam_id)
            if limit_row:
                dslamcommands = dslamcommands.filter(dslam=dslam)[:int(limit_row)]
            else:
                dslamcommands = dslamcommands.filter(dslam=dslam)
            return dslamcommands
        except:
            return []


class BulkCommand(views.APIView):
    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        user = request.user
        data = request.data
        commands = data.get('commands')
        command_objs = []
        for command in commands:
            command_dict = dict(
                Command.objects.filter(id=command.get('command_id')).values('id', 'text', 'type').first())
            command_dict['params'] = command.get('params')
            command_objs.append(command_dict)
        conditions = data.get('conditions')
        title = data.get('title')
        print()
        '======================================='
        print()
        title
        print()
        command_objs,
        print()
        conditions
        print()
        '======================================='
        result = utility.bulk_command(title, command_objs, conditions)
        return JsonResponse({'result': 'Commands is running with conditions'})


class MDFDSLAMViewSet(mixins.ListModelMixin,
                      mixins.CreateModelMixin,
                      mixins.RetrieveModelMixin,
                      mixins.UpdateModelMixin,
                      viewsets.GenericViewSet):
    serializer_class = MDFDSLAMSerializer
    permission_classes = (IsAuthenticated, HasAccessToDslamPort)
    #pagination_class = LargeResultsSetPagination

    def get_queryset(self):
        user = self.request.user
        queryset = MDFDSLAM.objects.all().order_by('identifier_key')
        telecom_center_id = self.request.query_params.get('search_telecom', None)
        telecom_mdf_id = self.request.query_params.get('search_telecom_mdf', None)
        telecom_identifier_key = self.request.query_params.get('search_identifier_key', None)
        sort_field = self.request.query_params.get('sort_field', None)

        if telecom_center_id:
            queryset = queryset.filter(telecom_center_id=telecom_center_id)

        if telecom_mdf_id:
            queryset = queryset.filter(telecom_center_mdf_id=telecom_mdf_id)

        if telecom_identifier_key:
            queryset = queryset.filter(identifier_key__istartswith=telecom_identifier_key)

        if sort_field:
            queryset = queryset.order_by(sort_field)

        return queryset

    @action(methods=['GET'], detail=True)
    def get_free_identifier(self, request):
        data = request.query_params
        telecom_id = data.get('telecom_id')
        dslam_id = data.get('dslam_id')
        identifier_key = data.get('identifier_key')
        if not telecom_id and not dslam_id:
            return Response({'result': 'pleas send telecom_id or dslam parameter'}, status=status.HTTP_400_BAD_REQUEST)

        if dslam_id:
            dslam_ids = [dslam_id, ]
        elif telecom_id:
            telecom_obj = TelecomCenter.objects.get(id=telecom_id)
            dslam_ids = DSLAM.objects.filter(telecom_center=telecom_obj).values_list('id')

        all_identifier = set(MDFDSLAM.objects.filter(dslam_id__in=dslam_ids).values_list('identifier_key', flat=True))

        reserved_identifier = set(
            CustomerPort.objects.filter(telecom_center_id=telecom_id).values_list('identifier_key', flat=True))
        free_identifier = list(all_identifier - reserved_identifier)

        if identifier_key:
            free_identifier = [item for item in free_identifier if identifier_key in item]
        return HttpResponse(json.dumps(free_identifier[0:10]), content_type='application/json; charset=UTF-8')

    def create(self, request, *args, **kwargs):
        telecom_center_id = request.data.get('telecom_center_id')
        calc_faulty_port = request.data.get('calc_faulty_port')

        if not telecom_center_id:
            return Response({'result': 'telecom id is not valid'}, status=status.HTTP_400_BAD_REQUEST)

        # get all telecom dslams
        telecom_center_obj = TelecomCenter.objects.get(id=telecom_center_id)
        prefix_bukht_name = telecom_center_obj.prefix_bukht_name
        dslams = DSLAM.objects.filter(telecom_center=telecom_center_obj).order_by('dslam_number')
        print((dslams.count()))
        # delete all bukht table on tc
        MDFDSLAM.objects.filter(telecom_center_id=telecom_center_id).delete()

        # Get all Bukht config TelecomCenterMDF
        telecomMDF_objs = TelecomCenterMDF.objects.filter(telecom_center=telecom_center_obj).order_by('priority', 'id')

        telecomBukht = []
        row_count = 1
        max_number = '4'

        # floor_connection_count = [[telecomMDF_obj.floor_count, telecomMDF_obj.connection_count] for telecomMDF_obj in telecomMDF_objs]
        # max_number = sum(item[0]*item[1] for item in floor_connection_count)

        for telecomMDF_obj in telecomMDF_objs:
            row_number = telecomMDF_obj.row_number
            conn_start = telecomMDF_obj.connection_start
            floor_start = telecomMDF_obj.floor_start

            conn_step = 1
            floor_step = 1

            # create floor range
            if telecomMDF_obj.floor_counting_status != "STANDARD":
                if telecomMDF_obj.floor_counting_status == "ODD":
                    if telecomMDF_obj.start_of_port % 2 == 0:
                        floor_start = floor_start + 1
                else:
                    if telecomMDF_obj.start_of_port % 2 == 1:
                        floor_start = floor_start + 1
                floor_step = 2

            floor_range = list(range(floor_start, telecomMDF_obj.floor_count + floor_start, floor_step))

            # create connection range
            if telecomMDF_obj.connection_counting_status != "STANDARD":
                if telecomMDF_obj.connection_counting_status == "ODD":
                    if telecomMDF_obj.connection_start % 2 == 0:
                        conn_start = telecomMDF_obj.connection_start + 1
                else:
                    if telecomMDF_obj.connection_connection_start % 2 == 1:
                        conn_start = telecomMDF_obj.connection_start + 1
                conn_step = 2

            connection_range = list(range(conn_start, conn_start + telecomMDF_obj.connection_count, conn_step))

            # create bukht table without dslam cart and port
            for floor_number in floor_range:
                for connection_number in connection_range:
                    telecomBukht.append({
                        'row_number': row_number,
                        'floor_number': floor_number,
                        'telecom_center_mdf_id': telecomMDF_obj.id,
                        'reseller': telecomMDF_obj.reseller,
                        'connection_number': connection_number,
                        'status_of_port': telecomMDF_obj.status_of_port,
                        'identifier_key': ('{0}-{1:0' + max_number + '}').format(str(prefix_bukht_name), row_count)
                    })
                    row_count += 1

        # add cart port dslam into bukht table
        telecomBukht_index = 0
        if dslams.count() > 0:
            faulty_ports = DSLAMPortFaulty.objects.filter(dslam_id__in=dslams.values_list('id', flat=True))
            for dslam in dslams:
                for dslam_cart in DSLAMCart.objects.filter(dslam=dslam).order_by('priority', 'id'):
                    slot_range = list(range(dslam_cart.cart_start, dslam_cart.cart_start + dslam_cart.cart_count))
                    port_range = list(range(dslam_cart.port_start, dslam_cart.port_start + dslam_cart.port_count))
                    try:
                        for slot in slot_range:
                            for port in port_range:
                                if calc_faulty_port:
                                    if faulty_ports.filter(slot_number=slot, port_number=port,
                                                           dslam_id=dslam.id).exists():
                                        continue
                                mdf_dslam = MDFDSLAM()
                                mdf_dslam.row_number = telecomBukht[telecomBukht_index].get('row_number')
                                mdf_dslam.telecom_center_mdf_id = telecomBukht[telecomBukht_index].get(
                                    'telecom_center_mdf_id')
                                mdf_dslam.floor_number = telecomBukht[telecomBukht_index].get('floor_number')
                                mdf_dslam.connection_number = telecomBukht[telecomBukht_index].get('connection_number')
                                mdf_dslam.identifier_key = telecomBukht[telecomBukht_index].get('identifier_key')
                                mdf_dslam.dslam_id = dslam.id
                                mdf_dslam.telecom_center_id = telecom_center_obj.id
                                mdf_dslam.slot_number = slot
                                mdf_dslam.port_number = port
                                mdf_dslam.status = telecomBukht[telecomBukht_index].get('status_of_port')
                                if mdf_dslam.status == 'RESELLER':
                                    reseller_obj = None
                                    reseller_obj = telecomBukht[telecomBukht_index].get('reseller')
                                    if reseller_obj:
                                        rp_obj, created = ResellerPort.objects.get_or_create(
                                            identifier_key=mdf_dslam.identifier_key, reseller=reseller_obj,
                                            telecom_center_id=mdf_dslam.telecom_center_id)
                                        mdf_dslam.reseller = reseller_obj
                                mdf_dslam.save()
                                telecomBukht_index += 1
                    except Exception as ex:
                        print()
                        ex
                        pass

        for row in telecomBukht[telecomBukht_index:]:
            mdf_dslam = MDFDSLAM()
            mdf_dslam.row_number = row.get('row_number')
            mdf_dslam.telecom_center_mdf_id = row.get('telecom_center_mdf_id')
            mdf_dslam.floor_number = row.get('floor_number')
            mdf_dslam.connection_number = row.get('connection_number')
            mdf_dslam.identifier_key = row.get('identifier_key')
            mdf_dslam.telecom_center_id = telecom_center_obj.id
            mdf_dslam.status = telecomBukht[telecomBukht_index].get('status_of_port')
            try:
                if mdf_dslam.status == 'RESELLER':
                    reseller_obj = None
                    reseller_obj = telecomBukht[telecomBukht_index].get('reseller')
                    if reseller_obj:
                        rp_obj, created = ResellerPort.objects.get_or_create(identifier_key=mdf_dslam.identifier_key,
                                                                             reseller=reseller_obj,
                                                                             telecom_center_id=mdf_dslam.telecom_center_id)
                        mdf_dslam.reseller = reseller_obj

                mdf_dslam.save()
            except Exception as ex:
                print()
                ex

        return Response({'result': 'Created mdf dslam'}, status=status.HTTP_201_CREATED)

    @action(methods=['GET'], detail=True)
    def download(self, request):
        def stream():
            buffer_ = StringIO.StringIO()
            writer = csv.writer(buffer_)
            writer.writerow(
                ['row number', 'floor number', 'connection number', 'card number', 'port nummber', 'identifier key',
                 'status'])
            rows = MDFDSLAM.objects.filter(telecom_center_id=telecom_center_id)
            for mdf_dslam_obj in rows:
                writer.writerow([mdf_dslam_obj.row_number, mdf_dslam_obj.floor_number, mdf_dslam_obj.connection_number, \
                                 mdf_dslam_obj.slot_number, mdf_dslam_obj.port_number, mdf_dslam_obj.identifier_key,
                                 mdf_dslam_obj.status])
                buffer_.seek(0)
                data = buffer_.read()
                buffer_.seek(0)
                buffer_.truncate()
                yield data

        data = request.query_params
        telecom_center_id = data.get('telecom_center_id')
        if telecom_center_id:
            response = StreamingHttpResponse(
                stream(), content_type='text/csv'
            )
            disposition = "attachment; filename=file.csv"
            response['Content-Disposition'] = disposition
            return response
        return None

    """
    Update a model instance.
    """

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        description = 'Update MDF DSLAM Status {0}: {1}'.format(instance.id, request.data)
        add_audit_log(request, 'MDFDSLAM', instance.id, 'Update MDF DSLAM', description)
        return Response(serializer.data)


class DSLAMBulkCommandResultViewSet(mixins.ListModelMixin,
                                    mixins.UpdateModelMixin,
                                    mixins.DestroyModelMixin,
                                    mixins.RetrieveModelMixin,
                                    viewsets.GenericViewSet):
    serializer_class = DSLAMBulkCommandResultSerializer
    permission_classes = (IsAuthenticated, HasAccessToDslamPort)
    pagination_class = LargeResultsSetPagination

    def get_queryset(self):
        user = self.request.user

        if user.type_contains(ADMIN) == False:
            return []
        queryset = DSLAMBulkCommandResult.objects.all().order_by('-created_at')
        title = self.request.query_params.get('title', None)
        if title:
            queryset = queryset.filter(title__icontains=title)

        return queryset


class DSLAMFaultyConfigViewSet(mixins.ListModelMixin,
                               mixins.CreateModelMixin,
                               mixins.UpdateModelMixin,
                               mixins.DestroyModelMixin,
                               mixins.RetrieveModelMixin,
                               viewsets.GenericViewSet):
    serializer_class = DSLAMFaultyConfigSerializer
    permission_classes = (IsAuthenticated,)
    queryset = DSLAMFaultyConfig.objects.all()

    def get_queryset(self):
        user = self.request.user
        dslam_id = self.request.query_params.get('dslam_id', None)
        queryset = self.queryset

        if dslam_id:
            queryset = queryset.filter(dslam_id=dslam_id)
        return queryset

    """
    Destroy a model instance.
    """

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        DSLAMPortFaulty.objects.filter(dslam_faulty_config=instance.id).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    """
    Create a model instance.
    """

    def create(self, request, *args, **kwargs):
        data = request.data
        slot_number_from = data.get('slot_number_from')
        slot_number_to = data.get('slot_number_to')
        port_number_from = data.get('port_number_from')
        port_number_to = data.get('port_number_to')
        dslam_id = data.get('dslam_id')
        faulty_configs = DSLAMFaultyConfig.objects.filter(dslam_id=dslam_id)
        for faulty_config in faulty_configs:
            if faulty_config.slot_number_from <= int(slot_number_from) and faulty_config.port_number_from <= int(
                    port_number_from):
                if faulty_config.slot_number_to >= int(slot_number_to) and faulty_config.port_number_to >= int(
                        port_number_to):
                    return Response({'result': 'Confilict Data'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        port_items = DSLAMPort.objects.filter(dslam_id=dslam_id).filter(slot_number__gte=slot_number_from,
                                                                        port_number__gte=port_number_from). \
            filter(slot_number__lte=slot_number_to).exclude(slot_number=slot_number_to,
                                                            port_number__gt=port_number_to).values('dslam_id',
                                                                                                   'slot_number',
                                                                                                   'port_number')
        DSLAMPortFaulty.objects.filter(dslam_id=dslam_id).delete()
        dslam_faulty_config_obj = DSLAMFaultyConfig.objects.get(id=serializer.data['id'])
        for port_item in port_items:
            try:
                dslam_id = port_item['dslam_id']
                port_number = port_item['port_number']
                slot_number = port_item['slot_number']
                portfaulty = DSLAMPortFaulty()
                portfaulty.dslam_id = dslam_id
                portfaulty.dslam_faulty_config = dslam_faulty_config_obj
                portfaulty.port_number = port_number
                portfaulty.slot_number = slot_number
                portfaulty.save()
            except Exception as ex:
                print()
                ex
                pass

        description = 'Add DSLAM Faulty Config {0}'.format(serializer.data['dslam_id'])
        add_audit_log(request, 'DSLAMFaultyConfig', serializer.data['id'], 'Add DSLAM Faulty Config', description)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    """
    Update a model instance.
    """

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        data = request.data
        slot_number_from = data.get('slot_number_from')
        slot_number_to = data.get('slot_number_to')
        port_number_from = data.get('port_number_from')
        port_number_to = data.get('port_number_to')
        dslam_id = data.get('dslam_id')

        faulty_configs = DSLAMFaultyConfig.objects.filter(dslam_id=dslam_id).exclude(id=instance.pk)
        for faulty_config in faulty_configs:
            if faulty_config.slot_number_from <= int(slot_number_from) and faulty_config.port_number_from <= int(
                    port_number_from):
                if faulty_config.slot_number_to >= int(slot_number_to) and faulty_config.port_number_to >= int(
                        port_number_to):
                    return Response({'result': 'Confilict Data'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        port_items = DSLAMPort.objects.filter(dslam_id=dslam_id).filter(slot_number__gte=slot_number_from,
                                                                        port_number__gte=port_number_from). \
            filter(slot_number__lte=slot_number_to).exclude(slot_number=slot_number_to,
                                                            port_number__gt=port_number_to).values('dslam_id',
                                                                                                   'slot_number',
                                                                                                   'port_number')

        DSLAMPortFaulty.objects.filter(dslam_faulty_config=instance).delete()
        dslam_faulty_config_obj = DSLAMFaultyConfig.objects.get(id=instance.id)
        for port_item in port_items:
            try:
                dslam_id = port_item['dslam_id']
                port_number = port_item['port_number']
                slot_number = port_item['slot_number']
                portfaulty = DSLAMPortFaulty()
                portfaulty.dslam_id = dslam_id
                portfaulty.dslam_faulty_config = dslam_faulty_config_obj
                portfaulty.port_number = port_number
                portfaulty.slot_number = slot_number
                portfaulty.save()
            except:
                pass

        description = 'Update DSLAM Falty Config {0}, params: {1}'.format(instance.id, request.data)
        add_audit_log(request, 'DSLAMFaltyConfig', instance.id, 'Update DSLAM Falty Config', description)
        return Response(serializer.data)


class DSLAMPortFaultyViewSet(mixins.ListModelMixin,
                             mixins.RetrieveModelMixin,
                             viewsets.GenericViewSet):
    serializer_class = DSLAMPortFaultySerializer
    permission_classes = (IsAuthenticated,)
    queryset = DSLAMPortFaulty.objects.all()

    def get_queryset(self):
        user = self.request.user
        dslam_id = self.request.query_params.get('dslam_id', None)
        queryset = self.queryset
        if dslam_id:
            queryset = queryset.filter(dslam_id=dslam_id).order_by('slot_number', 'port_number')
        return queryset


class QuickSearchView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def get(self, request, format=None):
        user = request.user
        value = request.query_params['value']

        dslam_objs = DSLAM.objects.all()
        port_objs = DSLAMPort.objects.all()
        mdf_dslam_objs = MDFDSLAM.objects.all()

        reseller_identifier_keys = None

        ports = []
        dslams = []

        if value:
            # search port
            user_ports = []
            mac_ports = []

            identifier_keys = list(
                CustomerPort.objects.filter(username__istartswith=value).values_list('identifier_key', flat=True))
            mdf_dslam_objs = mdf_dslam_objs.filter(identifier_key__in=identifier_keys)

            if user.type_contains(RESELLER):
                reseller = user.reseller
                reseller_identifier_keys = ResellerPort.objects.filter(reseller=reseller).values_list('identifier_key',
                                                                                                      flat=True)
                mdf_dslam_ports = mdf_dslam_objs.filter(identifier_key__in=reseller_identifier_keys).values('dslam_id',
                                                                                                            'slot_number',
                                                                                                            'port_number')
                dslam_ids = set([dslam.get('dslam_id') for dslam in mdf_dslam_ports])
                dslam_objs = dslam_objs.filter(id__in=dslam_ids).filter(
                    Q(ip=value) | Q(name__icontains=value) | Q(hostname__icontains=value)) | Q(fqdn__icontains=value)
            else:
                dslam_objs = dslam_objs.filter(
                    Q(ip=value) | Q(name__icontains=value) | Q(hostname__icontains=value) | Q(fqdn__icontains=value))
                mdf_dslam_ports = mdf_dslam_objs.values('dslam_id', 'slot_number', 'port_number')

            for mdf_dslam_port in mdf_dslam_ports:
                user_ports.append(port_objs.get(
                    dslam_id=mdf_dslam_port['dslam_id'],
                    slot_number=mdf_dslam_port['slot_number'],
                    port_number=mdf_dslam_port['port_number']).pk)

            if user_ports:
                port_objs = port_objs.filter(id__in=user_ports)

            mac_ports = list(
                DSLAMPortMac.objects.filter(mac_address__istartswith=value).values_list('port__id', flat=True))

            if mac_ports:
                port_objs = port_objs.filter(id__in=mac_ports)

            if user_ports or mac_ports:
                ports = [{'id': port.id, 'port_index': port.port_index, 'slot_number': port.slot_number,
                          'port_number': port.port_number, 'port_name': port.port_name, 'dslam_id': port.dslam.id,
                          'dslam_name': port.dslam.name, 'hostname': port.dslam.hostname} for port in port_objs]
                # end search port

            dslams = [{'id': dslam.id, 'name': dslam.name, 'hostname': dslam.hostname, 'ip': dslam.ip,
                       'type': dslam.dslam_type.name} for dslam in dslam_objs]
        return Response({'result': {'dslams': dslams, 'ports': ports}})


class RegisterPortAPIView(views.APIView):
    """
    "port":{
        "telecom_center_name": "string",
        "fqdn": "string",
        "card_number": "string",
        "port_number": "string"
    },
    "status": "RESELLER",
    "identifier_key":"string",
    "reseller": {
          "name": "string",
    },
    "subscriber":{
          "username": "string",
    }
    """

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        #return JsonResponse({'id': 201}, status=status.HTTP_201_CREATED)

        user = request.user
        data = request.data
        print(data)
        res = ''
        result = ''
        PVC = ''
        reseller_data = data.get('reseller')
        customer_data = data.get('subscriber')
        mdf_status = data.get('status')
        sid = ''
        PVC = ''
        pvc = ''
        if not mdf_status:
            mdf_status = 'BUSY'
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        identifier_key = data.get('identifier_key')
        if not identifier_key:
            identifier_key = str(random.randint(10000000, 9999999999999999))
        port_data = data.get('port')
        try:
            print(data)
            fqdn = port_data.get('fqdn')
            # return JsonResponse({'Result':'', 'id':201,}, status=status.HTTP_201_OK)

            if ('.z6' in fqdn):
                fqdn = fqdn.replace('.z6', '.Z6')

            # dslam_obj = DSLAM.objects.get(name=port_data.get('dslam_name'), telecom_center__name=port_data.get('telecom_center_name'))
            try:
                dslam_obj = DSLAM.objects.get(fqdn=fqdn)
                # return JsonResponse({'Result': serialize('json',[dslam_obj])}, status=status.HTTP_200_OK)
                telecomId = dslam_obj.telecom_center_id
                cityId = TelecomCenter.objects.get(id=telecomId).city_id
            except ObjectDoesNotExist as ex:
                try:
                    if ('.Z6' in fqdn):
                        fqdn = fqdn.replace('.Z6', '.z6')
                        dslam_obj = DSLAM.objects.get(fqdn=fqdn)
                        # return JsonResponse({'Result': dslam_obj.fqdn}, status=status.HTTP_200_OK)
                        telecomId = dslam_obj.telecom_center_id
                        cityId = TelecomCenter.objects.get(id=telecomId).city_id

                except ObjectDoesNotExist as ex:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    mail_info = Mail()
                    mail_info.from_addr = 'oss-notification@pishgaman.net'
                    mail_info.to_addr = 'oss-problems@pishgaman.net'
                    mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                        str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'),
                        port_data.get('card_number'), port_data.get('port_number'), ip)
                    mail_info.msg_subject = 'Error in RegisterPortAPIView'
                    # Mail.Send_Mail(mail_info)
                    return JsonResponse({'Result': 'Dslam Not Found. Please check FQDN again.'},
                                        status=status.HTTP_200_OK)

            telecom_mdf_obj = TelecomCenterMDF.objects.filter(telecom_center_id=dslam_obj.telecom_center.id)
            if telecom_mdf_obj:
                telecom_mdf_obj = telecom_mdf_obj.first()

            mdf_dslam_obj, mdf_dslam_updated = MDFDSLAM.objects.update_or_create(
                telecom_center_id=dslam_obj.telecom_center.id, telecom_center_mdf_id=telecom_mdf_obj.id,
                #### Check this whole line
                row_number=0, floor_number=0, connection_number=0,  ##### Check this whole line
                dslam_id=dslam_obj.id, slot_number=port_data.get('card_number'),
                port_number=port_data.get('port_number'),
                defaults={'status': mdf_status, 'identifier_key': identifier_key})
            # if mdf_dslam.status != 'FREE':
            #    return JsonResponse(
            #            {'result': 'port status is {0}'.format(mdf_dslam.status), 'id': -1}
            #            )
            # else:
            #    mdf_dslam.status = 'RESELLER'
            #    mdf_dslam.save()
            # identifier_key = mdf_dslam.identifier_key
        except ObjectDoesNotExist as ex:
            return JsonResponse({'result': str(ex), 'id': -1})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            try:
                return JsonResponse({'result': 'Error is {0}--{1}'.format(ex, exc_tb.tb_lineno)})
            except ObjectDoesNotExist as ex:
                mail_info = Mail()
                mail_info.from_addr = 'oss-notification@pishgaman.net'
                mail_info.to_addr = 'oss-problems@pishgaman.net'
                mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                    str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'), port_data.get('card_number'),
                    port_data.get('port_number'), ip)
                mail_info.msg_subject = 'Error in RegisterPortAPIView'
                Mail.Send_Mail(mail_info)
                return JsonResponse({'Result': ''}, status=status.HTTP_400_BAD_REQUEST)
        try:
            reseller_obj, reseller_created = Reseller.objects.get_or_create(name=reseller_data.get('name'))
            print()
            'reseller', reseller_obj

            customer_obj, customer_updated = CustomerPort.objects.update_or_create(
                username=customer_data.get('username'),
                defaults={'identifier_key': identifier_key, 'telecom_center_id': mdf_dslam_obj.telecom_center_id}
            )

            rp = ResellerPort()
            rp.identifier_key = identifier_key
            rp.status = 'ENABLE'
            rp.dslam_id = dslam_obj.id
            rp.dslam_slot = port_data.get('card_number')
            rp.dslam_port = port_data.get('port_number')
            rp.reseller = reseller_obj
            rp.telecom_center_id = mdf_dslam_obj.telecom_center_id
            rp.save()

            vlan_objs = Vlan.objects.filter(reseller=reseller_obj)
            print()
            'vlan_objs ->', vlan_objs
            # return JsonResponse({'vlan':  vlan_objs[0].vlan_id, 'vpi' : reseller_obj.vpi,'vci' : reseller_obj.vci})
            port_indexes = [{'slot_number': port_data.get('card_number'), 'port_number': port_data.get('port_number')}]
            pishParams = {
                "type": "dslamport",
                "is_queue": False,
                "vlan_id": 3900,
                "vlan_name": 'pte',
                "dslam_id": dslam_obj.id,
                "port_indexes": port_indexes,
                "username": customer_data.get('username'),
                "vpi": 0,
                "vci": 35,
            }
            # 11111
            params = {
                "type": "dslamport",
                "is_queue": False,
                "vlan_id": vlan_objs[0].vlan_id,
                "vlan_name": vlan_objs[0].vlan_name,
                "dslam_id": dslam_obj.id,
                "port_indexes": port_indexes,
                "username": customer_data.get('username'),
                "vpi": reseller_obj.vpi,
                "vci": reseller_obj.vci,
            }
            # return JsonResponse({'res' : params })
            result = ''

            try:
                if (cityId == 11111111111):
                    # return JsonResponse({'result': str('Email Sent To rt-network-access!!!'), 'id': 201}, status=status.HTTP_202_ACCEPTED)
                    q = "select * from shabdizrt where agent = '{0}' and ip='{1}' and fqdn='{2}' and card = '{3}' and port= '{4}' and status= '{5}'".format(
                        reseller_obj, dslam_obj.ip, dslam_obj.fqdn, port_data.get('card_number'),
                        port_data.get('port_number'), "sent")

                    cursor = connection.cursor()
                    cursor.execute(q)
                    rows = cursor.fetchall()
                    if (cursor.rowcount > 0):
                        return JsonResponse(
                            {'result': str('Email already has been sent to rt-access. Please wait for response.'),
                             'id': 201}, status=status.HTTP_202_ACCEPTED)

                    query = "INSERT INTO shabdizrt values('{0}','{1}', '{2}', '{3}', '{4}', '{5}')".format(reseller_obj,
                                                                                                           dslam_obj.ip,
                                                                                                           dslam_obj.fqdn,
                                                                                                           port_data.get(
                                                                                                               'card_number'),
                                                                                                           port_data.get(
                                                                                                               'port_number'),
                                                                                                           "sent")
                    # return JsonResponse({'result': str(cursor.rowcount), 'id': 201}, status=status.HTTP_202_ACCEPTED)

                    cursor = connection.cursor()
                    cursor.execute(query)

                    vpi = Reseller.objects.get(name=reseller_data['name']).vpi
                    vci = Reseller.objects.get(name=reseller_data['name']).vci
                    fromaddr = "oss-notification@pishgaman.net"
                    toaddr = "rt-network-access@pishgaman.net"
                    msg = MIMEMultipart()
                    msg['From'] = fromaddr
                    msg['To'] = toaddr
                    msg['Subject'] = "OSS Shabdiz Dslams"
                    body = 'Command: add to vlan for Shabdiz Dslams, IP: {0}, fqdn: {1} , Card: {2} , Port: {3} , VlanId: {4},Reseller: {5},Vpi: {6},Vci: {7}'.format(
                        dslam_obj.ip, dslam_obj.fqdn, port_data.get('card_number'), port_data.get('port_number'),
                        vlan_objs[0].vlan_id, reseller_data['name'], vpi, vci)
                    msg.attach(MIMEText(body, 'plain'))
                    server = smtplib.SMTP('mail.pishgaman.net', 25)
                    server.login("oss-notification@pishgaman.net", "Os123456")
                    text = msg.as_string()
                    server.sendmail(fromaddr, toaddr, text)
                    return JsonResponse({'result': str('Email Sent To rt-network-access!!!'), 'id': 201},
                                        status=status.HTTP_202_ACCEPTED)
                # elif(dslam_obj.dslam_type_id == 4):
                elif (dslam_obj.dslam_type_id == 4 or dslam_obj.dslam_type_id == 3 or dslam_obj.dslam_type_id == 5):
                    try:
                        dslamType = ''
                        if (dslam_obj.dslam_type_id == 4):
                            dslamType = 'Fiberhome'
                            if (dslam_obj.id == 1231):
                                sourceVlanId = '2576'
                            else:
                                sourceVlanId = '3900'
                        elif (dslam_obj.dslam_type_id == 3):
                            sourceVlanId = '3900'
                            dslamType = 'FiberhomeAN3300'

                        elif (dslam_obj.dslam_type_id == 5):
                            dslamType = 'FiberhomeAN5006'
                            sourceVlanId = '3900'

                        vlanName = str(reseller_obj).split('-')[1]
                        if (vlanName == 'didehban'):
                            vlanName = 'dideban'
                        if (vlanName == 'baharsamaneh'):
                            vlanName = 'baharsam'
                        if (vlanName == 'badrrayan'):
                            vlanName = 'badrray'
                        if (vlanName == 'farzanegan'):
                            vlanName = 'farzaneg'

                        url = 'http://172.28.244.138:9096/api/Telnet/telnet'
                        data = "{'type':'%s','dslam':'%s','telnetPort':'23','userName':'%s','password':'%s','access':'%s','sourceVlanId':'%s','vlanName':'%s','vlanId':'%s','untaggedPortList':'%s','vpiVci':'%s','card':'%s','port':'%s','command':'%s','terminalDelay':'600','requestTimeOut':'1500'}" % (
                            dslamType, dslam_obj.ip, dslam_obj.telnet_username, dslam_obj.telnet_password,
                            dslam_obj.access_name, sourceVlanId, vlanName, vlan_objs[0].vlan_id,
                            '0-{0}-{1}'.format(port_data.get('card_number'), port_data.get('port_number')),
                            '{0}/{1}'.format(Reseller.objects.get(name=reseller_data['name']).vpi,
                                             Reseller.objects.get(name=reseller_data['name']).vci),
                            port_data.get('card_number'), port_data.get('port_number'), 'addToVlan')
                        # return JsonResponse({'result':data })
                        fhresponse = requests.post(url, data=data, headers={"Content-Type": "application/json"})
                        print(fhresponse.json())
                        sid = fhresponse.json()

                    except Exception as ex:
                        exc_type, exc_obj, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        mail_info = Mail()
                        mail_info.from_addr = 'oss-notification@pishgaman.net'
                        mail_info.to_addr = 'oss-problems@pishgaman.net'
                        mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                            str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'),
                            port_data.get('card_number'), port_data.get('port_number'), ip)
                        mail_info.msg_subject = 'Error in RegisterPortAPIView'
                        # Mail.Send_Mail(mail_info)
                        print(str(ex))
                        return JsonResponse(
                            {'result': str('an error occurred. please try again. {0}'.format(str(exc_tb.tb_lineno)))})
                elif (dslam_obj.dslam_type_id == 7):
                    vpi = Reseller.objects.get(name=reseller_data['name']).vpi
                    vci = Reseller.objects.get(name=reseller_data['name']).vci
                    fromaddr = "oss-notification@pishgaman.net"
                    toaddr = "rt-network-access@pishgaman.net"
                    msg = MIMEMultipart()
                    msg['From'] = fromaddr
                    msg['To'] = toaddr
                    msg['Subject'] = "OSS Problem"
                    body = 'Command: add to vlan, IP: {0}, fqdn: {1} , Card: {2} , Port: {3} , VlanId: {4},Reseller: {5},Vpi: {6},Vci: {7}'.format(
                        dslam_obj.ip, dslam_obj.fqdn, port_data.get('card_number'), port_data.get('port_number'),
                        vlan_objs[0].vlan_id, reseller_data['name'], vpi, vci)
                    msg.attach(MIMEText(body, 'plain'))
                    server = smtplib.SMTP('mail.pishgaman.net', 25)
                    server.login("oss-notification@pishgaman.net", "Os123456")
                    text = msg.as_string()
                    server.sendmail(fromaddr, toaddr, text)
                    # return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})
                    return JsonResponse(
                        {'result': str('An Error Ocurred for add to vlan Command. Email Sent To rt-network-access!!!'),
                         'id': 201},
                        status=status.HTTP_202_ACCEPTED)


                else:
                    # res = utility.dslam_port_run_command(dslam_obj.id, 'delete from vlan', pishParams)
                    result = utility.dslam_port_run_command(dslam_obj.id, 'add to vlan', params)
                    pvc = utility.dslam_port_run_command(dslam_obj.id, 'port pvc show', params)

                    # result2 = utility.dslam_port_run_command(dslam_obj.id, 'add to vlan', params)

            except Exception as ex:
                if (dslam_obj.dslam_type_id == 5 or dslam_obj.dslam_type_id == 3 or dslam_obj.dslam_type_id == 7):
                    vpi = Reseller.objects.get(name=reseller_data['name']).vpi
                    vci = Reseller.objects.get(name=reseller_data['name']).vci
                    fromaddr = "oss-notification@pishgaman.net"
                    toaddr = "rt-network-access@pishgaman.net"
                    msg = MIMEMultipart()
                    msg['From'] = fromaddr
                    msg['To'] = toaddr
                    msg['Subject'] = "OSS Problem"
                    cc = ['oss-problems@pishgaman.net']
                    body = 'Command: add to vlan, IP: {0}, fqdn: {1} , Card: {2} , Port: {3} , VlanId: {4},Reseller: {5},Vpi: {6},Vci: {7}'.format(
                        dslam_obj.ip, dslam_obj.fqdn, port_data.get('card_number'), port_data.get('port_number'),
                        vlan_objs[0].vlan_id, reseller_data['name'], vpi, vci)
                    msg.attach(MIMEText(body, 'plain'))
                    server = smtplib.SMTP('mail.pishgaman.net', 25)
                    server.login("oss-notification@pishgaman.net", "Os123456")
                    text = msg.as_string()
                    server.sendmail(fromaddr, toaddr, text)
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    # return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})
                    return JsonResponse(
                        {'result': str('An Error Ocurred for add to vlan Command. Email Sent To rt-network-access!!!')},
                        status=status.HTTP_201_CREATED)

                else:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    # mail_info = Mail()
                    # mail_info.from_addr = 'oss-notification@pishgaman.net'
                    # mail_info.to_addr = 'oss-problems@pishgaman.net'
                    # mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                    #     str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'),
                    #     port_data.get('card_number'), port_data.get('port_number'), ip)
                    # mail_info.msg_subject = 'Error in RegisterPortAPIView'
                    # Mail.Send_Mail(mail_info)
                    return JsonResponse(
                        {'result': str('an error occurred. please try again. {0}'.format(str(exc_tb.tb_lineno)))})

            if vlan_objs.count() > 0:
                try:
                    port_obj = DSLAMPort.objects.get(dslam=dslam_obj, slot_number=port_data.get('card_number'),
                                                     port_number=port_data.get('port_number'))
                except ObjectDoesNotExist as ex:
                    dslam_port = DSLAMPort()
                    dslam_port.dslam = dslam_obj
                    dslam_port.created_at = datetime.now()
                    dslam_port.updated_at = datetime.now()
                    dslam_port.slot_number = port_data.get('card_number')
                    dslam_port.port_number = port_data.get('port_number')
                    dslam_port.port_index = '{0}{1}'.format(port_data.get('card_number'), port_data.get('port_number'))
                    dslam_port.port_name = 'adsl{0}-{1}'.format(port_data.get('card_number'),
                                                                port_data.get('port_number'))
                    dslam_port.admin_status = 'UNLOCK'
                    dslam_port.oper_status = 'NO-SYNC'
                    dslam_port.line_profile = '768/10240/IR'
                    dslam_port.selt_value = ''
                    dslam_port.uptime = '18:55:20'
                    dslam_port.upstream_snr = '192'
                    dslam_port.downstream_snr = '230'
                    dslam_port.upstream_attenuation = '69'
                    dslam_port.downstream_attenuation = '188'
                    dslam_port.upstream_attainable_rate = '2434'
                    dslam_port.downstream_attainable_rate = '157'
                    dslam_port.upstream_tx_rate = '93'
                    dslam_port.downstream_tx_rate = '1250'
                    dslam_port.upstream_snr_flag = 'good'
                    dslam_port.downstream_snr_flag = 'excellent'
                    dslam_port.upstream_attenuation_flag = 'outstanding'
                    dslam_port.downstream_attenuation_flag = 'outstanding'
                    dslam_port.vpi = 0
                    dslam_port.vci = 35
                    dslam_port.save()
                    port_obj = DSLAMPort.objects.get(dslam=dslam_obj, slot_number=port_data.get('card_number'),
                                                     port_number=port_data.get('port_number'))
                port_vlan_obj = DSLAMPortVlan()
                port_vlan_obj.vlan = vlan_objs.first()
                port_vlan_obj = port_obj
                port_vlan_obj.save()
            if dslam_obj.dslam_type_id == 4:
                if 'succeed' in sid:
                    return JsonResponse({'id': 201, 'res': sid, 'msg': 'port config has been done.'},
                                        status=status.HTTP_201_CREATED)
                else:
                    return JsonResponse({'result': 'Error', 'ErrorDesc': sid, 'id': 400, 'res': 'Error'},
                                        status=status.HTTP_400_BAD_REQUEST)
            if dslam_obj.dslam_type_id == 3:
                if ('added to vlan' in sid):
                    return JsonResponse({'PVC': PVC, 'id': 201, 'res': sid, 'msg': 'port config has been done.'},
                                        status=status.HTTP_201_CREATED)
                else:
                    return JsonResponse({'result': 'Error', 'ErrorDesc': sid, 'id': 400, 'res': 'Error'},
                                        status=status.HTTP_400_BAD_REQUEST)
            if dslam_obj.dslam_type_id == 1:
                # return JsonResponse({"result": PVC})

                if isinstance(pvc, str):
                    return JsonResponse({'PVC': pvc, 'id': 400, 'msg': 'port config has not been done.'},
                                        status=status.HTTP_400_BAD_REQUEST)
                print('++++++++++++++++++++++++++++++++')
                print(pvc)
                print('++++++++++++++++++++++++++++++++')
                for PVC in pvc['result']:
                    if PVC['pvc'] == '{0}-{1}-{2}/{3}'.format(port_data.get('card_number'),
                                                              port_data.get('port_number'), reseller_obj.vpi,
                                                              reseller_obj.vci) and PVC['pvid'] == vlan_objs[0].vlan_id:
                        print(PVC['pvc'])
                        return JsonResponse({'PVC': PVC, 'id': 201, 'res': sid, 'msg': 'port config has been done.'},
                                            status=status.HTTP_201_CREATED)

                else:
                    return JsonResponse({'PVC': pvc, 'id': 400, 'res': sid, 'msg': 'port config has not been done.'},
                                        status=status.HTTP_400_BAD_REQUEST)
            # return JsonResponse({'result':'Port is registered', 'PVC': PVC , 'id': 201, 'res': sid}, status=status.HTTP_201_CREATED)
            if dslam_obj.dslam_type_id == 5:
                if 'attach pvc profile name' in sid:
                    return JsonResponse({'id': 201, 'msg': 'port config has been done.'},
                                        status=status.HTTP_201_CREATED)
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # mail_info = Mail()
            # mail_info.from_addr = 'oss-notification@pishgaman.net'
            # mail_info.to_addr = 'oss-problems@pishgaman.net'
            # mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
            #     str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'), port_data.get('card_number'),
            #     port_data.get('port_number'), ip)
            # mail_info.msg_subject = 'Error in RegisterPortAPIView'
            # Mail.Send_Mail(mail_info)

            # return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})
            return JsonResponse(
                {'result': str('an error occurred. please try again. {0}-{1}'.format(str(ex), str(exc_tb.tb_lineno)))},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class showPort():
    dslamName = ''
    cammandName = ''
    slotPort = ''
    link = ''
    payloadrate = 0
    actualrate = 0
    attainablerate = 0
    noisemargin = 0
    attenuation = 0
    userProfile = ''


class DslanInfo():
    type = ''
    dslam = ''
    telnetPort = 23
    userName = ''
    password = ''
    access = ''
    card = ''
    port = ''
    command = ''
    terminalDelay = ''
    requestTimeOut = ''
    userProfile = ''


class fInfo():
    Date = ''
    cardInfo = ''
    OP_State = ''
    Standard = ''
    Latency = ''
    Stream_SNR_Margin = ''
    Rate = ''
    Stream_attenuation = ''
    Tx_power = ''
    Remote_vendor_ID = ''
    Power_management_state = ''
    Remote_vendor_version_ID = ''
    Loss_of_power = ''
    Errored_frames = ''
    Loss_of_signal = ''
    Error_seconds = ''
    Loss_by_HEC_collision = ''
    Forward_correct = ''
    Uncorrect = ''
    Attainable_rate = ''
    Interleaved_Delay = ''
    Remote_loss_of_link = ''


class RunCommandAPIView(views.APIView):
    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip = x_forwarded_for.split(',')[0]
            else:
                ip = request.META.get('REMOTE_ADDR')

            user = request.user
            data = request.data
            print(ip)
            # return JsonResponse({'Result': request.data.get('fqdn')}, status=status.HTTP_200_OK)
            s = []
            command = data.get('command', None)
            subscriber = data.get('subscriber')
            fqdn = request.data.get('fqdn')
            if ('.z6' in fqdn):
                fqdn = fqdn.replace('.z6', '.Z6')
            try:
                dslamObj = DSLAM.objects.get(fqdn=fqdn)
                dslam_id = dslamObj.id
                dslam_ip = dslamObj.ip
            except ObjectDoesNotExist as ex:
                try:
                    if ('.Z6' in fqdn):
                        fqdn = fqdn.replace('.Z6', '.z6')
                    dslamObj = DSLAM.objects.get(fqdn=fqdn)
                    dslam_id = dslamObj.id
                    dslam_ip = dslamObj.ip
                except Exception as ex:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    mail_info = Mail()
                    mail_info.from_addr = 'oss-notification@pishgaman.net'
                    mail_info.to_addr = 'oss-problems@pishgaman.net'
                    mail_info.msg_body = 'Error in RunCommandAPIViewin/Fiberhome2200 Line {0}.Error Is: {1}. fqdn = {2},card = {3}, port = {4}, command = {5}, subscriber = {6}. IP: {7}'.format(
                        str(exc_tb.tb_lineno), str(ex), request.data.get('fqdn'),
                        request.data.get('params').get('port_conditions').get('slot_number'),
                        request.data.get('params').get('port_conditions').get('port_number'), command, subscriber, ip)
                    mail_info.msg_subject = 'Error in RunCommandAPIView'
                    Mail.Send_Mail(mail_info)
                    return JsonResponse({'Result': ''}, status=status.HTTP_200_OK)

            userProfile = ''  # DSLAMPort.objects.get(port_number = request.data.get('params').get('port_conditions').get('port_number'), slot_number = request.data.get('params').get('port_conditions').get('slot_number'), dslam_id = dslam_id).line_profile
            params = data.get('params', None)
            try:
                dslam_obj = DSLAM.objects.get(id=dslam_id)
                fiber = DslanInfo()
                fiber.type = 'Fiberhome'
                fiber.dslam = dslam_obj.ip
                fiber.telnetPort = 23
                fiber.userName = dslam_obj.telnet_username
                fiber.password = dslam_obj.telnet_password
                fiber.access = dslam_obj.access_name
                fiber.card = request.data.get('params').get('port_conditions').get('slot_number')
                fiber.port = request.data.get('params').get('port_conditions').get('port_number')
                fiber.terminalDelay = '300'
                fiber.requestTimeOut = '1500'
                # ---------------------------------------------------Fiberhome2200---------------------------------------------------
                if (dslam_obj.dslam_type_id == 4):
                    try:
                        profile = request.data.get('params').get('new_lineprofile')
                        if (command == 'selt'):
                            return JsonResponse({'res': 'this command is not supported by this dslam'})
                        if (command == 'show linerate' or command == 'showPort' or command == 'showport'):
                            fiber.command = 'showPort'
                        if (command == 'profile adsl show' or command == 'showProfiles'):
                            fiber.command = 'showProfiles'
                        if (
                                command == 'setPortProfiles' or command == 'Set Port Profiles' or command == 'profile adsl set' or command == 'change lineprofile port'):
                            fiber.command = 'setPortProfiles'
                        url = 'http://5.202.129.61:9096/api/Telnet/telnet'
                        data = "{'type':'Fiberhome','dslam':'%s','telnetPort':'23','userName':'%s','password':'%s','access':'%s','card':'%s','port':'%s','command':'%s','profile':'%s','terminalDelay':'600','requestTimeOut':'1500'}" % (
                            fiber.dslam, fiber.userName, fiber.password, fiber.access, fiber.card, fiber.port,
                            fiber.command,
                            profile)
                        # return JsonResponse({'response':data})

                        fhresponse = requests.post(url, data=data, headers={"Content-Type": "application/json"})
                        sid = fhresponse.json()
                        try:
                            if (command == 'show linerate' or command == 'showPort' or command == 'showport'):
                                res = sid.split("\n\r")

                                fhome = fInfo()
                                fhome.Date = sid.split("\n\r")[0]
                                portInfo = [int(c) for c in res[3].split() if c.isdigit()]

                                return JsonResponse({'current_userProfile': userProfile,
                                                     'dslamName/cammandName': dslam_obj.name + '/' + command,
                                                     'date': res[0].split(":")[1] + res[0].split(":")[2] + ':' +
                                                             res[0].split(":")[3],
                                                     'slot/port': str(portInfo[1]) + '-' + str(portInfo[2]),
                                                     # 'OP_State' : res[4].split(":")[1],
                                                     # 'Standard' : res[5].split(":")[1],
                                                     # 'Latency' : res[6].split(":")[1],
                                                     'noisemarginDown': res[7].split(":")[1].split("/")[0],
                                                     'noisemarginUp': res[7].split(":")[1].split("/")[1].split(" ")[0],
                                                     'payloadrateDown': res[8].split(":")[1].split("/")[0],
                                                     'payloadrateUp': res[8].split(":")[1].split("/")[1].split(" ")[0],
                                                     'attenuationDown': res[9].split(":")[1].split("/")[0],
                                                     'attenuationUp': res[9].split(":")[1].split("/")[1].split(" ")[0],
                                                     # 'Tx power(D/U)' : res[10].split(":")[1].split("/")[0],
                                                     # 'Tx power(D/U)' : res[10].split(":")[1].split("/")[1].split(" ")[0],
                                                     # 'Remote vendor ID' : res[11].split(":")[1],
                                                     # 'Power management state' : res[12].split(":")[1],
                                                     # 'Remote vendor version ID' : res[13].split(":")[1],
                                                     # 'Loss of power(D)' : res[14].split(":")[1].split("/")[0],
                                                     # 'Loss of power(U)' : res[14].split(":")[1].split("/")[1].split(" ")[0],
                                                     # 'Loss of signal(D)' : res[15].split(":")[1].split("/")[0],
                                                     # 'Loss of signal(U)' : res[15].split(":")[1].split("/")[1].split(" ")[0],
                                                     # 'Error seconds(D)' : res[16].split(":")[1].split("/")[0],
                                                     # 'Error seconds(U)' : res[16].split(":")[1].split("/")[1].split(" ")[0],
                                                     # 'Loss by HEC collision(D)' : res[17].split(":")[1].split("/")[0],
                                                     # 'Loss by HEC collision(U)' : res[17].split(":")[1].split("/")[1].split(" ")[0],
                                                     # 'Forward correct(D)' : res[18].split(":")[1].split("/")[0],
                                                     # 'Forward correct(U)' : res[18].split(":")[1].split("/")[1],
                                                     # 'Uncorrect(D)' : res[19].split(":")[1].split("/")[0],
                                                     # 'Uncorrect(U)' : res[19].split(":")[1].split("/")[1],
                                                     'attainablerateDown': res[20].split(":")[1].split("/")[0],
                                                     'attainablerateUp': res[20].split(":")[1].split("/")[1],
                                                     # 'Interleaved Delay(D) ' : res[21].split(":")[1].split("/")[0],
                                                     # 'Interleaved Delay(U) ' : res[21].split(":")[1].split("/")[1],
                                                     # 'Remote loss of link' : res[22].split(":")[1],
                                                     })
                                return JsonResponse({'current_userProfile': userProfile, 'response': sid.split("\n\r")})
                            elif (command == 'profile adsl show' or command == 'showProfiles'):
                                return JsonResponse({'current_userProfile': userProfile, 'response': sid.split("\n\r"),
                                                     'DslamType': 'fiberhomeAN2200',
                                                     'message': 'please send profile Id'})
                            elif (command == 'set port profile' or command == 'profile Change'):
                                return JsonResponse({'current_userProfile': userProfile, 'response': sid,
                                                     'DslamType': 'fiberhomeAN2200'})

                            else:
                                return JsonResponse({'current_userProfile': userProfile, 'response': sid.split("\n\r"),
                                                     'DslamType': 'fiberhomeAN2200'})
                        except Exception as ex:
                            try:
                                if ('handshake' in sid):
                                    try:
                                        return JsonResponse({'slot/port': str(portInfo[1]) + '-' + str(portInfo[2]),
                                                             'current_userProfile': userProfile,
                                                             'OP_State': 'port is down',
                                                             'Power management state': 'UNKNOWN STATE'})
                                    except Exception as ex:
                                        return JsonResponse({'current_userProfile': userProfile, 'response': sid})
                            except Exception as ex:
                                exc_type, exc_obj, exc_tb = sys.exc_info()
                                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                mail_info = Mail()
                                mail_info.from_addr = 'oss-notification@pishgaman.net'
                                mail_info.to_addr = 'oss-problems@pishgaman.net'
                                mail_info.msg_body = 'Error in RunCommandAPIViewin/Fiberhome2200 Line {0}.Error Is: {1}. fqdn = {2},card = {3}, port = {4}, command = {5}, subscriber = {6}. IP: {7}'.format(
                                    str(exc_tb.tb_lineno), str(ex), request.data.get('fqdn'),
                                    request.data.get('params').get('port_conditions').get('slot_number'),
                                    request.data.get('params').get('port_conditions').get('port_number'), command,
                                    subscriber, ip)
                                mail_info.msg_subject = 'Error in RunCommandAPIView/Fiberhome2200'
                                Mail.Send_Mail(mail_info)
                                return JsonResponse({'current_userProfile': 'Error'})

                    except Exception as ex:
                        exc_type, exc_obj, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        mail_info = Mail()
                        mail_info.from_addr = 'oss-notification@pishgaman.net'
                        mail_info.to_addr = 'oss-problems@pishgaman.net'
                        mail_info.msg_body = 'Error in RunCommandAPIViewin/Fiberhome2200 Line {0}.Error Is: {1}. fqdn = {2},card = {3}, port = {4}, command = {5}, subscriber = {6}. IP: {7}'.format(
                            str(exc_tb.tb_lineno), str(ex), request.data.get('fqdn'),
                            request.data.get('params').get('port_conditions').get('slot_number'),
                            request.data.get('params').get('port_conditions').get('port_number'), command,
                            subscriber, ip)
                        mail_info.msg_subject = 'Error in RunCommandAPIView/Fiberhome2200'
                        Mail.Send_Mail(mail_info)
                        return JsonResponse({'current_userProfile': 'Error'})

                # ---------------------------------------------------Fiberhome5006---------------------------------------------------
                elif (dslam_obj.dslam_type_id == 5):
                    profile = request.data.get('params').get('new_lineprofile')
                    if (command == 'selt'):
                        return JsonResponse({'res': 'this command is not supported by this dslam'})
                    if (command == 'show linerate' or command == 'showPort' or command == 'showport'):
                        fiber.command = 'showPort'
                    if (command == 'profile adsl show' or command == 'showProfiles'):
                        fiber.command = 'showProfiles'
                    if (
                            command == 'set Profile' or command == 'setProfile' or command == 'setPortProfiles' or command == 'setProfile'):
                        fiber.command = 'setProfile'
                    url = 'http://5.202.129.88:9096/api/Telnet/telnet'
                    data = "{'type':'FiberhomeAN5006','dslam':'%s','telnetPort':'23','userName':'%s','password':'%s','access':'%s','card':'%s','port':'%s','command':'%s','profile':'%s','terminalDelay':'600','requestTimeOut':'1500'}" % (
                        fiber.dslam, fiber.userName, fiber.password, fiber.access, fiber.card, fiber.port,
                        fiber.command,
                        profile)
                    # return JsonResponse({'response':data})
                    fhresponse = requests.post(url, data=data, headers={"Content-Type": "application/json"})
                    sid = fhresponse.json()
                    res = sid.split("\r\n")
                    if (fiber.command == 'showProfiles'):
                        my_list = []
                        for x in res:
                            if ('name' in x):
                                my_list.append(x.split(': ')[1])
                        return JsonResponse(
                            {'current_userProfile': userProfile, 'response': my_list, 'DslamType': 'fiberhomeAN5006'})
                    return JsonResponse({'current_userProfile': userProfile, 'response': res})


                # ---------------------------------------------------Fiberhome3300---------------------------------------------------
                elif (dslam_obj.dslam_type_id == 3):
                    profile = request.data.get('params').get('new_lineprofile')
                    if (command == 'show linerate' or command == 'showPort' or command == 'showport'):
                        fiber.command = 'showPort'
                    if (command == 'profile adsl show' or command == 'showProfiles'):
                        fiber.command = 'showProfiles'
                    if (
                            command == 'setPortProfiles' or command == 'Set Port Profiles' or command == 'profile adsl set' or command == 'change lineprofile port'):
                        fiber.command = 'setProfile'

                    url = 'http://5.202.129.88:9096/api/Telnet/telnet'
                    data = "{'type':'FiberhomeAN3300','dslam':'%s','telnetPort':'23','userName':'%s','password':'%s','access':'%s','card':'%s','port':'%s','command':'%s','profile':'%s','terminalDelay':'600','requestTimeOut':'1500'}" % (
                        fiber.dslam, fiber.userName, fiber.password, fiber.access, fiber.card, fiber.port,
                        fiber.command,
                        profile)
                    # return JsonResponse({'response':data})
                    fhresponse = requests.post(url, data=data, headers={"Content-Type": "application/json"})
                    sid = fhresponse.json()
                    res = sid.split("\n\r")
                    if (command == 'show linerate' or command == 'showPort' or command == 'showPort'):
                        try:
                            if ('Handshake' in sid):
                                return JsonResponse({'DslamType': 'fiberhomeAN3300',
                                                     'Port state': sid.split("\r\n")[5].split(":")[1].split()[0],
                                                     'Link state': sid.split("\r\n")[6].split(":")[1].split()[0],
                                                     'Profile Name': sid.split("\r\n")[7].split(":")[1].split()[0],
                                                     'OP State': sid.split("\r\n")[9].split(":")[1].split()[0]})
                            else:
                                return JsonResponse({'DslamType': 'fiberhomeAN3300',
                                                     'Port state': sid.split("\r\n")[5].split(":")[1].split()[0],
                                                     'Link state': sid.split("\r\n")[6].split(":")[1].split()[0],
                                                     'Profile Name': sid.split("\r\n")[7].split(":")[1].split()[0],
                                                     'OP State': sid.split("\r\n")[9].split(":")[1].split()[0],
                                                     'DownStream rate': sid.split("\r\n")[10].split(":")[1].split()[0],
                                                     'DownStream attain rate':
                                                         sid.split("\r\n")[11].split(":")[1].split()[0],
                                                     'DownStream Margin': sid.split("\r\n")[12].split(":")[1].split()[
                                                         0],
                                                     'DownStream Attenuat': sid.split("\r\n")[13].split(":")[1].split()[
                                                         0],
                                                     'DownStream Tx Power': sid.split("\r\n")[14].split(":")[1].split()[
                                                         0],
                                                     'DownStream Int Delay':
                                                         sid.split("\r\n")[15].split(":")[1].split()[0],
                                                     'PM STATE': sid.split("\r\n")[15].split(":")[1].split()[0]})
                        except Exception as ex:
                            exc_type, exc_obj, exc_tb = sys.exc_info()
                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                            mail_info = Mail()
                            mail_info.from_addr = 'oss-notification@pishgaman.net'
                            mail_info.to_addr = 'oss-problems@pishgaman.net'
                            mail_info.msg_body = 'Error in RunCommandAPIViewin/Fiberhome3300 Line {0}.Error Is: {1}. fqdn = {2},card = {3}, port = {4}, command = {5}, subscriber = {6}. IP: {7}'.format(
                                str(exc_tb.tb_lineno), str(ex), request.data.get('fqdn'),
                                request.data.get('params').get('port_conditions').get('slot_number'),
                                request.data.get('params').get('port_conditions').get('port_number'), command,
                                subscriber, ip)
                            mail_info.msg_subject = 'Error in RunCommandAPIView/Fiberhome3300'
                            Mail.Send_Mail(mail_info)
                            return JsonResponse({'current_userProfile': userProfile, 'response': sid.split("\r\n")})

                    if (command == 'profile adsl show' or command == 'showProfiles'):
                        return JsonResponse({'current_userProfile': userProfile, 'response': sid.split("\r\n")})
                    if (
                            command == 'setPortProfiles' or command == 'Set Port Profiles' or command == 'profile adsl set' or command == 'change lineprofile port'):
                        return JsonResponse({'current_userProfile': userProfile, 'response': sid})

                # ---------------------------------------------------zyxel---------------------------------------------------
                elif (dslam_obj.dslam_type_id == 1):
                    description = 'Run Command {0} on DSLAM {1}'.format(
                        command,
                        dslam_obj.name,
                    )
                    res = utility.dslam_port_run_command(dslam_obj.pk, 'port Info', params)
                    if (command == 'show linerate' or command == 'showPort' or command == 'showport'):
                        command = 'show linerate'
                    elif (command == 'profile adsl show' or command == 'showProfiles'):
                        command = 'profile adsl show'
                    elif (command == 'open port' or command == 'port enable'):
                        command = 'port enable'
                    elif (command == 'close port' or command == 'port disable'):
                        command = 'port disable'
                    elif (command == 'show mac slot port' or command == 'showmacslotport'):
                        command = 'show mac slot port'
                    elif (command == 'show port with mac' or command == 'show port mac'):
                        command = 'show port with mac'

                    if not command:
                        return Response({'result': 'Command does not exits'}, status=status.HTTP_400_BAD_REQUEST)
                    params["username"] = user.username
                    if (
                            command == 'setPortProfiles' or command == 'Set Port Profiles' or command == 'profile adsl set' or command == 'setProfiles'):
                        command = 'profile adsl set'

                    result = utility.dslam_port_run_command(dslam_obj.pk, command, params)
                    if result:
                        if 'Busy' in result:
                            return Response({'result': result}, status=status.HTTP_400_BAD_REQUEST)
                        else:
                            description = 'Run Command {0} on DSLAM {1}'.format(
                                command,
                                dslam_obj.name)

                            add_audit_log(request, 'DSLAMCommand', None, 'Run Command On DSLAM Port', description)
                    if (command == "show linerate"):
                        try:
                            sp = showPort()
                            result = result['result'].split("\\r\\n")
                            sp.dslamNameAndcammandName = result[1]
                            sp.slotPort = result[2].split(",")[0].split("=")[1]
                            sp.link = result[3].replace(" ", "").split("=")[1]
                            payloadrate = result[6].split("=")[1]
                            sp.payloadrate = [int(s) for s in payloadrate.split() if s.isdigit()]
                            actualrate = result[7].split("=")[1]
                            sp.actualrate = [int(c) for c in actualrate.split() if c.isdigit()]
                            attainablerate = result[8].split("=")[1]
                            sp.attainablerate = [int(c) for c in attainablerate.split() if c.isdigit()]
                            noisemargin = result[9].split("=")[1]
                            sp.noisemargin = re.findall(r"[-+]?\d*\.\d+|\d+", noisemargin)
                            attenuation = result[10].split("=")[1]
                            sp.attenuation = re.findall(r"[-+]?\d*\.\d+|\d+", attenuation)
                            return JsonResponse({'current_userProfile': res['result'][7].split(':')[1],
                                                 'dslamName/cammandName': sp.dslamNameAndcammandName,
                                                 'slot/Port': sp.slotPort,
                                                 'link': sp.link,
                                                 'payloadrateUp': sp.payloadrate[0],
                                                 'payloadrateDown': sp.payloadrate[1],
                                                 'actualrateUp': sp.actualrate[0],
                                                 'actualrateDown': sp.actualrate[1],
                                                 'attainablerateUp': sp.attainablerate[0],
                                                 'attainablerateDown': sp.attainablerate[1],
                                                 'noisemarginUp': sp.noisemargin[0],
                                                 'noisemarginDown': sp.noisemargin[1],
                                                 'attenuationUp': sp.attenuation[0],
                                                 'attenuationDown': sp.attenuation[1],
                                                 }, status=status.HTTP_200_OK)
                        except Exception as ex:
                            try:
                                return JsonResponse({'current_userProfile': res['result'][6].split(':')[1],
                                                     'dslamName/cammandName': sp.dslamNameAndcammandName,
                                                     'slot/Port': sp.slotPort,
                                                     'link': sp.link,
                                                     'payloadrateUp': sp.payloadrate[0],
                                                     'payloadrateDown': sp.payloadrate[1],
                                                     'actualrateUp': sp.actualrate[0],
                                                     'actualrateDown': sp.actualrate[1],
                                                     'attainablerateUp': sp.attainablerate[0],
                                                     'attainablerateDown': sp.attainablerate[1],
                                                     'noisemarginUp': sp.noisemargin[0],
                                                     'noisemarginDown': sp.noisemargin[1],
                                                     'attenuationUp': sp.attenuation[0],
                                                     'attenuationDown': sp.attenuation[1],
                                                     }, status=status.HTTP_200_OK)

                            except Exception as ex:
                                try:
                                    return JsonResponse(
                                        {'current_userProfile': res['result'][6].split(':')[1], 'response': result})
                                except Exception as ex:
                                    return JsonResponse({'current_userProfile': res, 'response': result})

                    elif (command == "show lineinfo"):
                        return JsonResponse(
                            {'current_userProfile': userProfile, 'response': result['result'].split("\r\n")})
                    elif (command == "selt"):
                        return JsonResponse({'current_userProfile': userProfile, 'response': result})
                    elif (command == 'profile adsl show'):
                        response = result['result']
                        # my_list = []
                        # for item in response:
                        # my_list.append(item['name'])
                        return Response({'response': response})
                    elif (
                            command == 'port enable' or command == 'port disable' or command == 'port pvc show' or command == 'get config'):
                        return JsonResponse({'response': result})
                    elif (command == "show mac slot port"):
                        portPvc = utility.dslam_port_run_command(dslam_obj.pk, 'port pvc show', params)
                        res = utility.dslam_port_run_command(dslam_obj.pk, 'show linerate', params)
                        payload = res['result'].split('\r\n')[6].split("=")[1]
                        payloadrate = [int(s) for s in payload.split() if s.isdigit()]
                        return JsonResponse(
                            {'Port PVC': portPvc, 'response': result['result'], 'payloadrateUp': payloadrate[0],
                             'payloadrateDown': payloadrate[1]})
                    elif (command == "show linestat port"):
                        return JsonResponse({'link status': result['result'].split('\r\n')[5].split()[4],
                                             'up-time':
                                                 [x.strip(' ') for x in result['result'].split('\r\n')][5].split()[6]})
                    elif (command == 'show port with mac' or command == 'show port mac'):
                        return JsonResponse({'response': result['result']})

                        slot = result['result'][5].split()[2].replace('-', '')
                        port = result['result'][5].split()[3]
                        vid = result['result'][5].split()[0]
                        mac = result['result'][5].split()[1]
                        if (fiber.card == int(slot) and fiber.port == int(port)):
                            return JsonResponse({'Slot': slot, 'Port': port, 'Vid': vid, 'Mac': mac})
                        else:
                            s = get_user_info_from_ibs('2177901525')
                            return JsonResponse({'response': s})
                            url = 'http://5.202.129.160:2080/api/v1/dslamport/port_conflict_correction/'
                            data = "{'fqdn':'teh.tehn.abouzar.t.dsl.Z6000.01','old_fqdn':'teh.tehn.abouzar.t.dsl.Z6000.01','old_port':{'card_number':3,'port_number':10},'port':{'card_number':2,'port_number':4},'reseller':{'name':'Iran-mabna'},'old_reseller':{'name':'Iran-mabna'},'subscriber':{'username':''}}"
                            # return JsonResponse({'response':data})
                            fhresponse = requests.post(url, data=data, headers={"Content-Type": "application/json"})
                            sid = fhresponse.json()
                            return JsonResponse({'Slot': slot, 'Port': port, 'Vid': vid, 'Mac': mac,
                                                 'result': 'there is a port conflict.'})

                    else:
                        return JsonResponse({'response': result})

            except Exception as ex:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                # return Response({'result': result['result'],'Line': str(exc_tb.tb_lineno)}, status=status.HTTP_400_BAD_REQUEST)
                return JsonResponse({'result': str(
                    'an error occurred. please try again.{0}----{1}'.format(str(exc_tb.tb_lineno), str(ex)))})

        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # return JsonResponse({'result': str('an error occurred. please try again. {0}'.format(str(exc_tb.tb_lineno)))}, status=status.HTTP_202_ACCEPTED)
            try:

                mail_info = Mail()
                mail_info.from_addr = 'oss-notification@pishgaman.net'
                mail_info.to_addr = 'oss-problems@pishgaman.net'
                mail_info.msg_body = 'Error in RunCommandAPIViewin Line {0}.Error Is: {1}. fqdn = {2},card = {3}, port = {4}, command = {5}, subscriber = {6}. IP: {7}'.format(
                    str(exc_tb.tb_lineno), str(ex), request.data.get('fqdn'),
                    request.data.get('params').get('port_conditions').get('slot_number'),
                    request.data.get('params').get('port_conditions').get('port_number'), command, subscriber, ip)
                mail_info.msg_subject = 'Error in RunCommandAPIView'
                Mail.Send_Mail(mail_info)
                return JsonResponse(
                    {'result': str(
                        'an error occurred. please try again. {0}-----{1}'.format(str(exc_tb.tb_lineno), str(ex)))},
                    status=status.HTTP_202_ACCEPTED)

            except Exception as ex:
                return JsonResponse(
                    {'result': str('an error occurred. please try again. {0}'.format(str(exc_tb.tb_lineno)))},
                    status=status.HTTP_202_ACCEPTED)


class FreePortAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            identifier_key = CustomerPort.objects.get(username=data.get('username')).identifier_key
            if (identifier_key != ''):
                CustomerPort.objects.get(username=data.get('username')).delete()
                Status = MDFDSLAM.objects.get(identifier_key=identifier_key)
                Status.status = 'FREE'
                Status.save()
                return JsonResponse({'result': str(Status.status)})
            else:
                return JsonResponse({'result': str('ERROR')})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # return JsonResponse({'result': str(ex),'Line': str(exc_tb.tb_lineno)}, status=status.HTTP_202_ACCEPTED)
            add_audit_log(request, 'FreePort', None, 'FreePort', str(ex))
            return JsonResponse({'result': str('an error occurred. please try again')}, status=status.HTTP_202_ACCEPTED)


class ReservePortAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            identifier_key = CustomerPort.objects.get(username=data.get('username')).identifier_key
            if (identifier_key != ''):
                Status = MDFDSLAM.objects.get(identifier_key=identifier_key)
                Status.status = 'RESERVE'
                Status.save()
                return JsonResponse({'result': str(Status.status)})
            else:
                return JsonResponse({'result': str('ERROR')})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # return JsonResponse({'result': str(ex),'Line': str(exc_tb.tb_lineno)}, status=status.HTTP_202_ACCEPTED)
            add_audit_log(request, 'FreePort', None, 'FreePort', str(ex))
            return JsonResponse({'result': str('an error occurred. please try again')}, status=status.HTTP_202_ACCEPTED)


class RanjeNumberInquiryV2APIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            phone = data.get('phoneNumber')
            code = data.get('code')
            telcoId = data.get('telecom_center_id')
            freePorts = MDFDSLAM.objects.filter(status='FREE', telecom_center_id=telcoId)
            telecom_center = TelecomCenter.objects.get(id=89015)
            url = 'https://pishgaman.net/checknum.php?rnd=0.026210093973347837&thephone=%s&area=%s' % (phone, code)
            response = requests.get(url)
            return Response(
                {'result': response, 'Free Ports': freePorts.count(), 'telecom_center': str(telecom_center.name)},
                status=status.HTTP_200_OK)
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            add_audit_log(request, 'FreePort', None, 'FreePort', str(ex))
            # return JsonResponse({'result': str(ex),'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class RanjeNumberInquiryAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            phone = data.get('phoneNumber')
            code = data.get('code')
            username = code + phone
            url = 'http://5.202.129.88:2080/api/PhonePrefix/GetPrefixInfo?code=%s&phone=%s' % (code, phone)
            response = requests.get(url)
            res = response.json()
            telecom_center_id = re.findall(r"[-+]?\d*\.\d+|\d+", res.get('OssTelcoId'))[0]
            freePortsCount = MDFDSLAM.objects.filter(status='FREE', telecom_center_id=telecom_center_id).count()
            if (freePortsCount == 0):
                return JsonResponse(
                    {'Result': 'There is no free port in the this telecom center.', 'freePortsCount': freePortsCount},
                    status=status.HTTP_200_OK)
            if (res.get('ErrorCode')):
                return JsonResponse({'Result': res}, status=status.HTTP_200_OK)
            return JsonResponse({'Result': res,
                                 'freePortsCount': freePortsCount,
                                 }, status=status.HTTP_200_OK)
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            add_audit_log(request, 'FreePort', None, 'FreePort', str(ex))
            # return JsonResponse({'result': str(ex),'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class BukhtUpdateAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            phone = data.get('phoneNumber')
            code = data.get('code')
            username = code + phone
            if (username[:1] == '0'):
                username = username[1:11]
            new_Bukht = data.get('new_Bukht')
            identifier_key = CustomerPort.objects.get(username=username).identifier_key
            portInfo = MDFDSLAM.objects.get(identifier_key=identifier_key)
            if portInfo is None:
                return JsonResponse({'Result': 'Current Port Information Does Not Exist.'}, status=status.HTTP_200_OK)
            portInfo.row_number = new_Bukht.get('row_number')
            portInfo.floor_number = new_Bukht.get('floor_number')
            portInfo.connection_number = new_Bukht.get('connection_number')
            portInfo.slot_number = new_Bukht.get('slot_number')
            portInfo.port_number = new_Bukht.get('port_number')
            portInfo.save()
            return JsonResponse({'Result': serialize('json', [portInfo])}, status=status.HTTP_200_OK)
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            add_audit_log(request, 'BukhtUpdate', None, str(str(exc_tb.tb_lineno)), str(ex))
            # return JsonResponse({'result': str(ex),'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class GetPortHistoryAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            return JsonResponse({'result': str('OK')})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # add_audit_log(request, 'BukhtUpdate', None, str(str(exc_tb.tb_lineno)) , str(ex))
            # return JsonResponse({'result': str(ex),'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class GetAllFreePortsAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            username = data.get('username')
            if (username[:1] == '0'):
                username = username[1:11]
            identifier_key = CustomerPort.objects.get(username=username).identifier_key
            telecomCenterId = MDFDSLAM.objects.get(identifier_key=identifier_key).telecom_center_id
            freeports = MDFDSLAM.objects.filter(telecom_center_id=telecomCenterId, status='FREE')
            freeports_res = []
            for record in freeports:
                freeports_obj = dict(row_number=record.row_number,
                                     floor_number=record.floor_number,
                                     connection_number=record.connection_number,
                                     slot_number=record.slot_number,
                                     port_number=record.port_number,
                                     status=record.status,
                                     telecomCenterId=telecomCenterId)
                freeports_res.append(freeports_obj)
            return HttpResponse(json.dumps(freeports_res), content_type='application/json')
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # add_audit_log(request, 'BukhtUpdate', None, str(str(exc_tb.tb_lineno)) , str(ex))
            # return JsonResponse({'result': str(ex),'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class ChangeBukhtAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            username = data.get('username')
            if (username[:1] == '0'):
                username = username[1:11]
            oldToStatus = data.get('old_to_Status')
            newToStatus = data.get('new_to_Status')
            new_Bukht = data.get('new_Bukht')
            try:
                CustomerObj = CustomerPort.objects.get(username=username)
            except CustomerPort.DoesNotExist:
                return JsonResponse({'Result': 'User By this UserName Does Not Exist.'})
            bukhtInfo = MDFDSLAM.objects.get(identifier_key=CustomerObj.identifier_key)
            bukhtInfo.status = oldToStatus
            bukhtInfo.save()

            try:
                newbukhtInfo = MDFDSLAM.objects.get(row_number=new_Bukht.get('row_number'),
                                                    connection_number=new_Bukht.get('connection_number'),
                                                    floor_number=new_Bukht.get('floor_number'),
                                                    slot_number=new_Bukht.get('slot_number'),
                                                    port_number=new_Bukht.get('port_number'))

                newbukhtInfo.status = newToStatus
                newbukhtInfo.save()

            except MDFDSLAM.DoesNotExist:
                return JsonResponse({'Result': 'Current Port Information Does Not Exist.'})
            CustomerObj.identifier_key = newbukhtInfo.identifier_key
            CustomerObj.save()

            return JsonResponse({
                'result': 'New Bukht Is:row_number:{0}/connection_number:{1}/floor_number:{2}/slot_number:{3}/port_number:{4}'.format(
                    newbukhtInfo.row_number, newbukhtInfo.connection_number,
                    newbukhtInfo.floor_number, newbukhtInfo.slot_number, newbukhtInfo.port_number)})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # add_audit_log(request, 'BukhtUpdate', None, str(str(exc_tb.tb_lineno)) , str(ex))
            # return JsonResponse({'result': str(ex),'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class GetFreePortInfoAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            phone = data.get('phoneNumber')
            code = data.get('code')
            username = code + phone
            url = 'http://5.202.129.88:2080/api/PhonePrefix/GetPrefixInfo?code=%s&phone=%s' % (code, phone)
            response = requests.get(url)
            res = response.json()
            telecom_center_id = re.findall(r"[-+]?\d*\.\d+|\d+", res.get('OssTelcoId'))[0]
            freePortsCount = MDFDSLAM.objects.filter(status='FREE', telecom_center_id=telecom_center_id).count()
            if (freePortsCount == 0):
                return JsonResponse(
                    {'TelcoName': res.get('TelcoName'), 'Result': 'There is no free port in the this telecom center.',
                     'freePortsCount': freePortsCount}, status=status.HTTP_200_OK)
            freePortInfo = MDFDSLAM.objects.filter(status='FREE', telecom_center_id=telecom_center_id).first()
            dslam_fqdn = DSLAM.objects.get(id=freePortInfo.dslam_id).fqdn
            if (res.get('ErrorCode')):
                return JsonResponse({'Result': res}, status=status.HTTP_200_OK)
            return JsonResponse({'dslam_fqdn': dslam_fqdn,
                                 'slot_number': freePortInfo.slot_number,
                                 'port_number': freePortInfo.port_number,
                                 'row_number': freePortInfo.row_number,
                                 'connection_number': freePortInfo.connection_number,
                                 'floor_number': freePortInfo.floor_number
                                 }, status=status.HTTP_200_OK)
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            add_audit_log(request, ' GetFreePortInfo', None, ' GetFreePortInfo', str(ex))
            # return JsonResponse({'result': str(ex),'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class PortAssignAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            username = data.get('username')
            dslamObj = DSLAM.objects.get(fqdn=data.get('fqdn'))
            mdfdslamObj = MDFDSLAM.objects.get(dslam_id=dslamObj.id, telecom_center_id=dslamObj.telecom_center.id,
                                               row_number=data.get('row_number'),
                                               connection_number=data.get('connection_number'),
                                               floor_number=data.get('floor_number'),
                                               slot_number=data.get('slot_number'), port_number=data.get('port_number'))
            identifier_key = mdfdslamObj.identifier_key
            if (username[:1] == '0'):
                username = username[1:11]
            customerObj = CustomerPort.objects.get(username=username)
            mdfdslamObj.status = 'Assigned'
            mdfdslamObj.save()
            customerObj.identifier_key = identifier_key
            customerObj.save()
            add_audit_log(request, '', None, 'Port Assign',
                          'Assign Port to {0} in {1} with {2} Was successful'.format(username, datetime.now(),
                                                                                     identifier_key))
            return JsonResponse({'Result': 'Assign Port to {0} in {1} with {2} was successful'.format(username,
                                                                                                      datetime.now(),
                                                                                                      identifier_key)},
                                status=status.HTTP_200_OK)
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            add_audit_log(request, 'Assign Port to {0} in {1} failed'.format(username, datetime.now()), None,
                          'Port Assign', str(ex))
            # return JsonResponse({'result': str(ex),'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class AddCustomerAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            if (CustomerPort.objects.filter(username=data.get('username')).exists()):
                return JsonResponse({'result': 'UserName {0} already exists.'.format(data.get('username'))})
            customer = CustomerPort()
            customer.lastname = data.get('lastname')
            customer.firstname = data.get('firstname')
            customer.username = data.get('username')
            if (customer.username[:1] == '0'):
                customer.username = customer.username[1:11]
            customer.tel = data.get('tel')
            customer.mobile = data.get('mobile')
            customer.national_code = data.get('national_code')
            customer.email = data.get('email')
            customer.identifier_key = data.get('mobile')
            customer.telecom_center_id = data.get('telecom_center_id')
            customer.save()
            return JsonResponse({'result': 'Customer {0} in {1} has been created successfully.'.format(
                data.get('username'), datetime.now())})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            add_audit_log(request,
                          'Add Customer {0} Process in {1} failed'.format(data.get('username'), datetime.now()), None,
                          'Add Customer', str(ex))
            return JsonResponse({'result': str(ex), 'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class FetchShaskamInquiryAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            usrPass = "pishgaman-software:RFFE@56844!"
            b64Val = base64.b64encode(usrPass)
            res = requests.post('https://bssapi.pishgaman.net/v1/login', headers={"Authorization": "Basic %s" % b64Val},
                                verify=False)
            response = res.json()
            token = response.get('data').get('token')
            res2 = requests.get('https://bssapi.pishgaman.net/v1/dsl/service/fetch-shaskam-inquiry/',
                                headers={"Authorization": "Bearer %s" % token}, verify=False)
            return JsonResponse({'result': res2.json()})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # add_audit_log(request, u'Add Customer {0} Process in {1} failed'.format(data.get('username'),datetime.now()), None, 'Add Customer', str(ex))
            # return JsonResponse({'result': str(ex),'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class SetShaskamInquiryAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            inquiry_result = request.data.get('inquiry_result')
            dsl_service_id = request.data.get('dsl_service_id')
            description = request.data.get('description')
            # return JsonResponse({'result': description  })
            usrPass = "pishgaman-software:RFFE@56844!"
            b64Val = base64.b64encode(usrPass)
            res = requests.post('https://bssapi.pishgaman.net/v1/login', headers={"Authorization": "Basic %s" % b64Val},
                                verify=False)
            response = res.json()
            token = response.get('data').get('token')
            data = "{'dsl_service_id':'%s','inquiry_result':%s,'description':'%s'}" % (inquiry_result, 0, description)
            if (inquiry_result == 0):
                res2 = requests.post('https://bssapi.pishgaman.net/v1/dsl/service/set-shaskam-inquiry/',
                                     data={"dsl_service_id": dsl_service_id, "inquiry_result": inquiry_result,
                                           'description': description}, headers={"Authorization": "Bearer %s" % token},
                                     verify=False)
            else:
                res2 = requests.post('https://bssapi.pishgaman.net/v1/dsl/service/set-shaskam-inquiry/',
                                     data={"dsl_service_id": dsl_service_id, "inquiry_result": inquiry_result},
                                     headers={"Authorization": "Bearer %s" % token}, verify=False)
            return JsonResponse({'result': res2.json()})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            add_audit_log(request,
                          'Set Shaskam Inquiry for Dsl ServiceId {0} in {1} was Failed,'.format(dsl_service_id,
                                                                                                datetime.now()), None,
                          'SetShaskamInquiry', str(ex))
            # return JsonResponse({'result': str(ex),'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class DSLAMPortRunCommand2View(views.APIView):

    def post(self, request, format=None):
        user = request.user
        data = request.data
        command = data.get('command', None)
        dslam_id = data.get('dslam_id', None)
        params = data.get('params', None)
        try:
            dslam_obj = DSLAM.objects.get(id=dslam_id)

        except Exception as ex:
            return JsonResponse({'result': str(ex)}, status=status.HTTP_400_BAD_REQUEST)
        logging.info(command)
        if not command:
            return Response({'result': 'Command does not exits'}, status=status.HTTP_400_BAD_REQUEST)
        params["username"] = user.username
        logging.info(params)
        if (dslam_obj.dslam_type_id == 4 or dslam_obj.dslam_type_id == 3 or dslam_obj.dslam_type_id == 5):
            fiber = DslanInfo()
            if (dslam_obj.dslam_type_id == 4):
                fiber.type = 'Fiberhome'
            if (dslam_obj.dslam_type_id == 3):
                fiber.type = 'FiberhomeAN3300'
            if (dslam_obj.dslam_type_id == 5):
                fiber.type = 'FiberhomeAN5006'
            fiber.dslam = dslam_obj.ip
            fiber.telnetPort = 23
            fiber.userName = dslam_obj.telnet_username
            fiber.password = dslam_obj.telnet_password
            fiber.access = dslam_obj.access_name
            fiber.card = request.data.get('params').get('port_conditions').get('slot_number')
            fiber.port = request.data.get('params').get('port_conditions').get('port_number')
            fiber.terminalDelay = '300'
            fiber.requestTimeOut = '1500'
            url = 'http://5.202.129.88:9096/api/Telnet/telnet'
            data = "{'type':'%s','dslam':'%s','telnetPort':'23','userName':'%s','password':'%s','access':'%s','card':'%s','port':'%s','command':'%s','terminalDelay':'600','requestTimeOut':'%s'}" % (
                fiber.type, fiber.dslam, fiber.userName, fiber.password, fiber.access, fiber.card, fiber.port, command,
                fiber.requestTimeOut)
            # return JsonResponse({'result': data})
            fhresponse = requests.post(url, data=data, headers={"Content-Type": "application/json"})
            sid = fhresponse.json()
            res = sid
            return JsonResponse({'result': res, 'Date': JalaliDatetime(datetime.now()).strftime("%Y-%m-%d %H:%M:%S"),
                                 'command': command})


class GetProvincesAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            provinces = City.objects.filter(parent_id=None)
            provinces_json_data = serialize('json', provinces)
            return HttpResponse(provinces_json_data, content_type="application/json")
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # add_audit_log(request, u'Set Shaskam Inquiry for Dsl ServiceId {0} in {1} was Failed,'.format(dsl_service_id,datetime.now()), None, 'SetShaskamInquiry', str(ex))
            return JsonResponse({'result': str(ex), 'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class GetCitiesAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            provinceId = data.get('provinceId')
            cities = City.objects.filter(parent_id=provinceId)
            cities_json_data = serialize('json', cities)
            return HttpResponse(cities_json_data, content_type="application/json")
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # add_audit_log(request, u'Set Shaskam Inquiry for Dsl ServiceId {0} in {1} was Failed,'.format(dsl_service_id,datetime.now()), None, 'SetShaskamInquiry', str(ex))
            return JsonResponse({'result': str(ex), 'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class GetTelecomCentersAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            cityId = data.get('cityId')
            telecom_centers = TelecomCenter.objects.filter(city_id=cityId)
            telecom_centers_json_data = serialize('json', telecom_centers)
            return HttpResponse(telecom_centers_json_data, content_type="application/json")
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # add_audit_log(request, u'Set Shaskam Inquiry for Dsl ServiceId {0} in {1} was Failed,'.format(dsl_service_id,datetime.now()), None, 'SetShaskamInquiry', str(ex))
            return JsonResponse({'result': str(ex), 'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class GetDslamsAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            telecom_centerId = data.get('telecomCenterId')
            dslams = DSLAM.objects.filter(telecom_center_id=telecom_centerId)
            dslams_json_data = serialize('json', dslams)
            return HttpResponse(dslams_json_data, content_type="application/json")
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # add_audit_log(request, u'Set Shaskam Inquiry for Dsl ServiceId {0} in {1} was Failed,'.format(dsl_service_id,datetime.now()), None, 'SetShaskamInquiry', str(ex))
            return JsonResponse({'result': str(ex), 'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class port_condition():
    slot_number = 0
    port_number = 0
    connection_number = 0
    floor_number = 0
    row_number = 0


class param():
    command = ''
    type = ''
    is_queue = ''
    fqdn = ''
    dslam_number = 0
    slot_number = 0
    port_number = 0
    port_conditions = port_condition()


class GetUserPortInfoAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            username = data.get('username')
            command = data.get('command')
            if (username[:1] == '0'):
                username = username[1:11]
            identifier_key = CustomerPort.objects.get(username=username).identifier_key
            portInfo = MDFDSLAM.objects.get(identifier_key=identifier_key)
            dslam_obj = DSLAM.objects.get(id=portInfo.dslam_id)
            params = data.get('params')

            params = param()
            params.type = 'dslamport'
            params.is_queue = False
            params.fqdn = dslam_obj.fqdn
            params.dslam_id = dslam_obj.id
            params.command = command
            params.port_conditions = port_condition()
            params.port_conditions.slot_number = portInfo.slot_number
            params.port_conditions.port_number = portInfo.port_number
            # result = utility.dslam_port_run_command(dslam_obj.id, 'selt', params)

            # return JsonResponse({'result':params.port_conditions})
            params = json.dumps(params, default=lambda x: x.__dict__)

            # result = utility.dslam_port_run_command(dslam_obj.id, 'selt', params)
            # return JsonResponse({'params': params })

            return HttpResponse(params, content_type="application/json")
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # add_audit_log(request, u'Set Shaskam Inquiry for Dsl ServiceId {0} in {1} was Failed,'.format(dsl_service_id,datetime.now()), None, 'SetShaskamInquiry', str(ex))
            return JsonResponse({'result': str(ex), 'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class GetCommandInfoSnmp(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            command = data.get('command')
            dslam_id = data.get('dslam_id')
            # params = data.get('params')
            params = json.dumps(data, default=lambda x: x.__dict__)

            # result = utility.dslam_port_run_command(dslam_id, 'selt', data)
            # return JsonResponse({'params': params })

            result = utility.dslam_port_run_command(dslam_id, command, data)
            if (command == 'selt'):
                return Response({'response': result[0].get('loopEstimateLength')})
            elif (command == 'show linerate'):
                userProfile = DSLAMPort.objects.get(port_number=request.data.get('port_conditions').get('port_number'),
                                                    slot_number=request.data.get('port_conditions').get('slot_number'),
                                                    dslam_id=dslam_id).line_profile
                return Response({'response': result['result'].split('\r\n'), 'userProfile': userProfile})
            elif (command == 'profile adsl show'):
                response = result['result']
                my_list = []
                for item in response:
                    my_list.append(item['name'])
                return Response({'response': my_list})

        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # add_audit_log(request, u'Set Shaskam Inquiry for Dsl ServiceId {0} in {1} was Failed,'.format(dsl_service_id,datetime.now()), None, 'SetShaskamInquiry', str(ex))
            return JsonResponse({'result': str(ex), 'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class GetCommandInfoTelnet(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            command = 'selt'
            data = request.data
            username = data.get('username')
            if (username[:1] == '0'):
                username = username[1:11]
            identifier_key = CustomerPort.objects.get(username=username).identifier_key
            portInfo = MDFDSLAM.objects.get(identifier_key=identifier_key)
            dslam_obj = DSLAM.objects.get(id=portInfo.dslam_id)
            fiber = DslanInfo()
            if (dslam_obj.dslam_type_id == 5):
                fiber.type = 'FiberhomeAN5006'
            else:
                fiber.type = 'ZyXEL'
            fiber.dslam = dslam_obj.ip
            fiber.telnetPort = 23
            fiber.userName = dslam_obj.telnet_username
            fiber.password = dslam_obj.telnet_password
            fiber.access = dslam_obj.access_name
            fiber.card = portInfo.slot_number
            fiber.port = portInfo.port_number
            fiber.terminalDelay = '300'
            fiber.requestTimeOut = '1500'
            url = 'http://5.202.129.88:9096/api/Telnet/telnet'
            data = "{'type':'%s','dslam':'%s','telnetPort':'23','userName':'%s','password':'%s','access':'%s','card':'%s','port':'%s','command':'%s','terminalDelay':'600','requestTimeOut':'%s'}" % (
                fiber.type, fiber.dslam, fiber.userName, fiber.password, fiber.access, fiber.card, fiber.port, command,
                fiber.requestTimeOut)
            # return JsonResponse({'result': data})
            fhresponse = requests.post(url, data=data, headers={"Content-Type": "application/json"})
            sid = fhresponse.json()
            res = sid.split("\n\r")
            return JsonResponse({'result': res, 'Date': JalaliDatetime(datetime.now()).strftime("%Y-%m-%d %H:%M:%S"),
                                 'command': command})

        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # add_audit_log(request, u'Set Shaskam Inquiry for Dsl ServiceId {0} in {1} was Failed,'.format(dsl_service_id,datetime.now()), None, 'SetShaskamInquiry', str(ex))
            return JsonResponse({'result': str(ex), 'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class GetBukhtInfoAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            queryset = City.objects.filter(parent_id=49)
            cityId = 1830
            telecomCenterId = 89682
            dslamId = 74
            if cityId:
                telecom_centers = TelecomCenter.objects.filter(city_id=cityId)
            if telecomCenterId:
                dslam_obj = DSLAM.objects.filter(telecom_center_id=telecomCenterId)
            # mdfdslamObj = MDFDSLAM.objects.filter(telecom_center_id = telecom_center_id , dslam_id = dslamId)
            json_data = serialize('json', queryset)
            return HttpResponse(json_data, content_type="application/json")
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # add_audit_log(request, u'Set Shaskam Inquiry for Dsl ServiceId {0} in {1} was Failed,'.format(dsl_service_id,datetime.now()), None, 'SetShaskamInquiry', str(ex))
            return JsonResponse({'result': str(ex), 'Line': str(exc_tb.tb_lineno)})
            return JsonResponse({'result': str('an error occurred. please try again')})


class GetPortInfoByUserNameAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            username = data.get('username')
            if (username[:1] == '0'):
                username = username[1:11]
            identifier_key = CustomerPort.objects.get(username=username).identifier_key
            portInfo = MDFDSLAM.objects.get(identifier_key=identifier_key)
            dslam_obj = DSLAM.objects.get(id=portInfo.dslam_id)
            params = param()
            params.fqdn = dslam_obj.fqdn
            dslam_number = dslam_obj.fqdn[-1:]
            params.dslam_number = dslam_number
            params.port_conditions = port_condition()
            params.port_conditions.slot_number = portInfo.slot_number
            params.port_conditions.port_number = portInfo.port_number
            params.port_conditions.connection_number = portInfo.connection_number
            params.port_conditions.floor_number = portInfo.floor_number
            params.port_conditions.row_number = portInfo.row_number

            params = json.dumps(params, default=lambda x: x.__dict__)
            return HttpResponse(params, content_type="application/json")
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # add_audit_log(request, u'Set Shaskam Inquiry for Dsl ServiceId {0} in {1} was Failed,'.format(dsl_service_id,datetime.now()), None, 'SetShaskamInquiry', str(ex))
            # return JsonResponse({'result': str(ex),'Line': str(exc_tb.tb_lineno)},status=status.HTTP_400_BAD_REQUEST)
            return JsonResponse(
                {'result': str('an error occurred. please try again.{0}'.format(str(exc_tb.tb_lineno)))},
                status=status.HTTP_400_BAD_REQUEST)


class RegisterPortByResellerIdAPIView(views.APIView):
    """
    "port":{
        "telecom_center_name": "string",
        "fqdn": "string",
        "card_number": "string",
        "port_number": "string"
    },
    "status": "RESELLER",
    "identifier_key":"string",
    "reseller": {
          "name": "string",
    },
    "subscriber":{
          "username": "string",
    }
    """

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        user = request.user
        data = request.data
        print()
        data
        reseller_data = data.get('reseller')
        customer_data = data.get('subscriber')
        mdf_status = data.get('status')
        sid = ''
        PVC = ''
        pvc = ''
        if not mdf_status:
            mdf_status = 'BUSY'
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        identifier_key = data.get('identifier_key')
        if not identifier_key:
            identifier_key = str(time.time())
        port_data = data.get('port')
        try:
            fqdn = port_data.get('fqdn')
            if ('.z6' in fqdn):
                fqdn = fqdn.replace('.z6', '.Z6')

            # dslam_obj = DSLAM.objects.get(name=port_data.get('dslam_name'), telecom_center__name=port_data.get('telecom_center_name'))
            try:
                dslam_obj = DSLAM.objects.get(fqdn=fqdn)
                # return JsonResponse({'Result': serialize('json',[dslam_obj])}, status=status.HTTP_200_OK)
                telecomId = dslam_obj.telecom_center_id
                cityId = TelecomCenter.objects.get(id=telecomId).city_id
            except ObjectDoesNotExist as ex:
                try:
                    if ('.Z6' in fqdn):
                        fqdn = fqdn.replace('.Z6', '.z6')
                        dslam_obj = DSLAM.objects.get(fqdn=fqdn)
                        # return JsonResponse({'Result': dslam_obj.fqdn}, status=status.HTTP_200_OK)
                        telecomId = dslam_obj.telecom_center_id
                        cityId = TelecomCenter.objects.get(id=telecomId).city_id

                except ObjectDoesNotExist as ex:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    mail_info = Mail()
                    mail_info.from_addr = 'oss-notification@pishgaman.net'
                    mail_info.to_addr = 'oss-problems@pishgaman.net'
                    mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                        str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'),
                        port_data.get('card_number'), port_data.get('port_number'), ip)
                    mail_info.msg_subject = 'Error in RegisterPortAPIView'
                    Mail.Send_Mail(mail_info)
                    return JsonResponse({'Result': ''}, status=status.HTTP_200_OK)

            telecom_mdf_obj = TelecomCenterMDF.objects.filter(telecom_center_id=dslam_obj.telecom_center.id)
            if telecom_mdf_obj:
                telecom_mdf_obj = telecom_mdf_obj.first()

            mdf_dslam_obj, mdf_dslam__updated = MDFDSLAM.objects.update_or_create(
                telecom_center_id=dslam_obj.telecom_center.id, telecom_center_mdf_id=telecom_mdf_obj.id,
                #### Check this whole line
                row_number=0, floor_number=0, connection_number=0,  ##### Check this whole line
                dslam_id=dslam_obj.id, slot_number=port_data.get('card_number'),
                port_number=port_data.get('port_number'),
                defaults={'status': mdf_status, 'identifier_key': identifier_key})
            # if mdf_dslam.status != 'FREE':
            #    return JsonResponse(
            #            {'result': 'port status is {0}'.format(mdf_dslam.status), 'id': -1}
            #            )
            # else:
            #    mdf_dslam.status = 'RESELLER'
            #    mdf_dslam.save()
            # identifier_key = mdf_dslam.identifier_key
        except ObjectDoesNotExist as ex:
            return JsonResponse({'result': str(ex), 'id': -1})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            try:
                return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(dslam_obj.telecom_center.id),
                                     'testRes': serialize('json', telecom_mdf_obj)})
            except ObjectDoesNotExist as ex:
                mail_info = Mail()
                mail_info.from_addr = 'oss-notification@pishgaman.net'
                mail_info.to_addr = 'oss-problems@pishgaman.net'
                mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                    str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'), port_data.get('card_number'),
                    port_data.get('port_number'), ip)
                mail_info.msg_subject = 'Error in RegisterPortAPIView'
                Mail.Send_Mail(mail_info)
                return JsonResponse({'Result': ''}, status=status.HTTP_200_OK)
        try:
            reseller_obj, reseller_created = Reseller.objects.get_or_create(id=reseller_data.get('resellerId'))
            print()
            'reseller', reseller_obj

            customer_obj, customer_updated = CustomerPort.objects.update_or_create(
                username=customer_data.get('username'),
                defaults={'identifier_key': identifier_key, 'telecom_center_id': mdf_dslam_obj.telecom_center_id}
            )

            rp = ResellerPort()
            rp.identifier_key = identifier_key
            rp.status = 'ENABLE'
            rp.dslam_id = dslam_obj.id
            rp.dslam_slot = port_data.get('card_number')
            rp.dslam_port = port_data.get('port_number')
            rp.reseller = reseller_obj
            rp.telecom_center_id = mdf_dslam_obj.telecom_center_id
            rp.save()

            vlan_objs = Vlan.objects.filter(reseller=reseller_obj)
            print()
            'vlan_objs ->', vlan_objs
            # return JsonResponse({'result':  reseller_obj.vci, 'vpi' : reseller_obj.vpi})
            port_indexes = [{'slot_number': port_data.get('card_number'), 'port_number': port_data.get('port_number')}]
            pishParams = {
                "type": "dslamport",
                "is_queue": False,
                "vlan_id": 3900,
                "vlan_name": 'pte',
                "dslam_id": dslam_obj.id,
                "port_indexes": port_indexes,
                "username": customer_data.get('username'),
                "vpi": 0,
                "vci": 35,
            }

            params = {
                "type": "dslamport",
                "is_queue": False,
                "vlan_id": vlan_objs[0].vlan_id,
                "vlan_name": vlan_objs[0].vlan_name,
                "dslam_id": dslam_obj.id,
                "port_indexes": port_indexes,
                "username": customer_data.get('username'),
                "vpi": reseller_obj.vpi,
                "vci": reseller_obj.vci,
            }
            # return JsonResponse({'res' : params })
            result = ''

            try:
                if (cityId == 1637):
                    vpi = Reseller.objects.get(name=reseller_data['name']).vpi
                    vci = Reseller.objects.get(name=reseller_data['name']).vci
                    fromaddr = "oss-notification@pishgaman.net"
                    toaddr = "rt-network-access@pishgaman.net"
                    msg = MIMEMultipart()
                    msg['From'] = fromaddr
                    msg['To'] = toaddr
                    msg['Subject'] = "OSS Shabdiz Dslams"
                    body = 'Command: add to vlan for Shabdiz Dslams, IP: {0}, fqdn: {1} , Card: {2} , Port: {3} , VlanId: {4},Reseller: {5},Vpi: {6},Vci: {7}'.format(
                        dslam_obj.ip, dslam_obj.fqdn, port_data.get('card_number'), port_data.get('port_number'),
                        vlan_objs[0].vlan_id, reseller_data['name'], vpi, vci)
                    msg.attach(MIMEText(body, 'plain'))
                    server = smtplib.SMTP('mail.pishgaman.net', 25)
                    server.login("oss-notification@pishgaman.net", "Os123456")
                    text = msg.as_string()
                    server.sendmail(fromaddr, toaddr, text)
                    return JsonResponse({'result': str('Email Sent To rt-network-access!!!')},
                                        status=status.HTTP_202_ACCEPTED)
                elif (dslam_obj.dslam_type_id == 4):
                    try:
                        vlanName = str(reseller_obj).split('-')[1]
                        url = 'http://5.202.129.88:9096/api/Telnet/telnet'
                        data = "{'type':'Fiberhome','dslam':'%s','telnetPort':'23','userName':'%s','password':'%s','access':'%s','sourceVlanId':'%s','vlanName':'%s','vlanId':'%s','untaggedPortList':'%s','vpiVci':'%s','card':'%s','port':'%s','command':'%s','terminalDelay':'600','requestTimeOut':'1500'}" % (
                            dslam_obj.ip, dslam_obj.telnet_username, dslam_obj.telnet_password, dslam_obj.access_name,
                            '3900', vlanName, vlan_objs[0].vlan_id,
                            '0-{0}-{1}'.format(port_data.get('card_number'), port_data.get('port_number')),
                            '{0}/{1}'.format(Reseller.objects.get(id=reseller_data['resellerId']).vpi,
                                             Reseller.objects.get(id=reseller_data['resellerId']).vci),
                            port_data.get('card_number'), port_data.get('port_number'), 'addToVlan')
                        # return JsonResponse({'result':data })
                        fhresponse = requests.post(url, data=data, headers={"Content-Type": "application/json"})
                        sid = fhresponse.json()

                    except Exception as ex:
                        exc_type, exc_obj, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        mail_info = Mail()
                        mail_info.from_addr = 'oss-notification@pishgaman.net'
                        mail_info.to_addr = 'oss-problems@pishgaman.net'
                        mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                            str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'),
                            port_data.get('card_number'), port_data.get('port_number'), ip)
                        mail_info.msg_subject = 'Error in RegisterPortAPIView'
                        Mail.Send_Mail(mail_info)
                        return JsonResponse(
                            {'result': str('an error occurred. please try again. {0}'.format(str(exc_tb.tb_lineno)))})

                else:
                    res = utility.dslam_port_run_command(dslam_obj.id, 'delete from vlan', pishParams)
                    result = utility.dslam_port_run_command(dslam_obj.id, 'add to vlan', params)
                    PVC = utility.dslam_port_run_command(dslam_obj.id, 'port pvc show', params)

                    # result2 = utility.dslam_port_run_command(dslam_obj.id, 'add to vlan', params)

            except Exception as ex:
                if (dslam_obj.dslam_type_id == 5 or dslam_obj.dslam_type_id == 3 or dslam_obj.dslam_type_id == 7):
                    vpi = Reseller.objects.get(name=reseller_data['name']).vpi
                    vci = Reseller.objects.get(name=reseller_data['name']).vci
                    fromaddr = "oss-notification@pishgaman.net"
                    toaddr = "rt-network-access@pishgaman.net"
                    msg = MIMEMultipart()
                    msg['From'] = fromaddr
                    msg['To'] = toaddr
                    msg['Subject'] = "OSS Problem"
                    body = 'Command: add to vlan, IP: {0}, fqdn: {1} , Card: {2} , Port: {3} , VlanId: {4},Reseller: {5},Vpi: {6},Vci: {7}'.format(
                        dslam_obj.ip, dslam_obj.fqdn, port_data.get('card_number'), port_data.get('port_number'),
                        vlan_objs[0].vlan_id, reseller_data['name'], vpi, vci)
                    msg.attach(MIMEText(body, 'plain'))
                    server = smtplib.SMTP('mail.pishgaman.net', 25)
                    server.login("oss-notification@pishgaman.net", "Os123456")
                    text = msg.as_string()
                    server.sendmail(fromaddr, toaddr, text)
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    # return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})
                    return JsonResponse(
                        {'result': str('An Error Ocurred for add to vlan Command. Email Sent To rt-network-access!!!')},
                        status=status.HTTP_202_ACCEPTED)

                else:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    mail_info = Mail()
                    mail_info.from_addr = 'oss-notification@pishgaman.net'
                    mail_info.to_addr = 'oss-problems@pishgaman.net'
                    mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                        str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'),
                        port_data.get('card_number'), port_data.get('port_number'), ip)
                    mail_info.msg_subject = 'Error in RegisterPortAPIView'
                    Mail.Send_Mail(mail_info)
                    return JsonResponse(
                        {'result': str('an error occurred. please try again. {0}'.format(str(exc_tb.tb_lineno)))})

            if vlan_objs.count() > 0:
                try:
                    port_obj = DSLAMPort.objects.get(dslam=dslam_obj, slot_number=port_data.get('card_number'),
                                                     port_number=port_data.get('port_number'))
                except ObjectDoesNotExist as ex:
                    dslam_port = DSLAMPort()
                    dslam_port.dslam = dslam_obj
                    dslam_port.created_at = datetime.now()
                    dslam_port.updated_at = datetime.now()
                    dslam_port.slot_number = port_data.get('card_number')
                    dslam_port.port_number = port_data.get('port_number')
                    dslam_port.port_index = '{0}{1}'.format(port_data.get('card_number'), port_data.get('port_number'))
                    dslam_port.port_name = 'adsl{0}-{1}'.format(port_data.get('card_number'),
                                                                port_data.get('port_number'))
                    dslam_port.admin_status = 'UNLOCK'
                    dslam_port.oper_status = 'NO-SYNC'
                    dslam_port.line_profile = '768/10240/IR'
                    dslam_port.selt_value = ''
                    dslam_port.uptime = '18:55:20'
                    dslam_port.upstream_snr = '192'
                    dslam_port.downstream_snr = '230'
                    dslam_port.upstream_attenuation = '69'
                    dslam_port.downstream_attenuation = '188'
                    dslam_port.upstream_attainable_rate = '2434'
                    dslam_port.downstream_attainable_rate = '157'
                    dslam_port.upstream_tx_rate = '93'
                    dslam_port.downstream_tx_rate = '1250'
                    dslam_port.upstream_snr_flag = 'good'
                    dslam_port.downstream_snr_flag = 'excellent'
                    dslam_port.upstream_attenuation_flag = 'outstanding'
                    dslam_port.downstream_attenuation_flag = 'outstanding'
                    dslam_port.vpi = 0
                    dslam_port.vci = 35
                    dslam_port.save()
                    port_obj = DSLAMPort.objects.get(dslam=dslam_obj, slot_number=port_data.get('card_number'),
                                                     port_number=port_data.get('port_number'))
                port_vlan_obj = DSLAMPortVlan()
                port_vlan_obj.vlan = vlan_objs.first()
                port_vlan_obj = port_obj
                port_vlan_obj.save()

            return JsonResponse({'result': '', 'PVC': PVC, 'id': 201, 'res': sid}, status=status.HTTP_201_CREATED)
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            mail_info = Mail()
            mail_info.from_addr = 'oss-notification@pishgaman.net'
            mail_info.to_addr = 'oss-problems@pishgaman.net'
            mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'), port_data.get('card_number'),
                port_data.get('port_number'), ip)
            mail_info.msg_subject = 'Error in RegisterPortAPIView'
            Mail.Send_Mail(mail_info)

            # return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})
            return JsonResponse(
                {'result': str('an error occurred. please try again. {0}'.format(str(exc_tb.tb_lineno)))},
                status=status.HTTP_202_ACCEPTED)


class BitstreamFreePortAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            identifier_key = CustomerPort.objects.get(username=data.get('username')).identifier_key
            if (identifier_key != ''):
                CustomerPort.objects.get(username=data.get('username')).delete()
                Status = MDFDSLAM.objects.get(identifier_key=identifier_key)
                Status.status = 'FREE'
                Status.save()
                result = utility.dslam_port_run_command(dslam_obj.id, 'add to vlan', params)
            return JsonResponse({'result': 'OK'}, status=status.HTTP_201_CREATED)
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return JsonResponse(
                {'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno), 'ip': str(dslam_obj.ip)})

            return JsonResponse({'result': str('an error occurred. please try again')}, status=status.HTTP_202_ACCEPTED)


class port_condition2():
    slot_number = 0
    port_number = 0


class param2():
    command = ''
    type = ''
    is_queue = ''
    fqdn = ''
    slot_number = 0
    port_number = 0
    port_conditions = port_condition2()


class GetDslamBackupAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        # dslam_objs = DSLAM.objects.annotate(Count('fqdn')).order_by('fqdn')[:2]
        params = param()
        dslam_objs = DSLAM.objects.filter(dslam_type_id=1)
        zyxel_path = DSLAM_BACKUP_PATH + 'zyxel/'
        if not os.path.exists(zyxel_path):
            os.makedirs(zyxel_path)
        try:
            for dslam_obj in dslam_objs:
                params = param2()
                params.type = 'dslamport'
                params.is_queue = False
                params.fqdn = dslam_obj.fqdn
                params.command = 'get config'
                params.port_conditions = port_condition2()
                params.port_conditions.slot_number = 1
                params.port_conditions.port_number = 1
                params = json.dumps(params, default=lambda x: x.__dict__)
                try:
                    result = utility.dslam_port_run_command(dslam_obj.pk, 'get config', json.loads(params))
                    print(result)
                    if result.get('status', None) == 200:
                        with open(zyxel_path + "{0}@{1}_{2}.txt".format(
                            dslam_obj.fqdn, dslam_obj.ip, str(datetime.today().strftime('%Y-%m-%d-%H:%M:%S'))),
                                  "w") as backup_file:
                            backup_file.write(result.get('result'))
                        return JsonResponse({'result': 'OK try'}, status=status.HTTP_201_CREATED)
                    else:
                        with open(
                                zyxel_path + 'Error#{0}@{1}_{2}.txt'.format(
                                    dslam_obj.fqdn, dslam_obj.ip,str(datetime.today().strftime('%Y-%m-%d-%H:%M:%S'))), 'w') as text_file:
                            text_file.write("Error_ %s" % result.get('result'))
                except Exception as ex:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    with open(
                            zyxel_path + 'Error#{0}@{1}_{2}.txt'.format(dslam_obj.fqdn, dslam_obj.ip,
                                                              str(datetime.today().strftime('%Y-%m-%d-%H:%M:%S'))), 'w') as text_file:
                        text_file.write("Error_ %s" % str(ex))
                    return JsonResponse({'result': 'except'}, status=status.HTTP_201_CREATED)

            return JsonResponse({'result': 'OK'}, status=status.HTTP_201_CREATED)

        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return JsonResponse(
                {'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno), 'ip': str(dslam_obj.ip)})


class File:
    file_name = ''
    file_date = ''


class ShowDslamBackups(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            date_array = []
            for i in range(0, 10):
                date_array.append(str(datetime.now().date() - timedelta(i)))
            dslam_id = request.data.get('id')
            dslam_obj = DSLAM.objects.get(id=dslam_id)
            dslam_type = dslam_obj.dslam_type_id
            if dslam_type == 1:
                path = DSLAM_BACKUP_PATH + 'zyxel/'
            elif dslam_type == 2:
                path = DSLAM_BACKUP_PATH + 'huawei/'
            elif dslam_type == 3:
                path = DSLAM_BACKUP_PATH + 'fiberhome_3300/'
            elif dslam_type == 5:
                path = DSLAM_BACKUP_PATH + 'fiberhome_5006/'
            elif dslam_type == 7:
                path = DSLAM_BACKUP_PATH + 'zyxel_1248/'
            elif dslam_type == 4:
                path = DSLAM_BACKUP_PATH + 'fiberhome_2200/'
            else:
                return None
            dslam_ip = dslam_obj.ip
            dslam_fqdn = dslam_obj.fqdn
            filenames = []
            total = []
            directory = path
            for filename in os.listdir(directory):
                file = File()
                if filename.__contains__(dslam_fqdn) or filename.__contains__(dslam_ip):
                    for item in date_array:
                        if item in filename:
                            file.file_name = filename
                            file.file_date = filename.split('_')[1].split('.')[0]
                            # filenames.append(file)
                            total.append(file)
            print(total)
            return JsonResponse({'response': json.dumps(total, default=lambda o: o.__dict__,
                                                        sort_keys=True, indent=4)})

        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            return JsonResponse({'row': str(ex) + "  // " + str(exc_tb.tb_lineno)})


class DownloadDslamBackupFileAPIView(views.APIView):

    def post(self, request, format=None):
        try:
            backup_file_name = request.data.get('backup_file_name')
            dslam_id = request.data.get('id')
            dslam_obj = DSLAM.objects.get(id=dslam_id)
            dslam_type_id = dslam_obj.dslam_type_id
            if dslam_type_id == 1:
                path = DSLAM_BACKUP_PATH + 'zyxel/'
            elif dslam_type_id == 2:
                path = DSLAM_BACKUP_PATH + 'huawei/'
            elif dslam_type_id == 3:
                path = DSLAM_BACKUP_PATH + 'fiberhome_3300/'
            elif dslam_type_id == 5:
                path = DSLAM_BACKUP_PATH + 'fiberhome_5006/'
            elif dslam_type_id == 7:
                path = DSLAM_BACKUP_PATH + 'zyxel_1248/'
            elif dslam_type_id == 4:
                path = DSLAM_BACKUP_PATH + 'fiberhome_2200/'
            else:
                return None
            directory = path + backup_file_name
            f = open(directory, "r")
            return JsonResponse({'response': f.read()})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            return JsonResponse({'row': str(ex) + "  // " + str(exc_tb.tb_lineno)})


class SetTimeAllDslamsAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        # dslam_objs = DSLAM.objects.annotate(Count('fqdn')).order_by('fqdn')[:2]
        params = param()
        dslam_objs = DSLAM.objects.filter(dslam_type_id=1)

        try:
            for dslam_obj in dslam_objs:
                telecom_center_id = DSLAM.objects.get(id=dslam_obj.id).telecom_center_id
                telecom_center = TelecomCenter.objects.get(id=telecom_center_id).name
                # city_id = TelecomCenter.objects.get(id = telecom_center_id).city_id
                # city_name = City.objects.get(id = city_id).name
                params = param2()
                params.type = 'dslamport'
                params.is_queue = False
                params.fqdn = dslam_obj.fqdn
                params.command = 'get config'
                params.port_conditions = port_condition2()
                params.port_conditions.slot_number = 1
                params.port_conditions.port_number = 1
                params = json.dumps(params, default=lambda x: x.__dict__)
                try:
                    result = utility.dslam_port_run_command(dslam_obj.pk, 'set time', json.loads(params))
                except Exception as ex:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    os.mkdir('{0}/{1}-{2}'.format('/opt/PortMan/backsErrors', telecom_center, dslam_obj.ip))
                    with open('{0}/{1}-{2}/{3}-{4}.txt'.format('/opt/PortMan/timeErrors', telecom_center, dslam_obj.ip,
                                                               telecom_center, dslam_obj.ip), 'w') as text_file:
                        text_file.write("Error: %s" % str(ex))

            return JsonResponse({'result': 'OK'}, status=status.HTTP_201_CREATED)

        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return JsonResponse(
                {'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno), 'ip': str(dslam_obj.ip)})
            return JsonResponse({'result': str('an error occurred. please try again')}, status=status.HTTP_202_ACCEPTED)


class SendMailAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            # mail_info= Mail()
            # mail_info.from_addr = 'oss-notification@pishgaman.net'
            # mail_info.to_addr = 'oss-problems@pishgaman.net'
            # mail_info.msg_body = 'Hello'
            # mail_info.msg_subject= 'Hello'
            # Mail.Send_Mail(mail_info)
            data = request.data
            fromaddr = "oss-notification@pishgaman.net"
            toaddr = data.get('toaddr')
            msg = MIMEMultipart()
            msg['From'] = fromaddr
            msg['To'] = toaddr
            msg['Subject'] = data.get('subject')
            body = data.get('body')
            msg.attach(MIMEText(body, 'plain'))
            server = smtplib.SMTP('mail.pishgaman.net', 25)
            server.login("oss-notification@pishgaman.net", "Os123456")
            text = msg.as_string()
            server.sendmail(fromaddr, toaddr, text)
            return JsonResponse({'result': 'Email Sent'})

        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})


class GetPortInfoByIdAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            portId = data.get('portId')
            portObj = DSLAMPort.objects.get(id=portId)
            card = portObj.slot_number
            port = portObj.port_number
            fqdn = DSLAM.objects.get(id=portObj.dslam_id).fqdn
            vpi2 = portObj.vpi
            vci2 = portObj.vci
            if (vpi2 and vci2):
                resellerName = Reseller.objects.get(vpi=vpi2, vci=vci2).name
                return JsonResponse({'resellerName': resellerName, 'fqdn': fqdn, 'card': card, 'port': port})
            else:
                return JsonResponse({'Error': 'Vpi Or Vci not set.'})

        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})


class FiberHomeGetCardAPIView(views.APIView):
    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, format=None):
        status_cards_class = FiberHomeGetCardStatusService(request.data)
        status_cards = status_cards_class.get_status_cards()
        return JsonResponse({'result': status_cards})


class FiberHomeGetPortAPIView(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, format=None):
        print('FiberHomeGetCardAPIView')
        data = request.data
        command = 'Show Card'
        dslam_id = data.get('dslam_id', None)
        dslamObj = DSLAM.objects.get(id=dslam_id)
        params = data.get('params', None)
        dslam_type = dslamObj.dslam_type_id
        ports_info = []

        try:
            result = utility.dslam_port_run_command(dslamObj.pk, command, params)
            if dslam_type == 3:  ############################## fiberhomeAN3300 ##############################
                if "There is one of the following problems:" in result['result'] or \
                        "contact with core-access department" in result['result'] or \
                        "DSLAM Password is wrong!" in result['result'] or \
                        "Card number is out of range" in result['result']:
                    return JsonResponse({'result': result['result']})
                number_of_ports = 0
                slot_number = params['port_conditions']['slot_number']

                
                for item in result['result']:
                    info = item.strip().split(':')
                    slot_number_holder = info[0].split()
                    print(info, slot_number, slot_number_holder)
                    
                    try :
                        is_slot_matched = slot_number_holder[1] == str(slot_number)
                    except:
                        is_slot_matched = slot_number_holder[0] == str(slot_number)

                    if is_slot_matched:
                        port_info = {}
                        port_info['port_number'] = info[1].split()[0]
                        if 'DATA' in info[-1] or 'Data' in info[1]:
                            port_info['port_status'] = "ON"
                        elif 'attribute.' in info[-1] or 'attribute.' in info[1]:
                            port_info['port_status'] = "has not this attribute."
                        else:
                            port_info['port_status'] = "OFF"
                        number_of_ports += 1
                        port_info['up_time'] = None
                        port_info['protocol'] = None
                        ports_info.append(port_info)
                if len(ports_info) == 0:
                    return JsonResponse({'result': "There are no active ports on this card."},
                                        status=status.HTTP_200_OK)

                return JsonResponse({'number of ports': number_of_ports, 'result': ports_info})

            elif dslam_type == 4:  ############################## fiberhomeAN2200 ##############################
                if "This card is not configured" in result['result']:
                    return JsonResponse({'result': result}['result'])
                if "No card is defined on this port" in result['result']:
                    return JsonResponse({'result': 'There are no active ports on this card'}, status=status.HTTP_200_OK)
                result = [val for val in result['result'] if re.search(r'.prf', val)]
                number_of_ports = len(result)

                for item in result:
                    port_info = {}
                    port_info['port_number'] = item.split()[0]
                    if 'UP' in item:
                        port_info['port_status'] = 'ON'
                    else:
                        port_info['port_status'] = 'OFF'
                    port_info['up_time'] = None
                    port_info['protocol'] = None
                    ports_info.append(port_info)
                return JsonResponse({'number of ports': number_of_ports, 'result': ports_info},
                                    status=status.HTTP_200_OK)

            elif dslam_type == 5:  ############################## fiberhomeAN5006 ##############################
                if "The Card number maybe unavailable or does not exist." in result['result'] or \
                        "core-access department." in result['result']:
                    return JsonResponse({'result': result['result']}, status=status.HTTP_200_OK)
                if "Card number is out of range." in result['result']:
                    return JsonResponse({'result': result['result']}, status=status.HTTP_200_OK)
                number_of_ports = len(result['result'])
                for item in result['result']:
                    port_info = {}
                    port_info['port_number'] = item.split()[1]
                    if 'SHOWTIME' in item:
                        port_info['port_status'] = 'ON'
                    else:
                        port_info['port_status'] = 'OFF'
                    port_info['up_time'] = None
                    port_info['protocol'] = None
                    ports_info.append(port_info)
                return JsonResponse({'number of ports': number_of_ports, 'result': ports_info},
                                    status=status.HTTP_200_OK)
            elif dslam_type == 1:  ############################## zyxel ##############################
                result = [item for item in result['result'] if re.search(r"^\s*\d+", item)]
                number_of_ports = len(result)
                if number_of_ports == 0:
                    return JsonResponse({'result': 'There are no active ports on this card.'},
                                        status=status.HTTP_200_OK)
                for item in result:
                    port_info = {}
                    info = re.sub('\s{2,}', " ", item)
                    info = info.split('-')[1].lstrip().split(' ', 5)
                    port_info['port_number'] = info[0]

                    if 'up' in info[1]:
                        port_info['port_status'] = 'ON'
                    else:
                        port_info['port_status'] = 'OFF'

                    port_info['up_time'] = info[5]
                    port_info['protocol'] = info[4]
                    ports_info.append(port_info)

                return JsonResponse({'number of ports': number_of_ports, 'result': ports_info},
                                    status=status.HTTP_200_OK)

            elif dslam_type == 7:  ############################## zyxel1248 ##############################

                result = [item.split() for item in result['result'] if re.search(r'^\s+\d+', item)]
                number_of_ports = len(result)
                for info in result:
                    port_info = {}
                    port_info['port_number'] = info[0]
                    if info[1] != '-':
                        port_info['port_status'] = 'ON'
                        port_info['up_time'] = info[-2]
                    else:
                        port_info['port_status'] = 'OFF'
                        port_info['up_time'] = ""
                    port_info['protocol'] = None
                    ports_info.append(port_info)
                return JsonResponse({'number of ports': number_of_ports, 'result': ports_info},
                                    status=status.HTTP_200_OK)

            elif dslam_type == 2:  ############################## huawei ##############################
                result = [item for item in result['result'] if re.search(r'^\s+\d+', item)]
                number_of_ports = len(result)
                if number_of_ports == 0:
                    return JsonResponse({'result': 'There are no active ports on this card'}, status=status.HTTP_200_OK)
                for info in result:
                    port_info = {}
                    port_info['port_number'] = info.split()[0]
                    if 'Activated' in info:
                        port_info['port_status'] = 'ON'
                    else:
                        port_info['port_status'] = 'OFF'
                    port_info['up_time'] = None
                    port_info['protocol'] = None
                    ports_info.append(port_info)
                return JsonResponse({'number of ports': number_of_ports, 'result': ports_info},
                                    status=status.HTTP_200_OK)

        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(dslam_type, str(exc_tb.tb_lineno), fname)
            return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)},
                                status=status.HTTP_500_INTERNAL_SERVER_ERROR)


###### Get PVC and VLAN API
class GetPVCVlanAPIView(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, format=None):
        print('Get PVC VLAN')
        data = request.data
        command = 'Show VLAN'
        dslam_id = data.get('dslam_id', None)
        dslamObj = DSLAM.objects.get(id=dslam_id)
        params = data.get('params', None)
        dslam_type = dslamObj.dslam_type_id
        pvc_vlan = []
        vlan_info = {}

        try:
            result = utility.dslam_port_run_command(dslamObj.pk, command, params)
            if dslam_type == 3:  ############################## fiberhomeAN3300 ##############################
                result = utility.dslam_port_run_command(dslamObj.pk, "Show All VLANs", params)
                if "information" not in result[0]:
                    return Response({'result': result})
                cart_port = [val.split() for val in result]
                flag = 0
                for item in reversed(cart_port):
                    if item[0] == str(params['port_conditions']['slot_number']) and item[1] == str(
                            params['port_conditions']['port_number']):
                        vlan_info['pvc_num'] = item[2]
                        flag = 1
                    if 'vlan' in item and flag == 1:
                        vlan_info['vlan_id'] = item[2].split(":")[1]
                        vlan_info['vlan_name'] = item[1].split(":")[1]
                        break
                result = utility.dslam_port_run_command(dslamObj.pk, "show pvc", params)
                if "pvc" not in str(result[0]):
                    return Response({'result': result})
                for val in result:
                    vlan_info.update(val)
                pvc_vlan.append(vlan_info)
                return Response({'result': pvc_vlan})
                # return Response({'result': result})

            elif dslam_type == 4:  ############################## fiberhomeAN2200 ##############################
                # vlan_info['vlan_id'] = result['VLAN ID']
                # vlan_info['vlan_name'] = result['Name']
                # result = utility.dslam_port_run_command(dslamObj.pk, "show pvc", params)
                # result = result[-1]
                # result = re.split("\s{2,}", result)
                # vlan_info['pvc1'] = result[4].strip()
                # vlan_info['pvc2'] = result[6].strip()
                # pvc_vlan.append(vlan_info)
                # return JsonResponse({'result': pvc_vlan})
                return JsonResponse({'result': result})

            elif dslam_type == 5:  ############################## fiberhomeAN5006 ##############################
                vlan_info.update(result)
                result = utility.dslam_port_run_command(dslamObj.pk, "show pvc", params)
                if "Profile name" not in result:
                    return JsonResponse({'result': result})
                for val in result:
                    if "pvc index" not in val:
                        vlan_info[val] = result.get(val)
                    else:
                        if result.get(val)['vpi'] != "0" or result.get(val)["vci"] != "0":
                            vlan_info[val] = result.get(val)
                pvc_vlan.append(vlan_info)
                return JsonResponse({'result': pvc_vlan})
                # return JsonResponse({'result': result})

            elif dslam_type == 1:  ############################## zyxel ##############################
                result = utility.dslam_port_run_command(dslamObj.pk, 'port Info', params)
                if 'name' not in result[0]:
                    return JsonResponse({'result': result})
                result = result[-2:]
                pvc_count = 0
                for val in result:
                    if "pvc" in val:
                        vlan_info = {}
                        pvc_count += 1
                        vlan_info["pvc"] = pvc_count
                        vlan_info["vlan_id"] = val.split()[0]
                        vlan_info['vpi'] = val.split()[1].split("-")[-1].split("/")[0]
                        vlan_info['vci'] = val.split()[1].split("-")[-1].split("/")[1]
                        print(pvc_vlan)
                        pvc_vlan.append(vlan_info)
                return JsonResponse({'result': pvc_vlan})
                # return JsonResponse({'result': result})

        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})


class AddToVlanAPIView(views.APIView):  # 000000

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, format=None):
        data = request.data
        # if data['port'].get('old_vlan_name', None) is None and 'an2200' in data['port'].get('fqdn', None):
        #     return Response(dict(results="The old_vlan_name field is required."), status=status.HTTP_400_BAD_REQUEST)
        port_data = data.get('port', None)
        fqdn = port_data.get('fqdn', None)
        reseller_data = data.get('reseller')
        customer_data = data.get('subscriber')
        dslam_obj = DSLAM.objects.get(fqdn=str(fqdn).lower())
        mdf_status = data.get('status')
        identifier_key = data.get('identifier_key')
        if not identifier_key:
            identifier_key = str(random.randint(10000000, 9999999999999999))
        dslam_type = dslam_obj.dslam_type_id
        if not mdf_status:
            mdf_status = 'BUSY'
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        log_port_data = "/".join([val for val in port_data.values()])
        log_username = customer_data.get('username')
        log_date = datetime.now()
        log_reseller_name = reseller_data.get('name')
        try:
            if '.z6' in fqdn:
                fqdn = fqdn.replace('.z6', '.Z6')

            # dslam_obj = DSLAM.objects.get(name=port_data.get('dslam_name'), telecom_center__name=port_data.get('telecom_center_name'))
            try:
                dslam_obj = DSLAM.objects.get(fqdn=fqdn)
                # return JsonResponse({'Result': serialize('json',[dslam_obj])}, status=status.HTTP_200_OK)
                telecom_id = dslam_obj.telecom_center_id
                city_id = TelecomCenter.objects.get(id=telecom_id).city_id
            except ObjectDoesNotExist as ex:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                log_params = PortmanLogging.prepare_variables(self, log_port_data, log_username, 'add to vlan', '',
                                                              log_date,
                                                              ip, 'Register Port', False,
                                                              str(ex) + '/' + str(exc_tb.tb_lineno),
                                                              log_reseller_name)
                PortmanLogging('', log_params)
                try:
                    if '.Z6' in fqdn:
                        fqdn = fqdn.replace('.Z6', '.z6')
                        dslam_obj = DSLAM.objects.get(fqdn=fqdn)
                        # return JsonResponse({'Result': dslam_obj.fqdn}, status=status.HTTP_200_OK)
                        telecom_id = dslam_obj.telecom_center_id
                        city_id = TelecomCenter.objects.get(id=telecom_id).city_id

                except ObjectDoesNotExist as ex:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    log_params = PortmanLogging.prepare_variables(self, log_port_data, log_username, 'add to vlan', '',
                                                                  log_date,
                                                                  ip, 'Register Port', False,
                                                                  str(ex) + '/' + str(exc_tb.tb_lineno),
                                                                  log_reseller_name)
                    PortmanLogging('', log_params)
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    mail_info = Mail()
                    mail_info.from_addr = 'oss-notification@pishgaman.net'
                    mail_info.to_addr = 'oss-problems@pishgaman.net'
                    mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                        str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'),
                        port_data.get('card_number'), port_data.get('port_number'), ip)
                    mail_info.msg_subject = 'Error in RegisterPortAPIView'
                    Mail.Send_Mail(mail_info)
                    return JsonResponse({'Result': 'Dslam Not Found. Please check FQDN again.'},
                                        status=status.HTTP_200_OK)

            telecom_mdf_obj = TelecomCenterMDF.objects.filter(telecom_center_id=dslam_obj.telecom_center.id)
            print((dslam_obj.telecom_center.id))
            if telecom_mdf_obj:
                telecom_mdf_obj = telecom_mdf_obj.first()
            print((dslam_obj.telecom_center))
            mdf_dslam_obj, mdf_dslam_updated = MDFDSLAM.objects.update_or_create(
                telecom_center_id=dslam_obj.telecom_center.id,
                telecom_center_mdf_id=telecom_mdf_obj.id,
                #### Check this whole line
                row_number=0, floor_number=0, connection_number=0,  ##### Check this whole line
                dslam_id=dslam_obj.id, slot_number=port_data.get('card_number'),
                port_number=port_data.get('port_number'),
                defaults={'status': mdf_status, 'identifier_key': identifier_key})
            # if mdf_dslam.status != 'FREE':
            #    return JsonResponse(
            #            {'result': 'port status is {0}'.format(mdf_dslam.status), 'id': -1}
            #            )
            # else:
            #    mdf_dslam.status = 'RESELLER'
            #    mdf_dslam.save()
            # identifier_key = mdf_dslam.identifier_key
        except ObjectDoesNotExist as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            log_params = PortmanLogging.prepare_variables(self, log_port_data, log_username, 'add to vlan', '',
                                                          log_date,
                                                          ip, 'Register Port', False,
                                                          str(ex) + '/' + str(exc_tb.tb_lineno),
                                                          log_reseller_name)
            PortmanLogging('', log_params)
            return JsonResponse({'result': str(ex), 'id': -1})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            log_params = PortmanLogging.prepare_variables(self, log_port_data, log_username, 'add to vlan', '',
                                                          log_date,
                                                          ip, 'Register Port', False,
                                                          str(ex) + '/' + str(exc_tb.tb_lineno),
                                                          log_reseller_name)
            PortmanLogging('', log_params)
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            try:
                return JsonResponse({'result': 'Error is {0}--{1}'.format(ex, exc_tb.tb_lineno)})
            except ObjectDoesNotExist as ex:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                log_params = PortmanLogging.prepare_variables(self, log_port_data, log_username, 'add to vlan', '',
                                                              log_date,
                                                              ip, 'Register Port', False,
                                                              str(ex) + '/' + str(exc_tb.tb_lineno),
                                                              log_reseller_name)
                PortmanLogging('', log_params)
                mail_info = Mail()
                mail_info.from_addr = 'oss-notification@pishgaman.net'
                mail_info.to_addr = 'oss-problems@pishgaman.net'
                mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                    str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'), port_data.get('card_number'),
                    port_data.get('port_number'), ip)
                mail_info.msg_subject = 'Error in RegisterPortAPIView'
                Mail.Send_Mail(mail_info)
                return JsonResponse({'Result': ''}, status=status.HTTP_200_OK)

        try:
            reseller_obj, reseller_created = Reseller.objects.get_or_create(name=reseller_data.get('name'))

            customer_obj, customer_updated = CustomerPort.objects.update_or_create(
                username=customer_data.get('username'),
                defaults={'identifier_key': identifier_key, 'telecom_center_id': mdf_dslam_obj.telecom_center_id}
            )

            rp = ResellerPort()
            rp.identifier_key = identifier_key
            rp.status = 'ENABLE'
            rp.dslam_id = dslam_obj.id
            rp.dslam_slot = port_data.get('card_number')
            rp.dslam_port = port_data.get('port_number')
            rp.reseller = reseller_obj
            rp.telecom_center_id = mdf_dslam_obj.telecom_center_id
            rp.save()

            vlan_objs = Vlan.objects.filter(reseller=reseller_obj)
            if not vlan_objs:
                raise ValidationError(dict(erorr=f"Vlan {reseller_obj.name} dose not exist."))
            print()
            'vlan_objs ->', vlan_objs
            # return JsonResponse({'vlan':  vlan_objs[0].vlan_id, 'vpi' : reseller_obj.vpi,'vci' : reseller_obj.vci})
            port_indexes = [{'slot_number': port_data.get('card_number'), 'port_number': port_data.get('port_number')}]
            pishParams = {
                "type": "dslamport",
                "is_queue": False,
                "vlan_id": 3900,
                "vlan_name": 'pte',
                "dslam_id": dslam_obj.id,
                "port_indexes": port_indexes,
                "username": customer_data.get('username'),
                "vpi": 0,
                "vci": 35,
            }
            vlan_name = vlan_objs[0].vlan_name
            if dslam_type == 3 or dslam_type == 4 or dslam_type == 5:
                vlan_name = str(reseller_obj).split('-')[1]
                if vlan_name == 'didehban':
                    vlan_name = 'dideban'
                if vlan_name == 'baharsamaneh':
                    vlan_name = 'baharsam'
                if vlan_name == 'badrrayan':
                    vlan_name = 'badrray'
                if vlan_name == 'farzanegan':
                    vlan_name = 'farzaneg'
                if vlan_name == 'fanap':
                    vlan_name = 'fanap'
            params = {
                "type": "dslamport",
                "is_queue": False,
                "vlan_id": vlan_objs[0].vlan_id,
                "vlan_name": vlan_name,
                "dslam_id": dslam_obj.id,
                "port_indexes": port_indexes,
                "username": customer_data.get('username'),
                "vpi": reseller_obj.vpi,
                "vci": reseller_obj.vci,
                "old_vlan_name": port_data.get('old_vlan_name', 'pte')
            }

            # res = utility.dslam_port_run_command(dslam_obj.id, 'delete from vlan', pishParams)
            result = utility.dslam_port_run_command(dslam_obj.id, 'add to vlan', params)
            # PVC = utility.dslam_port_run_command(dslam_obj.id, 'port pvc show', params)

            # result2 = utility.dslam_port_run_command(dslam_obj.id, 'add to vlan', params)
            log_status = True if isinstance(result, dict) else False
            log_params = PortmanLogging.prepare_variables(self, log_port_data, log_username, 'add to vlan', result,
                                                          log_date,
                                                          ip, 'Register Port', log_status, '', log_reseller_name)
            PortmanLogging(result, log_params)
            return Response({'result': result, 'id': 201}, status=status.HTTP_201_CREATED)

        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            log_params = PortmanLogging.prepare_variables(self, log_port_data, log_username, 'add to vlan', '',
                                                          log_date,
                                                          ip, 'Register Port', False,
                                                          str(ex) + '/' + str(exc_tb.tb_lineno),
                                                          log_reseller_name)
            PortmanLogging('', log_params)
            return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})


class FiberHomeCommandAPIView(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, format=None):
        device_ip = get_device_ip(request)
        data = request.data
        # print('============================================================================')
        # print(device_ip)
        # print(data)
        # print('============================================================================')
        command = data.get('command', None)
        command = command_recognise(command)
        fqdn = request.data.get('fqdn')
        dslamObj = DSLAM.objects.get(fqdn=str(fqdn).lower())
        params = data.get('params', None)
        dslam_type = dslamObj.dslam_type_id
        log_port_data = f"{fqdn}/{params['port_conditions']['slot_number']}/{params['port_conditions']['port_number']}"
        log_username = params.get('username')
        log_date = datetime.now()
        operator = params.get('operator')
        try:
            result = utility.dslam_port_run_command(dslamObj.pk, command, params)
            log_status = True if isinstance(result, dict) else False
            log_params = PortmanLogging.prepare_variables(self, log_port_data, log_username, command, result, log_date,
                                                          device_ip, 'Run Command', log_status, operator, '')
            PortmanLogging(result, log_params)
            # if dslam_type == 1:  ################################### zyxel ###################################
            #     return JsonResponse({'Result': dslam_type})
            if dslam_type == 1:  ################################### zyxel ###################################
                if not command:
                    return Response({'result': 'Command does not exits'}, status=status.HTTP_400_BAD_REQUEST)
                if result:
                    if 'Busy' in result:
                        return Response({'result': result}, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        description = 'Run Command {0} on DSLAM {1}'.format(command, dslamObj.name)

                        add_audit_log(request, 'DSLAMCommand', None, 'Run Command On DSLAM Port', description)
                port_info = utility.dslam_port_run_command(dslamObj.pk, 'port Info', params)
                current_user_profile = ""
                alarm_prof = ""
                port_state = ""
                for item in port_info['result']:
                    if 'prof' in item and 'alarm prof' not in item:
                        current_user_profile = item.split(':')[1]
                    if 'state' in item:
                        port_state = item.split(':')[1]
                        break

                return JsonResponse(
                    {'response': result, 'current_user_profile': current_user_profile, 'port_state': port_state})


            elif dslam_type == 2:  ############################## huawei ##############################
                if command == 'display_sippstnuser_reg_state':
                    card = 0
                    params['port_conditions']['slot_number'] = card
                    result = utility.dslam_port_run_command(dslamObj.pk, command, params)
                    while 'defined!' not in result.get('result'):
                        print(f"**************** {card} ************")
                        if 'specified section' not in result.get('result'):

                            for item in result.get('result'):
                                # phone_number =re.search(r"\(?\d{5}\)?[-.\s]?\d{3}[-.\s]?\d{4}", item)

                                values = item.split()
                                phone_number = values[-1]
                                register_state = values[-2]
                                mg_id = values[3]
                                port = values[2].replace("/", "")
                                slot = values[1].replace("/", "")
                                print(f"{phone_number}-{register_state}-{mg_id}-{port}-{slot}")
                                # rows = [[slot, port, mg_id, register_state, phone_number]]
                                # with open('data.csv', 'a', encoding='UTF8') as f:
                                #     writer = csv.writer(f)
                                #     writer.writerows(rows)
                                Order.objects.create(rastin_order_id=0, ranjePhoneNumber=phone_number,
                                                     username=phone_number,
                                                     slot_number=int(slot), port_number=int(port), telco_row=0,
                                                     telco_column=0, telco_connection=1, order_contact_id=0, user_id=0,
                                                     status_id=2, dslam_id=dslamObj.id, port_type_id=2)
                        card += 1
                        params['port_conditions']['slot_number'] = card
                        result = utility.dslam_port_run_command(dslamObj.pk, command, params)
                    return JsonResponse(
                        {'response': dict(result='Data added successfully.'), 'DslamType': 'huawei'})

                port_info = utility.dslam_port_run_command(dslamObj.pk, 'show profile by port', params)
                current_user_profile = ''
                if 'profile_name' not in port_info:
                    return JsonResponse(
                        {'response': result, 'current_user_profile': '', 'DslamType': 'huawei'})
                current_user_profile = port_info['profile_name']
                return JsonResponse(
                    {'response': result, 'current_user_profile': current_user_profile, 'DslamType': 'huawei'})

            elif dslam_type == 3:  ############################## fiberhomeAN3300 ##############################

                port_info = utility.dslam_port_run_command(dslamObj.pk, 'show linerate', params)
                current_user_profile = ''
                if 'profile_name' not in port_info:
                    return JsonResponse(
                        {'response': result, 'current_user_profile': '', 'DslamType': 'fiberhomeAN3300'})
                current_user_profile = port_info['profile_name']
                return JsonResponse(
                    {'response': result, 'current_user_profile': current_user_profile,
                     'DslamType': 'fiberhomeAN3300'})
            elif dslam_type == 4:  ############################## fiberhomeAN2200 ##############################

                port_info = utility.dslam_port_run_command(dslamObj.pk, 'show profile by port', params)
                current_user_profile = ''
                if 'profile_name' not in port_info:
                    return JsonResponse(
                        {'response': result, 'current_user_profile': '', 'DslamType': 'fiberhomeAN2200'})
                current_user_profile = port_info['profile_name']
                return JsonResponse(
                    {'response': result, 'current_user_profile': current_user_profile, 'DslamType': 'fiberhomeAN2200'})

            elif dslam_type == 5:  ############################## fiberhomeAN5006 ##############################
                if command == 'show mac':
                    from .services.fiber_home_get_card import FiberHomeGetCardStatusService
                    cards_status_class = FiberHomeGetCardStatusService(request.data)
                    cards_status = cards_status_class.get_status_cards()
                    params['cards_status'] = cards_status
                    result = utility.dslam_port_run_command(dslamObj.pk, command, params)
                port_info = utility.dslam_port_run_command(dslamObj.pk, 'show profile by port', params)
                current_user_profile = ''
                if 'profile_name' not in port_info:
                    return JsonResponse(
                        {'response': result, 'current_user_profile': '', 'DslamType': 'fiberhomeAN5006'})
                current_user_profile = port_info['profile_name']
                return JsonResponse(
                    {'response': result, 'current_user_profile': current_user_profile, 'DslamType': 'fiberhomeAN5006'})


            elif dslam_type == 7:  ########################### zyxel1248 ##########################
                port_info = utility.dslam_port_run_command(dslamObj.pk, 'show profile by port', params)
                current_user_profile = ''
                if 'profile_name' not in port_info:
                    return JsonResponse(
                        {'response': result, 'current_user_profile': '', 'DslamType': 'zyxel1248'})
                current_user_profile = port_info['profile_name']
                return JsonResponse(
                    {'response': result, 'current_user_profile': current_user_profile, 'DslamType': 'zyxel1248'})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            log_params = PortmanLogging.prepare_variables(self, log_port_data, log_username, command, '', log_date,
                                                          device_ip, 'Run Command', False,
                                                          operator, str(ex) + '/' + str(exc_tb.tb_lineno))
            PortmanLogging('', log_params)
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})


class RunCommandByIPAPIView(views.APIView):
    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, format=None):
        try:
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip = x_forwarded_for.split(',')[0]
            else:
                ip = request.META.get('REMOTE_ADDR')

            user = request.user
            data = request.data

            s = []
            command = data.get('command', None)
            subscriber = data.get('subscriber')
            dslam_ip = data.get('dslam_ip')
            dslamObj = DSLAM.objects.get(ip=dslam_ip)
            dslam_id = dslamObj.id

            userProfile = ''  # DSLAMPort.objects.get(port_number = request.data.get('params').get('port_conditions').get('port_number'), slot_number = request.data.get('params').get('port_conditions').get('slot_number'), dslam_id = dslam_id).line_profile
            params = data.get('params', None)
            try:
                dslam_obj = DSLAM.objects.get(id=dslam_id)
                fiber = DslanInfo()
                fiber.type = 'Fiberhome'
                fiber.dslam = dslam_obj.ip
                fiber.telnetPort = 23
                fiber.userName = dslam_obj.telnet_username
                fiber.password = dslam_obj.telnet_password
                fiber.access = dslam_obj.access_name
                fiber.card = request.data.get('params').get('port_conditions').get('slot_number')
                fiber.port = request.data.get('params').get('port_conditions').get('port_number')
                fiber.terminalDelay = '300'
                fiber.requestTimeOut = '1500'
                # ---------------------------------------------------Fiberhome2200---------------------------------------------------
                if (dslam_obj.dslam_type_id == 4):
                    profile = request.data.get('params').get('new_lineprofile')
                    if (command == 'selt'):
                        return JsonResponse({'res': 'this command is not supported by this dslam'})
                    if (command == 'show linerate' or command == 'showPort' or command == 'showPort'):
                        fiber.command = 'showPort'
                    if (command == 'profile adsl show' or command == 'showProfiles'):
                        fiber.command = 'showProfiles'
                    if (command == 'setPortProfiles' or command == 'Set Port Profiles' or command == 'test'):
                        fiber.command = 'setPortProfiles'
                    url = 'http://5.202.129.88:9096/api/Telnet/telnet'
                    data = "{'type':'Fiberhome','dslam':'%s','telnetPort':'23','userName':'%s','password':'%s','access':'%s','card':'%s','port':'%s','command':'%s','profile':'%s','terminalDelay':'600','requestTimeOut':'1500'}" % (
                        fiber.dslam, fiber.userName, fiber.password, fiber.access, fiber.card, fiber.port,
                        fiber.command,
                        profile)
                    # return JsonResponse({'current_userProfile':userProfile ,'response':data   })

                    fhresponse = requests.post(url, data=data, headers={"Content-Type": "application/json"})
                    sid = fhresponse.json()
                    try:
                        if (command == 'show linerate' or command == 'showPort' or command == 'showPort'):
                            res = sid.split("\n\r")

                            fhome = fInfo()
                            fhome.Date = sid.split("\n\r")[0]
                            portInfo = [int(c) for c in res[3].split() if c.isdigit()]

                            return JsonResponse({'current_userProfile': userProfile,
                                                 'dslamName/cammandName': dslam_obj.name + '/' + command,
                                                 'date': res[0].split(":")[1] + res[0].split(":")[2] + ':' +
                                                         res[0].split(":")[3],
                                                 'slot/port': str(portInfo[1]) + '-' + str(portInfo[2]),
                                                 # 'OP_State' : res[4].split(":")[1],
                                                 # 'Standard' : res[5].split(":")[1],
                                                 # 'Latency' : res[6].split(":")[1],
                                                 'noisemarginDown': res[7].split(":")[1].split("/")[0],
                                                 'noisemarginUp': res[7].split(":")[1].split("/")[1].split(" ")[0],
                                                 'payloadrateDown': res[8].split(":")[1].split("/")[0],
                                                 'payloadrateUp': res[8].split(":")[1].split("/")[1].split(" ")[0],
                                                 'attenuationDown': res[9].split(":")[1].split("/")[0],
                                                 'attenuationUp': res[9].split(":")[1].split("/")[1].split(" ")[0],
                                                 # 'Tx power(D/U)' : res[10].split(":")[1].split("/")[0],
                                                 # 'Tx power(D/U)' : res[10].split(":")[1].split("/")[1].split(" ")[0],
                                                 # 'Remote vendor ID' : res[11].split(":")[1],
                                                 # 'Power management state' : res[12].split(":")[1],
                                                 # 'Remote vendor version ID' : res[13].split(":")[1],
                                                 # 'Loss of power(D)' : res[14].split(":")[1].split("/")[0],
                                                 # 'Loss of power(U)' : res[14].split(":")[1].split("/")[1].split(" ")[0],
                                                 # 'Loss of signal(D)' : res[15].split(":")[1].split("/")[0],
                                                 # 'Loss of signal(U)' : res[15].split(":")[1].split("/")[1].split(" ")[0],
                                                 # 'Error seconds(D)' : res[16].split(":")[1].split("/")[0],
                                                 # 'Error seconds(U)' : res[16].split(":")[1].split("/")[1].split(" ")[0],
                                                 # 'Loss by HEC collision(D)' : res[17].split(":")[1].split("/")[0],
                                                 # 'Loss by HEC collision(U)' : res[17].split(":")[1].split("/")[1].split(" ")[0],
                                                 # 'Forward correct(D)' : res[18].split(":")[1].split("/")[0],
                                                 # 'Forward correct(U)' : res[18].split(":")[1].split("/")[1],
                                                 # 'Uncorrect(D)' : res[19].split(":")[1].split("/")[0],
                                                 # 'Uncorrect(U)' : res[19].split(":")[1].split("/")[1],
                                                 'attainablerateDown': res[20].split(":")[1].split("/")[0],
                                                 'attainablerateUp': res[20].split(":")[1].split("/")[1],
                                                 # 'Interleaved Delay(D) ' : res[21].split(":")[1].split("/")[0],
                                                 # 'Interleaved Delay(U) ' : res[21].split(":")[1].split("/")[1],
                                                 # 'Remote loss of link' : res[22].split(":")[1],
                                                 })
                            return JsonResponse({'current_userProfile': userProfile, 'response': sid.split("\n\r")})
                        elif (command == 'profile adsl show' or command == 'showProfiles'):
                            return JsonResponse({'current_userProfile': userProfile, 'response': sid.split("\n\r")})
                        elif (command == 'set port profile' or command == 'profile Change'):
                            return JsonResponse({'current_userProfile': userProfile, 'response': sid})

                        else:
                            return JsonResponse({'current_userProfile': userProfile, 'response': sid.split("\n\r")})
                    except Exception as ex:
                        try:
                            return JsonResponse({'current_userProfile': userProfile, 'response': sid.split("\n\r")})
                        except Exception as ex:
                            exc_type, exc_obj, exc_tb = sys.exc_info()
                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                            mail_info = Mail()
                            mail_info.from_addr = 'oss-notification@pishgaman.net'
                            mail_info.to_addr = 'oss-problems@pishgaman.net'
                            mail_info.msg_body = 'Error in RunCommandAPIViewin/Fiberhome2200 Line {0}.Error Is: {1}. fqdn = {2},card = {3}, port = {4}, command = {5}, subscriber = {6}. IP: {7}'.format(
                                str(exc_tb.tb_lineno), str(ex), request.data.get('fqdn'),
                                request.data.get('params').get('port_conditions').get('slot_number'),
                                request.data.get('params').get('port_conditions').get('port_number'), command,
                                subscriber, ip)
                            mail_info.msg_subject = 'Error in RunCommandAPIView/Fiberhome2200'
                            Mail.Send_Mail(mail_info)



                # ---------------------------------------------------Fiberhome5006---------------------------------------------------
                elif (dslam_obj.dslam_type_id == 5):
                    if (command == 'selt'):
                        return JsonResponse({'res': 'this command is not supported by this dslam'})
                    if (command == 'show linerate' or command == 'showPort'):
                        fiber.command = 'showPort'
                    if (command == 'profile adsl show' or command == 'showProfiles'):
                        fiber.command = 'showProfiles'
                    url = 'http://5.202.129.88:9096/api/Telnet/telnet'
                    data = "{'type':'FiberhomeAN5006','dslam':'%s','telnetPort':'23','userName':'%s','password':'%s','access':'%s','card':'%s','port':'%s','command':'%s','terminalDelay':'600','requestTimeOut':'1500'}" % (
                        fiber.dslam, fiber.userName, fiber.password, fiber.access, fiber.card, fiber.port,
                        fiber.command)
                    fhresponse = requests.post(url, data=data, headers={"Content-Type": "application/json"})
                    sid = fhresponse.json()
                    res = sid.split("\n\r")
                    return JsonResponse({'current_userProfile': userProfile, 'response': sid.split("\r\n")})
                elif (dslam_obj.dslam_type_id == 3):
                    if (command == 'show linerate' or command == 'showPort'):
                        fiber.command = 'showPort'

                    url = 'http://5.202.129.88:9096/api/Telnet/telnet'
                    data = "{'type':'FiberhomeAN3300','dslam':'%s','telnetPort':'23','userName':'%s','password':'%s','access':'%s','card':'%s','port':'%s','command':'%s','terminalDelay':'600','requestTimeOut':'1500'}" % (
                        fiber.dslam, fiber.userName, fiber.password, fiber.access, fiber.card, fiber.port,
                        fiber.command)
                    fhresponse = requests.post(url, data=data, headers={"Content-Type": "application/json"})
                    sid = fhresponse.json()
                    res = sid.split("\n\r")
                    return JsonResponse({'current_userProfile': userProfile, 'response': sid.split("\r\n")})
                    # ---------------------------------------------------huawei---------------------------------------------------
                    if (dslam_obj.dslam_type_id == 2):
                        try:
                            if (
                                    command == 'show mac slot port' or command == 'show mac with port' or command == 'show mac'):
                                command = 'show mac slot port'
                            if not command:
                                return Response({'result': 'Command does not exits'},
                                                status=status.HTTP_400_BAD_REQUEST)
                            result = utility.dslam_port_run_command(dslam_obj.pk, command, params)
                            return JsonResponse({'response': result['result'].split('\r\n')[31].split()[2]})
                        except Exception as ex:
                            try:
                                if ('There is not any MAC address record' in result['result']):
                                    return JsonResponse({'response': 'There is not any MAC address record'})
                                return JsonResponse({'response': result['result'].split('\r\n')})
                            except Exception as ex:
                                exc_type, exc_obj, exc_tb = sys.exc_info()
                                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                return JsonResponse({'result': str(
                                    'an error occurred. please try again.{0}'.format(str(exc_tb.tb_lineno)))})

                    # ---------------------------------------------------zyxel---------------------------------------------------
                    description = 'Run Command {0} on DSLAM {1}'.format(
                        command,
                        dslam_obj.name,
                    )
            except Exception as ex:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                return Response({'result': str(ex), 'Line': str(exc_tb.tb_lineno)}, status=status.HTTP_400_BAD_REQUEST)
                return JsonResponse({'result': str('an error occurred. please try again')})

            if (command == 'show linerate' or command == 'showPort'):
                command = 'show linerate'
            elif (command == 'profile adsl show' or command == 'showProfiles'):
                command = 'profile adsl show'
            elif (command == 'open port' or command == 'port enable'):
                command = 'port enable'
            elif (command == 'close port' or command == 'port disable'):
                command = 'port disable'
            elif (command == 'show mac slot port' or command == 'showmacslotport'):
                command = 'show mac slot port'

            if not command:
                return Response({'result': 'Command does not exits'}, status=status.HTTP_400_BAD_REQUEST)
            params["username"] = user.username
            result = utility.dslam_port_run_command(dslam_obj.pk, command, params)
            if result:
                if 'Busy' in result:
                    return Response({'result': result}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    description = 'Run Command {0} on DSLAM {1}'.format(
                        command,
                        dslam_obj.name)

                    add_audit_log(request, 'DSLAMCommand', None, 'Run Command On DSLAM Port', description)
            if (command == "show linerate"):
                try:
                    sp = showPort()
                    result = result['result'].split("\r\n")
                    sp.dslamNameAndcammandName = result[1]
                    sp.slotPort = result[2].split(",")[0].split("=")[1]
                    sp.link = result[3].replace(" ", "").split("=")[1]
                    payloadrate = result[6].split("=")[1]
                    sp.payloadrate = [int(s) for s in payloadrate.split() if s.isdigit()]
                    actualrate = result[7].split("=")[1]
                    sp.actualrate = [int(c) for c in actualrate.split() if c.isdigit()]
                    attainablerate = result[8].split("=")[1]
                    sp.attainablerate = [int(c) for c in attainablerate.split() if c.isdigit()]
                    noisemargin = result[9].split("=")[1]
                    sp.noisemargin = re.findall(r"[-+]?\d*\.\d+|\d+", noisemargin)
                    attenuation = result[10].split("=")[1]
                    sp.attenuation = re.findall(r"[-+]?\d*\.\d+|\d+", attenuation)
                    return JsonResponse({
                        'dslamName/cammandName': sp.dslamNameAndcammandName,
                        'slot/Port': sp.slotPort,
                        'link': sp.link,
                        'payloadrateUp': sp.payloadrate[0],
                        'payloadrateDown': sp.payloadrate[1],
                        'actualrateUp': sp.actualrate[0],
                        'actualrateDown': sp.actualrate[1],
                        'attainablerateUp': sp.attainablerate[0],
                        'attainablerateDown': sp.attainablerate[1],
                        'noisemarginUp': sp.noisemargin[0],
                        'noisemarginDown': sp.noisemargin[1],
                        'attenuationUp': sp.attenuation[0],
                        'attenuationDown': sp.attenuation[1],
                        'current_userProfile': userProfile
                    }, status=status.HTTP_200_OK)
                except Exception as ex:
                    return JsonResponse({'current_userProfile': userProfile, 'response': result})
            elif (command == "show lineinfo"):
                return JsonResponse({'current_userProfile': userProfile, 'response': result['result'].split("\r\n")})
            elif (command == "selt"):
                return JsonResponse({'current_userProfile': userProfile, 'response': result})
            elif (command == 'profile adsl show'):
                response = result['result']
                my_list = []
                for item in response:
                    my_list.append(item['name'])
                return Response({'response': my_list})
            elif (
                    command == 'port enable' or command == 'port disable' or command == 'port pvc show' or command == 'get config'):
                return JsonResponse({'response': result})
            elif (command == "show mac slot port"):
                portPvc = utility.dslam_port_run_command(dslam_obj.pk, 'port pvc show', params)
                res = utility.dslam_port_run_command(dslam_obj.pk, 'show linerate', params)
                payload = res['result'].split('\r\n')[6].split("=")[1]
                payloadrate = [int(s) for s in payload.split() if s.isdigit()]
                return JsonResponse({'Port PVC': portPvc, 'response': result['result'], 'payloadrateUp': payloadrate[0],
                                     'payloadrateDown': payloadrate[1]})
            else:
                return JsonResponse({'response': result})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            # return JsonResponse({'result': str('an error occurred. please try again. {0}'.format(str(exc_tb.tb_lineno)))}, status=status.HTTP_202_ACCEPTED)
            mail_info = Mail()
            mail_info.from_addr = 'oss-notification@pishgaman.net'
            mail_info.to_addr = 'oss-problems@pishgaman.net'
            mail_info.msg_body = 'Error in RunCommandAPIViewin Line {0}.Error Is: {1}. fqdn = {2},card = {3}, port = {4}, command = {5}, subscriber = {6}. IP: {7}'.format(
                str(exc_tb.tb_lineno), str(ex), request.data.get('fqdn'),
                request.data.get('params').get('port_conditions').get('slot_number'),
                request.data.get('params').get('port_conditions').get('port_number'), command, subscriber, ip)
            mail_info.msg_subject = 'Error in RunCommandAPIView'
            Mail.Send_Mail(mail_info)
            return JsonResponse(
                {'result': str('an error occurred. please try again. {0}'.format(str(exc_tb.tb_lineno)))},
                status=status.HTTP_202_ACCEPTED)


class GetDslamBackupByIdAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            dslam_id = data.get('id')
            return JsonResponse({'result': dslam_id})
        except Exception as ex:
            return JsonResponse({'result': str('an error occurred. please try again.')},
                                status=status.HTTP_202_ACCEPTED)


class UpdateProfileAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            dslamPotr_obj = DSLAMPort.objects.get(id=request.data.get('portId'))
            dslam_obj = DSLAM.objects.get(id=dslamPotr_obj.dslam_id)

            params = param()
            params.fqdn = dslam_obj.fqdn
            # dslam_number = dslam_obj.fqdn[-1:]
            params.type = 'dslamport'
            params.is_queue = False
            # params.dslam_number = dslam_number
            params.port_conditions = port_condition()
            params.port_conditions.slot_number = dslamPotr_obj.slot_number
            params.port_conditions.port_number = dslamPotr_obj.port_number
            params = json.dumps(params, default=lambda x: x.__dict__)
            # return JsonResponse({'result': params }, status=status.HTTP_202_ACCEPTED)
            res = utility.dslam_port_run_command(dslam_obj.pk, 'port Info', json.loads(params))
            profile = res['result'][7].split(':')[1]
            dslamPotr_obj.line_profile = profile
            dslamPotr_obj.save()
            return JsonResponse({'result': profile}, status=status.HTTP_202_ACCEPTED)
        except Exception as ex:
            try:
                profile = res['result'][6].split(':')[1]
                dslamPotr_obj.line_profile = profile
                dslamPotr_obj.save()
                return JsonResponse({'result': profile}, status=status.HTTP_202_ACCEPTED)
            except Exception as ex:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                return JsonResponse({'result': str(ex), 'line': str(exc_tb.tb_lineno)},
                                    status=status.HTTP_400_BAD_REQUEST)


class GetPortDownstreamAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        # dslam_objs = DSLAM.objects.annotate(Count('fqdn')).order_by('fqdn')[:2]
        params = param()
        dslam_objs = DSLAM.objects.filter(dslam_type_id=1, ip__startswith='172')

        try:
            for dslam_obj in dslam_objs:
                return JsonResponse({'result': str(dslam_obj.ip)},
                                    status=status.HTTP_400_BAD_REQUEST)

                telecom_center_id = DSLAM.objects.get(id=dslam_obj.id).telecom_center_id
                telecom_center = TelecomCenter.objects.get(id=telecom_center_id).name
                params = param2()
                params.type = 'dslamport'
                params.is_queue = False
                params.fqdn = dslam_obj.fqdn
                params.port_conditions = port_condition2()
                params.port_conditions.slot_number = 1
                params.port_conditions.port_number = 1
                params = json.dumps(params, default=lambda x: x.__dict__)
                dslam_result = utility.dslam_port_run_command(dslam_obj.pk, 'show linerate', json.loads(params))
                if ('inactive' in dslam_result['result']):
                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(dslam_obj.ip,
                                                                                                      card, port,
                                                                                                      'inactive!!!',
                                                                                                      datetime.now())
                    cursor = connection.cursor()
                    cursor.execute(query)
                elif ('downstream' in dslam_result['result']):

                    for card in range(1, 33):
                        for port in range(1, 73):
                            try:
                                params = param2()
                                params.type = 'dslamport'
                                params.is_queue = False
                                params.fqdn = dslam_obj.fqdn
                                params.port_conditions = port_condition2()
                                params.port_conditions.slot_number = card
                                params.port_conditions.port_number = port
                                slot_number = card
                                port_number = port
                                params = json.dumps(params, default=lambda x: x.__dict__)
                                port_obj = DSLAMPort.objects.get(dslam=dslam_obj, slot_number=slot_number,
                                                                 port_number=port_number)
                                result = utility.dslam_port_run_command(dslam_obj.pk, 'show linerate',
                                                                        json.loads(params))
                                if ('inactive' in result['result']):
                                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                                        dslam_obj.ip, card, port, 'inactive!!!', datetime.now())
                                    cursor = connection.cursor()
                                    cursor.execute(query)
                                elif ('downstream' in result['result']):
                                    payloadrate = result['result'].split('\r\n')[6].split("=")[1]
                                    payloadrate1 = [int(s) for s in payloadrate.split() if s.isdigit()]
                                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                                        dslam_obj.ip, card, port, 'Error!!!', datetime.now())
                                    cursor = connection.cursor()
                                    cursor.execute(query)
                                else:
                                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                                        dslam_obj.ip, card, port, 'Error!!!', datetime.now())
                                    cursor = connection.cursor()
                                    cursor.execute(query)
                            except ObjectDoesNotExist as ex:
                                dslam_port = DSLAMPort()
                                dslam_port.dslam = dslam_obj
                                dslam_port.created_at = datetime.now()
                                dslam_port.updated_at = datetime.now()
                                dslam_port.slot_number = slot_number
                                dslam_port.port_number = port_number
                                dslam_port.port_index = '{0}{1}'.format(slot_number, port_number)
                                dslam_port.port_name = 'adsl{0}-{1}'.format(slot_number, port_number)
                                dslam_port.admin_status = 'UNLOCK'
                                dslam_port.oper_status = 'NO-SYNC'
                                dslam_port.line_profile = '768/10240/IR'
                                dslam_port.selt_value = ''
                                dslam_port.uptime = '18:55:20'
                                dslam_port.upstream_snr = '192'
                                dslam_port.downstream_snr = '230'
                                dslam_port.upstream_attenuation = '69'
                                dslam_port.downstream_attenuation = '188'
                                dslam_port.upstream_attainable_rate = '2434'
                                dslam_port.downstream_attainable_rate = '157'
                                dslam_port.upstream_tx_rate = '93'
                                dslam_port.downstream_tx_rate = '1250'
                                dslam_port.upstream_snr_flag = 'good'
                                dslam_port.downstream_snr_flag = 'excellent'
                                dslam_port.upstream_attenuation_flag = 'outstanding'
                                dslam_port.downstream_attenuation_flag = 'outstanding'
                                dslam_port.vpi = 0
                                dslam_port.vci = 35
                                dslam_port.save()
                                port_obj = DSLAMPort.objects.get(dslam=dslam_obj, slot_number=slot_number,
                                                                 port_number=port_number)
                                result = utility.dslam_port_run_command(dslam_obj.pk, 'show linerate',
                                                                        json.loads(params))
                                if ('inactive' in result['result']):
                                    payloadrate = result['result'].split('\r\n')[6].split("=")[1]
                                    payloadrate1 = [int(s) for s in payloadrate.split() if s.isdigit()]
                                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                                        dslam_obj.ip, card, port, 'inactive!!!', datetime.now())
                                    cursor = connection.cursor()
                                    cursor.execute(query)
                                elif ('downstream' in result['result']):
                                    payloadrate = result['result'].split('\r\n')[6].split("=")[1]
                                    payloadrate1 = [int(s) for s in payloadrate.split() if s.isdigit()]
                                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                                        dslam_obj.ip, card, port, payloadrate1[1], datetime.now())
                                    cursor = connection.cursor()
                                    cursor.execute(query)

                                else:
                                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                                        dslam_obj.ip, card, port, 'Error!!!', datetime.now())
                                    cursor = connection.cursor()
                                    cursor.execute(query)

                            except Exception as ex:
                                exc_type, exc_obj, exc_tb = sys.exc_info()
                                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                return JsonResponse({'result': str(
                                    'an error occurred. please try again. {0}//{1}//{2}//{3}-{4}'.format(str(ex), str(
                                        exc_tb.tb_lineno), dslam_obj.ip, slot_number, port_number))},
                                    status=status.HTTP_202_ACCEPTED)
                                return JsonResponse(
                                    {'result': str(ex), 'dslam': dslam_obj.ip, 'slot': slot_number,
                                     'port': port_number},
                                    status=status.HTTP_202_ACCEPTED)
                else:
                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                        dslam_obj.ip, card, port, 'Error5!!!', datetime.now())
                    cursor = connection.cursor()
                    cursor.execute(query)
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return JsonResponse(
                {'result': str('an error occurred. please try again. {0}//{1}'.format(str(ex)), str(exc_tb.tb_lineno))},
                status=status.HTTP_202_ACCEPTED)


class SaveLineStatsAPIView(views.APIView):
    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        # dslam_objs = DSLAM.objects.annotate(Count('fqdn')).order_by('fqdn')[:2]
        params = param()
        dslam_objs = DSLAM.objects.filter(dslam_type_id=1, ip__startswith='172')

        try:
            for dslam_obj in dslam_objs:
                # return JsonResponse({'result': str(dslam_obj.ip)}, status=status.HTTP_400_BAD_REQUEST)
                query = "INSERT INTO IPs values('{0}')".format(dslam_obj.ip)
                # return JsonResponse({'result': str(query)}, status=status.HTTP_400_BAD_REQUEST)
                cursor = connection.cursor()
                cursor.execute(query)
                try:
                    telecom_center_id = DSLAM.objects.get(id=dslam_obj.id).telecom_center_id
                    telecom_center = TelecomCenter.objects.get(id=telecom_center_id).name
                    params = param2()
                    params.type = 'dslamport'
                    params.is_queue = False
                    params.fqdn = dslam_obj.fqdn
                    params.port_conditions = port_condition2()
                    params.port_conditions.slot_number = 1
                    params.port_conditions.port_number = 1
                    params = json.dumps(params, default=lambda x: x.__dict__)
                    dslam_result = utility.dslam_port_run_command(dslam_obj.pk, 'show linerate', json.loads(params))
                    if ('inactive' in dslam_result['result']):
                        query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(dslam_obj.ip,
                                                                                                          0, 0,
                                                                                                          'inactive!!!',
                                                                                                          datetime.now())
                        cursor = connection.cursor()
                        cursor.execute(query)
                    elif ('downstream' in dslam_result['result']):

                        for card in range(1, 33):
                            try:
                                params = param2()
                                params.type = 'dslamport'
                                params.is_queue = False
                                params.fqdn = dslam_obj.fqdn
                                params.port_conditions = port_condition2()
                                params.port_conditions.slot_number = card
                                params.port_conditions.port_number = 1
                                slot_number = card
                                port_number = 1
                                params = json.dumps(params, default=lambda x: x.__dict__)
                                result = utility.dslam_port_run_command(dslam_obj.pk, 'show linestat slot',
                                                                        json.loads(params))
                                if ('slot must be active' in result['result']):
                                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                                        dslam_obj.ip, card, 0, 'slot must be active!!!', datetime.now())
                                    cursor = connection.cursor()
                                    cursor.execute(query)
                                elif ('slot is between' in result['result']):
                                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                                        dslam_obj.ip, card, 0, 'There is not slot {0} !!!'.format(card), datetime.now())
                                    cursor = connection.cursor()
                                    cursor.execute(query)
                                elif ('is msc slot' in result['result']):
                                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                                        dslam_obj.ip, card, 0, 'slot {0} is msc slot!!!'.format(card), datetime.now())
                                    cursor = connection.cursor()
                                    cursor.execute(query)

                                else:
                                    res = result['result'].replace(
                                        "    Press any key to continue, 'n' to nopause,'e' to exit", '')
                                    # return JsonResponse({'result': res.split('\r\n') }, status=status.HTTP_400_BAD_REQUEST)

                                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                                        dslam_obj.ip, card, 0, res, datetime.now())
                                    cursor = connection.cursor()
                                    cursor.execute(query)
                            except ObjectDoesNotExist as ex:
                                dslam_port = DSLAMPort()
                                dslam_port.dslam = dslam_obj
                                dslam_port.created_at = datetime.now()
                                dslam_port.updated_at = datetime.now()
                                dslam_port.slot_number = slot_number
                                dslam_port.port_number = port_number
                                dslam_port.port_index = '{0}{1}'.format(slot_number, port_number)
                                dslam_port.port_name = 'adsl{0}-{1}'.format(slot_number, port_number)
                                dslam_port.admin_status = 'UNLOCK'
                                dslam_port.oper_status = 'NO-SYNC'
                                dslam_port.line_profile = '768/10240/IR'
                                dslam_port.selt_value = ''
                                dslam_port.uptime = '18:55:20'
                                dslam_port.upstream_snr = '192'
                                dslam_port.downstream_snr = '230'
                                dslam_port.upstream_attenuation = '69'
                                dslam_port.downstream_attenuation = '188'
                                dslam_port.upstream_attainable_rate = '2434'
                                dslam_port.downstream_attainable_rate = '157'
                                dslam_port.upstream_tx_rate = '93'
                                dslam_port.downstream_tx_rate = '1250'
                                dslam_port.upstream_snr_flag = 'good'
                                dslam_port.downstream_snr_flag = 'excellent'
                                dslam_port.upstream_attenuation_flag = 'outstanding'
                                dslam_port.downstream_attenuation_flag = 'outstanding'
                                dslam_port.vpi = 0
                                dslam_port.vci = 35
                                dslam_port.save()
                                port_obj = DSLAMPort.objects.get(dslam=dslam_obj, slot_number=slot_number,
                                                                 port_number=port_number)
                                result = utility.dslam_port_run_command(dslam_obj.pk, 'show linestat slot',
                                                                        json.loads(params))
                                if ('inactive' in result['result']):
                                    payloadrate = result['result'].split('\r\n')[6].split("=")[1]
                                    payloadrate1 = [int(s) for s in payloadrate.split() if s.isdigit()]
                                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                                        dslam_obj.ip, card, 0, 'inactive!!!', datetime.now())
                                    cursor = connection.cursor()
                                    cursor.execute(query)
                                elif ('downstream' in result['result']):
                                    payloadrate = result['result'].split('\r\n')[6].split("=")[1]
                                    payloadrate1 = [int(s) for s in payloadrate.split() if s.isdigit()]
                                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                                        dslam_obj.ip, card, 0, payloadrate1[1], datetime.now())
                                    cursor = connection.cursor()
                                    cursor.execute(query)

                                else:
                                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                                        dslam_obj.ip, card, 0, 'Error3!!!', datetime.now())
                                    cursor = connection.cursor()
                                    cursor.execute(query)

                            except Exception as ex:
                                exc_type, exc_obj, exc_tb = sys.exc_info()
                                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                return JsonResponse({'result': str(
                                    'an error occurred. please try again. {0}//{1}//{2}//{3}-{4}'.format(str(ex), str(
                                        exc_tb.tb_lineno), dslam_obj.ip, slot_number, port_number))},
                                    status=status.HTTP_202_ACCEPTED)
                                return JsonResponse(
                                    {'result': str(ex), 'dslam': dslam_obj.ip, 'slot': slot_number,
                                     'port': port_number},
                                    status=status.HTTP_202_ACCEPTED)
                    else:
                        query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(
                            dslam_obj.ip, 0, 0, 'Dslam Error!!!', datetime.now())
                        cursor = connection.cursor()
                        cursor.execute(query)
                except Exception as ex:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    # return JsonResponse({'result': str(dslam_obj.ip)}, status=status.HTTP_400_BAD_REQUEST)
                    query = "INSERT INTO down_stream values('{0}','{1}', '{2}', '{3}', '{4}')".format(dslam_obj.ip, 0,
                                                                                                      0, 'Line: ' + str(
                            exc_tb.tb_lineno), datetime.now())
                    # return JsonResponse({'result': str(dslam_obj.ip)}, status=status.HTTP_400_BAD_REQUEST)

                    cursor = connection.cursor()
                    cursor.execute(query)
            return JsonResponse(
                {'result': 'Done'},
                status=status.HTTP_400_BAD_REQUEST)

        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return JsonResponse(
                {'result': str(ex), 'Line': str(exc_tb.tb_lineno)},
                status=status.HTTP_400_BAD_REQUEST)


class GetSeltByFqdnAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')

        try:
            data = request.data
            fqdn = data.get('fqdn')
            dslam_obj = DSLAM.objects.get(fqdn=fqdn)
            if (dslam_obj.dslam_type_id != 1):
                return JsonResponse({'result': 'this command is not supported by this dslam'},
                                    status=status.HTTP_400_BAD_REQUEST)
            slot_number = data.get('slot_number')
            port_number = data.get('port_number')
            params = param()
            params.fqdn = fqdn
            params.type = 'dslamport'
            params.is_queue = False
            params.port_conditions = port_condition()
            params.port_conditions.slot_number = slot_number
            params.port_conditions.port_number = port_number
            params = json.dumps(params, default=lambda x: x.__dict__)
            dslam_result = utility.dslam_port_run_command(dslam_obj.pk, 'selt', json.loads(params))

            return JsonResponse(
                {'result': dslam_result},
                status=status.HTTP_200_OK)
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            mail_info = Mail()
            mail_info.from_addr = 'oss-notification@pishgaman.net'
            mail_info.to_addr = 'oss-problems@pishgaman.net'
            mail_info.msg_body = 'Error in GetSeltByFqdnAPIView Line {0}.Error Is: {1}. fqdn = {2},card = {3}, port = {4}, command = {5}, subscriber = {6}. System: {7}'.format(
                str(exc_tb.tb_lineno), str(ex), data.get('fqdn'),
                data.get('slot_number'),
                data.get('port_number'), 'Selt From Rastin', 'Rastin Bss', ip)
            mail_info.msg_subject = 'Error in GetSeltByFqdnAPIView'
            Mail.Send_Mail(mail_info)

            return JsonResponse(
                {'result': str('An Error has occurred!!'), 'Error Code': str(exc_tb.tb_lineno)},
                status=status.HTTP_400_BAD_REQUEST)


class PortConflictCorrectionAPIView(views.APIView):
    """
    "port":{
        "telecom_center_name": "string",
        "fqdn": "string",
        "card_number": "string",
        "port_number": "string"
    },
    "status": "RESELLER",
    "identifier_key":"string",
    "reseller": {
          "name": "string",
    },
    "subscriber":{
          "username": "string",
    }
    """

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        # return JsonResponse({'id': 201}, status=status.HTTP_201_CREATED)

        user = request.user
        data = request.data
        res = ''
        result = ''
        PVC = ''
        print()
        data
        reseller_data = data.get('reseller')
        old_reseller_data = data.get('old_reseller')
        customer_data = data.get('subscriber')
        mdf_status = data.get('status')
        sid = ''
        PVC = ''
        pvc = ''
        if not mdf_status:
            mdf_status = 'BUSY'
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        identifier_key = data.get('identifier_key')
        if not identifier_key:
            identifier_key = str(time.time())
        port_data = data.get('port')
        old_port_data = data.get('old_port')
        try:
            fqdn = data.get('fqdn')
            old_fqdn = data.get('old_fqdn')

            # return JsonResponse({'Result':fqdn, 'id':201})

            if ('.z6' in fqdn):
                fqdn = fqdn.replace('.z6', '.Z6')
            if ('.z6' in old_fqdn):
                old_fqdn = old_fqdn.replace('.z6', '.Z6')

            # dslam_obj = DSLAM.objects.get(name=port_data.get('dslam_name'), telecom_center__name=port_data.get('telecom_center_name'))
            try:
                dslam_obj = DSLAM.objects.get(fqdn=fqdn)
                old_dslam_obj = DSLAM.objects.get(fqdn=old_fqdn)

                # return JsonResponse({'Result': serialize('json',[dslam_obj])}, status=status.HTTP_200_OK)
                telecomId = dslam_obj.telecom_center_id
                cityId = TelecomCenter.objects.get(id=telecomId).city_id
            except ObjectDoesNotExist as ex:
                try:
                    if ('.Z6' in fqdn):
                        fqdn = fqdn.replace('.Z6', '.z6')
                        dslam_obj = DSLAM.objects.get(fqdn=fqdn)
                        # return JsonResponse({'Result': dslam_obj.fqdn}, status=status.HTTP_200_OK)
                        telecomId = dslam_obj.telecom_center_id
                        cityId = TelecomCenter.objects.get(id=telecomId).city_id

                except ObjectDoesNotExist as ex:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    mail_info = Mail()
                    mail_info.from_addr = 'oss-notification@pishgaman.net'
                    mail_info.to_addr = 'oss-problems@pishgaman.net'
                    mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                        str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'),
                        port_data.get('card_number'), port_data.get('port_number'), ip)
                    mail_info.msg_subject = 'Error in RegisterPortAPIView'
                    Mail.Send_Mail(mail_info)
                    return JsonResponse({'Result': 'Dslam Not Found. Please check FQDN again.'},
                                        status=status.HTTP_200_OK)

            telecom_mdf_obj = TelecomCenterMDF.objects.filter(telecom_center_id=dslam_obj.telecom_center.id)
            if telecom_mdf_obj:
                telecom_mdf_obj = telecom_mdf_obj.first()

            mdf_dslam_obj, mdf_dslam__updated = MDFDSLAM.objects.update_or_create(
                telecom_center_id=dslam_obj.telecom_center.id, telecom_center_mdf_id=telecom_mdf_obj.id,
                #### Check this whole line
                row_number=0, floor_number=0, connection_number=0,  ##### Check this whole line
                dslam_id=dslam_obj.id, slot_number=port_data.get('card_number'),
                port_number=port_data.get('port_number'),
                defaults={'status': mdf_status, 'identifier_key': identifier_key})
            # if mdf_dslam.status != 'FREE':
            #    return JsonResponse(
            #            {'result': 'port status is {0}'.format(mdf_dslam.status), 'id': -1}
            #            )
            # else:
            #    mdf_dslam.status = 'RESELLER'
            #    mdf_dslam.save()
            # identifier_key = mdf_dslam.identifier_key
        except ObjectDoesNotExist as ex:
            return JsonResponse({'result': str(ex), 'id': -1})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            try:
                return JsonResponse({'result': 'Error is {0}--{1}'.format(ex, exc_tb.tb_lineno)})
            except ObjectDoesNotExist as ex:
                mail_info = Mail()
                mail_info.from_addr = 'oss-notification@pishgaman.net'
                mail_info.to_addr = 'oss-problems@pishgaman.net'
                mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                    str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'), port_data.get('card_number'),
                    port_data.get('port_number'), ip)
                mail_info.msg_subject = 'Error in RegisterPortAPIView'
                Mail.Send_Mail(mail_info)
                return JsonResponse({'Result': ''}, status=status.HTTP_200_OK)
        try:
            reseller_obj, reseller_created = Reseller.objects.get_or_create(name=reseller_data.get('name'))
            print()
            'reseller', reseller_obj
            old_reseller_obj = Reseller.objects.get(name=old_reseller_data.get('name'))

            customer_obj, customer_updated = CustomerPort.objects.update_or_create(
                username=customer_data.get('username'),
                defaults={'identifier_key': identifier_key, 'telecom_center_id': mdf_dslam_obj.telecom_center_id}
            )

            rp = ResellerPort()
            rp.identifier_key = identifier_key
            rp.status = 'ENABLE'
            rp.dslam_id = dslam_obj.id
            rp.dslam_slot = port_data.get('card_number')
            rp.dslam_port = port_data.get('port_number')
            rp.reseller = reseller_obj
            rp.telecom_center_id = mdf_dslam_obj.telecom_center_id
            rp.save()

            vlan_objs = Vlan.objects.filter(reseller=reseller_obj)
            old_vlan_objs = Vlan.objects.filter(reseller=old_reseller_obj)

            print()
            'vlan_objs ->', vlan_objs
            # return JsonResponse({'vlan':  vlan_objs[0].vlan_id, 'vpi' : reseller_obj.vpi,'vci' : reseller_obj.vci})
            port_indexes = [{'slot_number': port_data.get('card_number'), 'port_number': port_data.get('port_number')}]
            old_port_indexes = [
                {'slot_number': old_port_data.get('card_number'), 'port_number': old_port_data.get('port_number')}]

            pishParams = {
                "type": "dslamport",
                "is_queue": False,
                "vlan_id": 3900,
                "vlan_name": 'pte',
                "dslam_id": old_dslam_obj.id,
                "port_indexes": old_port_indexes,
                "username": customer_data.get('username'),
                "vpi": 0,
                "vci": 35,
            }

            old_params = {
                "type": "dslamport",
                "is_queue": False,
                "vlan_id": old_vlan_objs[0].vlan_id,
                "vlan_name": old_vlan_objs[0].vlan_name,
                "dslam_id": old_dslam_obj.id,
                "port_indexes": old_port_indexes,
                "username": customer_data.get('username'),
                "vpi": old_reseller_obj.vpi,
                "vci": old_reseller_obj.vci,
            }

            params = {
                "type": "dslamport",
                "is_queue": False,
                "vlan_id": vlan_objs[0].vlan_id,
                "vlan_name": vlan_objs[0].vlan_name,
                "dslam_id": dslam_obj.id,
                "port_indexes": port_indexes,
                "username": customer_data.get('username'),
                "vpi": reseller_obj.vpi,
                "vci": reseller_obj.vci,
            }
            # return JsonResponse({'res' : params })
            result = ''

            try:
                if (cityId == 1637):
                    # return JsonResponse({'result': str('Email Sent To rt-network-access!!!'), 'id': 201}, status=status.HTTP_202_ACCEPTED)
                    q = "select * from shabdizrt where agent = '{0}' and ip='{1}' and fqdn='{2}' and card = '{3}' and port= '{4}' and status= '{5}'".format(
                        reseller_obj, dslam_obj.ip, dslam_obj.fqdn, port_data.get('card_number'),
                        port_data.get('port_number'), "sent")

                    cursor = connection.cursor()
                    cursor.execute(q)
                    rows = cursor.fetchall()
                    if (cursor.rowcount > 0):
                        return JsonResponse(
                            {'result': str('Email already has been sent to rt-access. Please wait for response.'),
                             'id': 201}, status=status.HTTP_202_ACCEPTED)

                    query = "INSERT INTO shabdizrt values('{0}','{1}', '{2}', '{3}', '{4}', '{5}')".format(reseller_obj,
                                                                                                           dslam_obj.ip,
                                                                                                           dslam_obj.fqdn,
                                                                                                           port_data.get(
                                                                                                               'card_number'),
                                                                                                           port_data.get(
                                                                                                               'port_number'),
                                                                                                           "sent")
                    # return JsonResponse({'result': str(cursor.rowcount), 'id': 201}, status=status.HTTP_202_ACCEPTED)

                    cursor = connection.cursor()
                    cursor.execute(query)

                    vpi = Reseller.objects.get(name=reseller_data['name']).vpi
                    vci = Reseller.objects.get(name=reseller_data['name']).vci
                    fromaddr = "oss-notification@pishgaman.net"
                    toaddr = "rt-network-access@pishgaman.net"
                    msg = MIMEMultipart()
                    msg['From'] = fromaddr
                    msg['To'] = toaddr
                    msg['Subject'] = "OSS Shabdiz Dslams"
                    body = 'Command: add to vlan for Shabdiz Dslams, IP: {0}, fqdn: {1} , Card: {2} , Port: {3} , VlanId: {4},Reseller: {5},Vpi: {6},Vci: {7}'.format(
                        dslam_obj.ip, dslam_obj.fqdn, port_data.get('card_number'), port_data.get('port_number'),
                        vlan_objs[0].vlan_id, reseller_data['name'], vpi, vci)
                    msg.attach(MIMEText(body, 'plain'))
                    server = smtplib.SMTP('mail.pishgaman.net', 25)
                    server.login("oss-notification@pishgaman.net", "Os123456")
                    text = msg.as_string()
                    server.sendmail(fromaddr, toaddr, text)
                    return JsonResponse({'result': str('Email Sent To rt-network-access!!!'), 'id': 201},
                                        status=status.HTTP_202_ACCEPTED)
                # elif(dslam_obj.dslam_type_id == 4):
                elif (dslam_obj.dslam_type_id == 4 or dslam_obj.dslam_type_id == 3):
                    try:
                        dslamType = ''
                        if (dslam_obj.dslam_type_id == 4):
                            dslamType = 'Fiberhome'
                            if (dslam_obj.id == 1231):
                                sourceVlanId = '2576'
                            else:
                                sourceVlanId = '3900'
                        elif (dslam_obj.dslam_type_id == 3):
                            sourceVlanId = '3900'
                            dslamType = 'FiberhomeAN3300'

                        vlanName = str(reseller_obj).split('-')[1]
                        if (vlanName == 'didehban'):
                            vlanName = 'dideban'
                        if (vlanName == 'baharsamaneh'):
                            vlanName = 'baharsam'
                        if (vlanName == 'badrrayan'):
                            vlanName = 'badrray'

                        url = 'http://5.202.129.88:9096/api/Telnet/telnet'
                        data = "{'type':'%s','dslam':'%s','telnetPort':'23','userName':'%s','password':'%s','access':'%s','sourceVlanId':'%s','vlanName':'%s','vlanId':'%s','untaggedPortList':'%s','vpiVci':'%s','card':'%s','port':'%s','command':'%s','terminalDelay':'600','requestTimeOut':'1500'}" % (
                            dslamType, dslam_obj.ip, dslam_obj.telnet_username, dslam_obj.telnet_password,
                            dslam_obj.access_name, sourceVlanId, vlanName, vlan_objs[0].vlan_id,
                            '0-{0}-{1}'.format(port_data.get('card_number'), port_data.get('port_number')),
                            '{0}/{1}'.format(Reseller.objects.get(name=reseller_data['name']).vpi,
                                             Reseller.objects.get(name=reseller_data['name']).vci),
                            port_data.get('card_number'), port_data.get('port_number'), 'addToVlan')
                        # return JsonResponse({'result':data })
                        fhresponse = requests.post(url, data=data, headers={"Content-Type": "application/json"})
                        sid = fhresponse.json()

                    except Exception as ex:
                        exc_type, exc_obj, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        mail_info = Mail()
                        mail_info.from_addr = 'oss-notification@pishgaman.net'
                        mail_info.to_addr = 'oss-problems@pishgaman.net'
                        mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                            str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'),
                            port_data.get('card_number'), port_data.get('port_number'), ip)
                        mail_info.msg_subject = 'Error in RegisterPortAPIView'
                        Mail.Send_Mail(mail_info)
                        return JsonResponse(
                            {'result': str('an error occurred. please try again. {0}'.format(str(exc_tb.tb_lineno)))})
                elif (dslam_obj.dslam_type_id == 5 or dslam_obj.dslam_type_id == 7):
                    vpi = Reseller.objects.get(name=reseller_data['name']).vpi
                    vci = Reseller.objects.get(name=reseller_data['name']).vci
                    fromaddr = "oss-notification@pishgaman.net"
                    toaddr = "rt-network-access@pishgaman.net"
                    msg = MIMEMultipart()
                    msg['From'] = fromaddr
                    msg['To'] = toaddr
                    msg['Subject'] = "OSS Problem"
                    body = 'Command: add to vlan, IP: {0}, fqdn: {1} , Card: {2} , Port: {3} , VlanId: {4},Reseller: {5},Vpi: {6},Vci: {7}'.format(
                        dslam_obj.ip, dslam_obj.fqdn, port_data.get('card_number'), port_data.get('port_number'),
                        vlan_objs[0].vlan_id, reseller_data['name'], vpi, vci)
                    msg.attach(MIMEText(body, 'plain'))
                    server = smtplib.SMTP('mail.pishgaman.net', 25)
                    server.login("oss-notification@pishgaman.net", "Os123456")
                    text = msg.as_string()
                    server.sendmail(fromaddr, toaddr, text)
                    # return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})
                    return JsonResponse(
                        {'result': str('An Error Ocurred for add to vlan Command. Email Sent To rt-network-access!!!'),
                         'id': 201},
                        status=status.HTTP_202_ACCEPTED)


                else:
                    res1 = utility.dslam_port_run_command(old_dslam_obj.id, 'delete from vlan', old_params)
                    res = utility.dslam_port_run_command(old_dslam_obj.id, 'add to vlan', pishParams)
                    result = utility.dslam_port_run_command(dslam_obj.id, 'add to vlan', params)
                    PVC = utility.dslam_port_run_command(dslam_obj.id, 'port pvc show', params)

                    # result2 = utility.dslam_port_run_command(dslam_obj.id, 'add to vlan', params)

            except Exception as ex:
                if (dslam_obj.dslam_type_id == 5 or dslam_obj.dslam_type_id == 3 or dslam_obj.dslam_type_id == 7):
                    vpi = Reseller.objects.get(name=reseller_data['name']).vpi
                    vci = Reseller.objects.get(name=reseller_data['name']).vci
                    fromaddr = "oss-notification@pishgaman.net"
                    toaddr = "rt-network-access@pishgaman.net"
                    msg = MIMEMultipart()
                    msg['From'] = fromaddr
                    msg['To'] = toaddr
                    msg['Subject'] = "OSS Problem"
                    cc = ['oss-problems@pishgaman.net']
                    body = 'Command: add to vlan, IP: {0}, fqdn: {1} , Card: {2} , Port: {3} , VlanId: {4},Reseller: {5},Vpi: {6},Vci: {7}'.format(
                        dslam_obj.ip, dslam_obj.fqdn, port_data.get('card_number'), port_data.get('port_number'),
                        vlan_objs[0].vlan_id, reseller_data['name'], vpi, vci)
                    msg.attach(MIMEText(body, 'plain'))
                    server = smtplib.SMTP('mail.pishgaman.net', 25)
                    server.login("oss-notification@pishgaman.net", "Os123456")
                    text = msg.as_string()
                    server.sendmail(fromaddr, toaddr, text)
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    # return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})
                    return JsonResponse(
                        {'result': str('An Error Ocurred for add to vlan Command. Email Sent To rt-network-access!!!')},
                        status=status.HTTP_202_ACCEPTED)

                else:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    mail_info = Mail()
                    mail_info.from_addr = 'oss-notification@pishgaman.net'
                    mail_info.to_addr = 'oss-problems@pishgaman.net'
                    mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                        str(exc_tb.tb_lineno), str(ex), reseller_data, fqdn,
                        port_data.get('card_number'), port_data.get('port_number'), ip)
                    mail_info.msg_subject = 'Error in RegisterPortAPIView'
                    Mail.Send_Mail(mail_info)
                    return JsonResponse(
                        {'result': str('an error occurred. please try again. {0}'.format(str(exc_tb.tb_lineno)))})

            if vlan_objs.count() > 0:
                try:
                    port_obj = DSLAMPort.objects.get(dslam=dslam_obj, slot_number=port_data.get('card_number'),
                                                     port_number=port_data.get('port_number'))
                except ObjectDoesNotExist as ex:
                    dslam_port = DSLAMPort()
                    dslam_port.dslam = dslam_obj
                    dslam_port.created_at = datetime.now()
                    dslam_port.updated_at = datetime.now()
                    dslam_port.slot_number = port_data.get('card_number')
                    dslam_port.port_number = port_data.get('port_number')
                    dslam_port.port_index = '{0}{1}'.format(port_data.get('card_number'), port_data.get('port_number'))
                    dslam_port.port_name = 'adsl{0}-{1}'.format(port_data.get('card_number'),
                                                                port_data.get('port_number'))
                    dslam_port.admin_status = 'UNLOCK'
                    dslam_port.oper_status = 'NO-SYNC'
                    dslam_port.line_profile = '768/10240/IR'
                    dslam_port.selt_value = ''
                    dslam_port.uptime = '18:55:20'
                    dslam_port.upstream_snr = '192'
                    dslam_port.downstream_snr = '230'
                    dslam_port.upstream_attenuation = '69'
                    dslam_port.downstream_attenuation = '188'
                    dslam_port.upstream_attainable_rate = '2434'
                    dslam_port.downstream_attainable_rate = '157'
                    dslam_port.upstream_tx_rate = '93'
                    dslam_port.downstream_tx_rate = '1250'
                    dslam_port.upstream_snr_flag = 'good'
                    dslam_port.downstream_snr_flag = 'excellent'
                    dslam_port.upstream_attenuation_flag = 'outstanding'
                    dslam_port.downstream_attenuation_flag = 'outstanding'
                    dslam_port.vpi = 0
                    dslam_port.vci = 35
                    dslam_port.save()
                    port_obj = DSLAMPort.objects.get(dslam=dslam_obj, slot_number=port_data.get('card_number'),
                                                     port_number=port_data.get('port_number'))
                port_vlan_obj = DSLAMPortVlan()
                port_vlan_obj.vlan = vlan_objs.first()
                port_vlan_obj = port_obj
                port_vlan_obj.save()
            if (dslam_obj.dslam_type_id == 4):
                if ('succeed' in sid):
                    return JsonResponse({'id': 201, 'res': sid, 'msg': 'port config has been done.'},
                                        status=status.HTTP_201_CREATED)
                else:
                    return JsonResponse({'result': 'Error', 'ErrorDesc': sid, 'id': 400, 'res': 'Error'},
                                        status=status.HTTP_400_BAD_REQUEST)
            if (dslam_obj.dslam_type_id == 3):
                if ('added to vlan' in sid):
                    return JsonResponse({'PVC': PVC, 'id': 201, 'res': sid, 'msg': 'port config has been done.'},
                                        status=status.HTTP_201_CREATED)
                else:
                    return JsonResponse({'result': 'Error', 'ErrorDesc': sid, 'id': 400, 'res': 'Error'},
                                        status=status.HTTP_400_BAD_REQUEST)
            if (dslam_obj.dslam_type_id == 1):
                if ('No results was returned' in PVC['result']):
                    return JsonResponse({'result': 'Error', 'ErrorDesc': PVC['result'], 'id': 400, 'res': 'Error'},
                                        status=status.HTTP_400_BAD_REQUEST)
                else:
                    return JsonResponse({'PVC': PVC, 'id': 201, 'res': sid, 'msg': 'port config has been done.'},
                                        status=status.HTTP_201_CREATED)
            else:
                return JsonResponse({'PVC': PVC, 'id': 201, 'res': sid, 'msg': 'port config has been done.'},
                                    status=status.HTTP_201_CREATED)
                # return JsonResponse({'result':'Port is registered', 'PVC': PVC , 'id': 201, 'res': sid}, status=status.HTTP_201_CREATED)
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            mail_info = Mail()
            mail_info.from_addr = 'oss-notification@pishgaman.net'
            mail_info.to_addr = 'oss-problems@pishgaman.net'
            mail_info.msg_body = 'Error in RegisterPortAPIView in Line {0}.Error Is: "{1}". Request is From {2} .fqdn = {3}, Cart = {4}, Port = {5},Ip: {6}'.format(
                str(exc_tb.tb_lineno), str(ex), reseller_data, port_data.get('fqdn'), port_data.get('card_number'),
                port_data.get('port_number'), ip)
            mail_info.msg_subject = 'Error in RegisterPortAPIView'
            Mail.Send_Mail(mail_info)

            # return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})
            return JsonResponse(
                {'result': str('an error occurred. please try again. {0}'.format(str(exc_tb.tb_lineno)))},
                status=status.HTTP_202_ACCEPTED)


"""from rtkit import set_logging
import logging

set_logging('debug')
logger = logging.getLogger('rtkit')

RT_REST_URL = 'https://ticket.service.pishgaman.net/REST/1.0/'
RT_USER = 'm.taher'
RT_PASSWORD = 'Aa123456'
RT_QUEUE_CEM = 19
RT_QUEUE_COMPLAINTS = 106
#############################
rt_resource = RTResource(RT_REST_URL, RT_USER, RT_PASSWORD, CookieAuthenticator)


class CreateTicketAPIView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        data = request.data
        try:
            ticket_id = submit_ticket_rt(data.get('subject'), data.get('queue', None), data.get('content'))
            return Response({'ticket_id': str(ticket_id)}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def submit_ticket_rt(subject, department, content):
    if department and department.lower() == 'sla':
        queue = 'SLA'
    else:
        queue = 'Complaint'
    content = {'content': {'Queue': queue, 'Subject': subject.replace('<br>', '\n').encode('utf-8'),
                           'Text': content.replace('<br>', '\n').encode('utf-8'), }}
    ticket_id = -1
    try:
        response = rt_resource.post(path='ticket/new', payload=content, )
        logger.info(response.parsed)
        logger.info('=================')
        ticket_id = response.parsed[0][0][1].split('/')[1]
    except RTResourceError as e:
        logger.error(e.response.status_int)
        logger.error(e.response.status)
        logger.error(e.response.parsed)
    return ticket_id


class GetTicketInfoAPIView(views.APIView):
    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request):
        data = request.data
        logger.debug('GetTicketInfoAPIView: %s' % str(data))
        ticket_id = data.get('ticket_id')
        t_info = get_ticket_info_rt(ticket_id)
        # t_info = get_ticket_info(ticket_id)
        return Response({'message': 'ok', 'ticket_info': t_info}, status=status.HTTP_200_OK)


def get_ticket_info_rt(ticket_id):
    ticket_id = '%s' % ticket_id
    # ticket_id = ticket_id[-4:]
    ticket_info = {'subject': '', 'posts': [], 'status': 'N/A', 'last_update': 'N/A'}
    try:
        response = rt_resource.get(path='ticket/%s/show' % ticket_id)
        print()
        response
        for r in response.parsed:
            for t in r:
                if t[0] == 'LastUpdated':
                    ticket_info['last_update'] = t[1]
                elif t[0] == 'Status':
                    ticket_info['status'] = t[1]
                elif t[0] == 'Subject':
                    ticket_info['subject'] = t[1]

                logger.info(t)
    except RTResourceError as e:
        print()
        e
        logger.error(e.response.status_int)
        logger.error(e.response.status)
        logger.error(e.response.parsed)

    reply_list = []
    try:
        response = rt_resource.get(path='ticket/%s/history?format=l' % ticket_id)
        for r in response.parsed:
            is_reply = False
            content = ''
            for t in r:
                if t[0] == 'Content':
                    content = t[1]
                elif t[0] == 'Type' and t[1] == 'Correspond':
                    is_reply = True
                logger.info(t)
            if is_reply and content:
                reply_list.append(content)
    except RTResourceError as e:
        print()
        e

    ticket_info['posts'] = reply_list
    print()
    '==============='
    print()
    ticket_info
    print()
    '==============='
    return ticket_info
"""


# Dana ticketing API
class AddTicketDanaAPIView(views.APIView):
    def post(self, request, format=None):
        data = request.data
        try:
            url = 'http://ticketing.pishgaman.net/api/v1/Token/GetToken'
            r_data = "{'username':'%s','password':'%s','secret':'%s'}" % (
                data['username'], data['password'], 'F7409BC306790047040899006710AF2D93E6A317')
            # return JsonResponse({'response':data})

            fhresponse = requests.post(url, data=r_data, headers={"Content-Type": "application/json"})
            sid = fhresponse.json()
            if (sid['ResultMessageList'] == []):
                bearer = sid['ResultData']['access_token']
                url = "http://ticketing.pishgaman.net/api/v1/Ticket"
                if (data['username'] == 'SLA'):
                    payload = "{'TemplateAttributeId':'e659e12c-0fac-49d2-b752-06a8d156cfa8','SaveMethod':'69e0b98b-185b-434f-be3b-d5d2f550a0ee','Status':'28cb76a6-d999-4ee3-887c-66792287453d','ApplicantId':'4b8e7077-d7f7-4137-a431-f9c3f14d24db','Title':'%s','DoneDate':null,'RecordColor':null,'LastResponseDate':null,'ITStaffId':null,'ProblemManagementId':null,'AssetId':null,'IsRemoved':null,'ITStaffGroupID':'b994c735-4b44-417c-af79-54d119efeb48','Type':null,'SatifacationRate':null,'InfluenceLevel':null,'GroupId':null,'Urgency':null,'DueDate':null,'Priority':null,'TicketBody':{'TicketBody':{'Body':'%s','IsPublic':true}}}" % (
                        data['Title'].encode('utf-8'), data['Body'].encode('utf-8'))
                elif (data['username'] == 'Complaint'):
                    payload = "{'TemplateAttributeId':'e659e12c-0fac-49d2-b752-06a8d156cfa8','SaveMethod':'69e0b98b-185b-434f-be3b-d5d2f550a0ee','Status':'28cb76a6-d999-4ee3-887c-66792287453d','ApplicantId':'bfc9a9cc-50ce-4d70-b5ee-d1f3b5d3163c','Title':'%s','DoneDate':null,'RecordColor':null,'LastResponseDate':null,'ITStaffId':null,'ProblemManagementId':null,'AssetId':null,'IsRemoved':null,'ITStaffGroupID':'3065480b-067e-436b-815c-93bc37b39f5e','Type':null,'SatifacationRate':null,'InfluenceLevel':null,'GroupId':null,'Urgency':null,'DueDate':null,'Priority':null,'TicketBody':{'TicketBody':{'Body':'%s','IsPublic':true}}}" % (
                        data['Title'].encode('utf-8'), data['Body'].encode('utf-8'))
                else:
                    return JsonResponse({'response': 'User Not Found.'})

                headers = {
                    'Content-Type': 'application/json',
                    'Authorization': 'bearer ' + bearer
                }
                response = requests.request("POST", url, headers=headers, data=payload)
                response_json = response.json()
                ticket_guid = response_json['ResultId']

                dana_ticket_info = get_ticket_info_dana(ticket_guid, bearer)
                return JsonResponse({'response': response_json,
                                     'TicketNumber': dana_ticket_info['ResultData']['result'][0]['TicketNumber']},
                                    status=status.HTTP_200_OK)
            else:
                return JsonResponse({'response': 'false'})


        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return Response({'message': str(ex) + "  " + str(exc_tb.tb_lineno)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def get_ticket_info_dana(ticket_guid, bearer):
    url = 'http://ticketing.pishgaman.net/api/v1/Ticket/' + ticket_guid
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'bearer ' + bearer
    }

    response = requests.get(url, headers=headers)
    response_json = response.json()
    return response_json


class GetTicketDetailsDanaAPIView(views.APIView):
    def post(self, request, format=None):
        data = request.data
        try:
            ticket_id = data['ticket_id']
            token_url = 'http://ticketing.pishgaman.net/api/v1/Token/GetToken'
            r_data = "{'username':'%s','password':'%s','secret':'%s'}" % (
                data['username'], data['password'], data['secret'])
            r_response = requests.post(token_url, data=r_data, headers={"Content-Type": "application/json"})
            sid = r_response.json()
            if (sid['ResultMessageList'] == []):
                bearer = sid['ResultData']['access_token']
                url = "https://ticketing.pishgaman.net/api/v1/Ticket/" + str(ticket_id)
                response = requests.get(url, headers={"Content-Type": "application/json",
                                                      'Authorization': 'bearer ' + bearer})
                res_json = response.json()
                ticket_guid = res_json['ResultData'].get('result')[0]['TicketId']
                ticket_details = get_ticket_detail_dana(ticket_guid, bearer)
            return JsonResponse({'response': ticket_details['ResultData']})

            return JsonResponse({'ticket_id': data['ticket_id']})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return Response({'message': str(ex) + "  " + str(exc_tb.tb_lineno)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def get_ticket_detail_dana(ticket_guid, bearer):
    url = 'https://ticketing.pishgaman.net/api/v1/Tickets/GetTicketBodyWithAttachment/{0}'.format(ticket_guid)
    response = requests.get(url, headers={"Content-Type": "application/json", 'Authorization': 'bearer ' + bearer})
    response_json = response.json()
    return response_json


def get_user_info_from_ibs(username):
    login_url = 'http://172.28.163.22:8280/authenticate/v1/requesttoken'
    login_data = "{'Username':'oss','Password':'74e#$pRe;F'}"
    response = requests.post(login_url, data=login_data, headers={"Content-Type": "application/json",
                                                                  "Authorization": "Bearer f6eccc80-1667-3a5e-88f1-089ec684cc4d"})
    response_json = response.json()
    getuserinfo_url = 'http://172.28.163.22:8280/ibs/v1/getuserinfo'
    getuserinfo_data = '{"NormalUsername":%s}' % (str(username))
    # return getuserinfo_data

    userinfo_response = requests.post(getuserinfo_url, data=getuserinfo_data,
                                      headers={"Content-Type": "application/json",
                                               "Authorization": "Bearer f6eccc80-1667-3a5e-88f1-089ec684cc4d",
                                               "token": '' + response_json['token']})
    userinfo_response_json = userinfo_response.json()
    for key in userinfo_response_json['result']:
        if userinfo_response_json['result'][key]['online_status'] == False:
            return "Error : this user is offline"
    if (userinfo_response_json['error'] and 'does not exists' in userinfo_response_json['error']):
        return "Error" + str(userinfo_response_json['error'])
    for value in userinfo_response_json['result']:
        return userinfo_response_json['result'][value]['attrs']['limit_mac']


class DSLAMRunICMPCommandByFqdnView(views.APIView):

    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        self.user = request.user
        data = request.data
        icmp_type = data.get('icmp_type')
        params = data.get('params')
        fqdn = request.data.get('fqdn')

        if ('.z6' in fqdn):
            try:
                fqdn = fqdn.replace('.z6', '.Z6')
                dslamObj = DSLAM.objects.get(fqdn=fqdn)
                dslam_id = dslamObj.id

            except ObjectDoesNotExist as ex:
                try:
                    if ('.Z6' in fqdn):
                        fqdn = fqdn.replace('.Z6', '.z6')
                    dslamObj = DSLAM.objects.get(fqdn=fqdn)
                    dslam_id = dslamObj.id
                    dslam_ip = dslamObj.ip
                except Exception as ex:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    mail_info = Mail()
                    mail_info.from_addr = 'oss-notification@pishgaman.net'
                    mail_info.to_addr = 'oss-problems@pishgaman.net'
                    mail_info.msg_body = 'Error in RunCommandAPIViewin/Fiberhome2200 Line {0}.Error Is: {1}. fqdn = {2},card = {3}, port = {4}, command = {5}, subscriber = {6}. IP: {7}'.format(
                        str(exc_tb.tb_lineno), str(ex), request.data.get('fqdn'),
                        request.data.get('params').get('port_conditions').get('slot_number'),
                        request.data.get('params').get('port_conditions').get('port_number'), command, subscriber, ip)
                    mail_info.msg_subject = 'Error in RunCommandAPIView'
                    Mail.Send_Mail(mail_info)
                    return JsonResponse({'Result': ''}, status=status.HTTP_200_OK)

        result = utility.run_icmp_command(dslam_id, icmp_type, params)

        dslam_name = DSLAM.objects.get(id=dslam_id).name

        description = 'Run {0} Command on DSLAM {1} with params: {2}'.format(icmp_type, dslam_name, params)
        add_audit_log(request, 'DSLAMICMP', None, 'Run ICMP Command on DSLAM', description)

        return Response({'result': result}, status=status.HTTP_201_CREATED)


def get_device_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


class CheckNetworkBulkAvailability(views.APIView):
    def get_permissions(self):
        return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            params = data.get('params')
            dslam_obj = DSLAM.objects.get(id=74)
            dslam_objs = DSLAM.objects.all().order_by('last_sync_duration', 'created_at', 'last_sync')
            for dslam_obj in dslam_objs:
                result = utility.run_icmp_command(dslam_obj.id, 'ping', params)
                if (result['received'] == '0' and result['packet_loss'] == '100'):
                    query = "INSERT INTO device_availability_icmp values('{0}','{1}', '{2}', '{3}', '{4}', '{5}')".format(
                        dslam_obj.ip,
                        dslam_obj.fqdn,
                        'DSLAM', result['received'], result['packet_loss'], 'time_Out')
                elif (result['received'] == '4' and result['packet_loss'] == '0'):
                    query = "INSERT INTO device_availability_icmp values('{0}','{1}', '{2}', '{3}', '{4}', '{5}')".format(
                        dslam_obj.ip,
                        dslam_obj.fqdn,
                        'DSLAM', result['received'], result['packet_loss'], 'accessible')
                else:
                    query = "INSERT INTO device_availability_icmp values('{0}','{1}', '{2}', '{3}', '{4}', '{5}')".format(
                        dslam_obj.ip,
                        dslam_obj.fqdn,
                        'DSLAM', result['received'], result['packet_loss'], 'unknown')

                cursor = connection.cursor()
                cursor.execute(query)
            return Response({'result': "OK"}, status=status.HTTP_201_CREATED)
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return Response({'message': str(ex) + "  " + str(exc_tb.tb_lineno)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)





class CheckPortConflict(views.APIView):

    # def get_permissions(self):
    #  return (permissions.IsAuthenticated(),)

    def post(self, request, format=None):
        try:
            data = request.data
            username = data.get('username')
            fqdn = data.get('fqdn')
            slot = data.get('slot')
            port = data.get('port')
            dslamObj = DSLAM.objects.get(fqdn=str(fqdn).lower())
            ibs_cls = GetDataFromIBS()
            ibs_obj = ibs_cls.get_user_info_from_ibs(username)
            is_online = ibs_obj.get('is_online')
            if is_online:
                mac = ibs_obj.get('mac')
                print(mac)
                if 'Error' in mac:
                    return JsonResponse({'Error': mac})

                params = param2()
                params.type = 'dslamport'
                params.is_queue = False
                params.mac = mac
                params.fqdn = fqdn
                params.command = 'show port with mac'
                params.port_conditions = port_condition2()
                params.port_conditions.slot_number = slot
                params.port_conditions.port_number = port
                params_json = json.dumps(params, default=lambda x: x.__dict__)
                result = utility.dslam_port_run_command(dslamObj.pk, params.command, json.loads(params_json))
                if result.get('status') == 200:
                    dslam_card = result['result'].split(',')[0].split()[-1]
                    dslam_port = result['result'].split(',')[1].split()[-1]
                    if dslam_card == str(slot) and dslam_port == str(port):
                        return JsonResponse({'response': {'result': f'IBS-MAC: {ibs_obj.get("mac")}\\n'
                                                                    f'This port does not have Conflict'}},
                                            status=status.HTTP_200_OK)
                    else:
                        msg = f"IBS-MAC: {ibs_obj.get('mac')}\\n" \
                              f"This port has conflict , the slot/port entered is: {slot}/{port} but correct slot/port is : {dslam_card}/{dslam_port}"
                        mail = Mail()
                        mail.from_addr = 'oss-notification@pishgaman.net'
                        mail.to_addr = 'M.Mohammadian@pishgaman.net', 'F.azad@pishgaman.net'
                        mail.msg_body = msg
                        mail.msg_subject = 'check_conflict'
                        Mail.Send_Mail(mail)
                        return JsonResponse({'response': {'result': msg}},
                                            status=status.HTTP_200_OK)

                else:
                    return JsonResponse({'response': result}, status=status.HTTP_200_OK)
            else:
                return JsonResponse({'response': {'result': ibs_obj.get('error')}}, status=status.HTTP_200_OK)

        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return Response({'message': str(ex) + "  // " + str(exc_tb.tb_lineno)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class LoadDslamPorts(views.APIView):

    def get(self, request, format=None):
        try:
            data = request.data
            print(request)
            dslam_id = request.query_params['dslam_id']
            slot_count = request.query_params['slot_count']
            port_count = request.query_params['port_count']
            del_query = 'DELETE from "portInfo"'
            cursor = connection.cursor()
            cursor.execute(del_query)
            query = 'INSERT INTO "public"."portInfo"("Card", "Port", "dslam_id") select card.c, port.p,{0} from (select generate_series(1, {1}) as c) card cross join (select generate_series(1, {2}) as p) port'.format(
                dslam_id, slot_count, port_count)
            cursor = connection.cursor()
            cursor.execute(query)
            fd = open('dslam/insert_dslam_port.sql', 'r')
            sqlFile = fd.read()
            fd.close()
            insert_query = sqlFile
            cursor = connection.cursor()
            cursor.execute(insert_query)
            return JsonResponse({'row': port_count})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return Response({'message': str(ex)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GetDslamPorts(views.APIView):

    def post(self, request, format=None):
        try:
            data = request.data
            dslam_id = data.get('dslam_id')
            slot = data.get('slot_number')
            port = data.get('port_number')
            dslamObj = DSLAM.objects.get(id=dslam_id)
            params = param2()
            params.type = 'dslamport'
            params.is_queue = False
            params.fqdn = dslamObj.fqdn
            params.command = 'lcman show'
            params.port_conditions = port_condition2()
            params.port_conditions.slot_number = slot
            params.port_conditions.port_number = port
            params_json = json.dumps(params, default=lambda x: x.__dict__)
            result = utility.dslam_port_run_command(dslamObj.pk, params.command, json.loads(params_json))
            return JsonResponse({'row': result})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            exc_type, exc_obj, exc_tb = sys.exc_info()
            return str(ex) + "  // " + str(exc_tb.tb_lineno)


class DslamCommandsV2APIView(views.APIView):  # 111111111111

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, format=None):
        device_ip = str(get_device_ip(request))
        data = request.data
        command = data.get('command', None)
        command = command_recognise(command)
        dslam_id = request.data.get('dslam_id')
        dslamObj = DSLAM.objects.get(id=dslam_id)
        params = data.get('params', None)
        params['device_ip'] = device_ip
        dslam_type = dslamObj.dslam_type_id
        fqdn = dslamObj.fqdn
        log_port_data = f"{fqdn}/{params['port_conditions']['slot_number']}/{params['port_conditions']['port_number']}"
        log_username = params.get('username')
        log_date = datetime.now()
        operator = params.get('operator')
        try:
            result = utility.dslam_port_run_command(dslamObj.pk, command, params)
            log_status = True if isinstance(result, dict) else False
            log_params = PortmanLogging.prepare_variables(self, log_port_data, log_username, command, result, log_date,
                                                          device_ip, 'Run Command', log_status, operator, '')
            PortmanLogging(result, log_params)
            # if dslam_type == 1:  ################################### zyxel ###################################
            #     return JsonResponse({'Result': dslam_type})
            if dslam_type == 1:  ################################### zyxel ###################################
                if not command:
                    return Response({'result': 'Command does not exits'}, status=status.HTTP_400_BAD_REQUEST)
                if result:
                    if 'Busy' in result:
                        return Response({'response': result}, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        description = 'Run Command {0} on DSLAM {1}'.format(command, dslamObj.name)

                        add_audit_log(request, 'DSLAMCommand', None, 'Run Command On DSLAM Port', description)

                return JsonResponse({'response': result})

            elif dslam_type == 2:  # huawei
                return JsonResponse({'response': result, 'DslamType': 'huawei'})
            elif dslam_type == 3:  ############################## fiberhomeAN3300 ##############################
                return JsonResponse({'response': result, 'DslamType': 'fiberhomeAN3300'})

            elif dslam_type == 4:  ############################## fiberhomeAN2200 ##############################

                return JsonResponse({'response': result})

            elif dslam_type == 5:  ############################## fiberhomeAN5006 ##############################
                if command == 'show mac':
                    from .services.fiber_home_get_card import FiberHomeGetCardStatusService
                    cards_status_class = FiberHomeGetCardStatusService(request.data)
                    cards_status = cards_status_class.get_status_cards()
                    params['cards_status'] = cards_status
                    print(params)
                    result = utility.dslam_port_run_command(dslamObj.pk, command, params)
                    return JsonResponse({'response': result, 'DslamType': 'fiberhomean5006'})
                if command == 'Show VLAN':
                    return JsonResponse({'response': result.split("\\r\\n"), 'DslamType': 'fiberhomeAN5006'})
                return JsonResponse({'response': result, 'DslamType': 'fiberhomeAN5006'})

            elif dslam_type == 7:  ########################### zyxel1248 ##########################

                port_info = utility.dslam_port_run_command(dslamObj.pk, 'show profile by port', params)

                current_user_profile = ''

                if 'profile_name' not in port_info:
                    return JsonResponse(

                        {'response': result, 'current_user_profile': '', 'DslamType': 'zyxel1248'})

                current_user_profile = port_info['profile_name']

                return JsonResponse(

                    {'response': result, 'current_user_profile': current_user_profile, 'DslamType': 'zyxel1248'})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            log_params = PortmanLogging.prepare_variables(self, log_port_data, log_username, command, '', log_date,
                                                          device_ip, 'Run Command', False,
                                                          operator, str(ex) + '/' + str(exc_tb.tb_lineno))
            PortmanLogging('', log_params)
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})


class UploadRentedPort(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, forman=None):
        # Read excel file items
        excel_file = request.FILES['file']
        if ".xlsx" not in excel_file.name:
            return Response({"result": "File is not a excel(xlsx). Please upload only xlsx files."})
        wb = openpyxl.load_workbook(excel_file)
        for sheet in wb.sheetnames:
            worksheet = wb[sheet]

            excel_data = list()
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
            try:
                i = 1
                for val in excel_data[1:]:
                    rented_port = Rented_port()
                    rented_port.agent_name = val[0]
                    rented_port.city_name = val[1]
                    rented_port.telecom_name = val[2]
                    rented_port.dslam_number = val[3]
                    rented_port.card = val[4]
                    rented_port.port = val[5]
                    rented_port.telco_row = val[6]
                    rented_port.telco_column = val[7]
                    rented_port.telco_connection = val[8]
                    rented_port.save()
                    i = i + 1
            except Exception as ex:
                # return Response({"result": str(ex) + str(i), "sheet": sheet})
                return Response({"result": "Excel format has a problem.", "sheet": sheet, "row": str(i)})

        return Response({"result": "Upload Completed."})


class RentedPortAPIView(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, format=None):
        device_ip = get_device_ip(request)
        data = request.data
        command = data.get('command', None)
        command = command_recognise(command)
        params = data.get('params', None)
        try:
            dslam_obj = Rented_port.objects.get(agent_name=data.get('agent_name'),
                                                city_name=data.get('city_name'),
                                                telecom_name=data.get('telecom_name'),
                                                dslam_number=data.get('dslam_number'),
                                                card=params['port_conditions']['slot_number'],
                                                port=params['port_conditions']['port_number'])
        except ObjectDoesNotExist:
            return Response({"Result": "You don't have access to this Card & Port. Please check parameters."})
        fqdn = dslam_obj.fqdn
        dslamObj = DSLAM.objects.get(fqdn=fqdn.lower())
        dslam_type = dslamObj.dslam_type_id
        try:
            result = utility.dslam_port_run_command(dslamObj.pk, command, params)
            # if dslam_type == 1:  ################################### zyxel ###################################
            #     return JsonResponse({'Result': dslam_type})
            if dslam_type == 1:  ################################### zyxel ###################################
                if not command:
                    return Response({'result': 'Command does not exits'}, status=status.HTTP_400_BAD_REQUEST)
                if result:
                    if 'Busy' in result:
                        return Response({'result': result}, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        description = 'Run Command {0} on DSLAM {1}'.format(command, dslamObj.name)

                        add_audit_log(request, 'DSLAMCommand', None, 'Run Command On DSLAM Port', description)
                port_info = utility.dslam_port_run_command(dslamObj.pk, 'port Info', params)
                current_userProfile = ""
                port_state = ""
                for item in port_info['result']:
                    if 'prof' in item and 'alarm prof' not in item:
                        current_userProfile = item.split(':')[1]
                    if 'state' in item:
                        port_state = item.split(':')[1]
                        break

                return JsonResponse(
                    {'response': result, 'current_userProfile': current_userProfile, 'port_state': port_state})

            elif dslam_type == 2:  # huawei

                return JsonResponse({'response': result, 'DslamType': 'huawei'})

            elif dslam_type == 3:  ############################## fiberhomeAN3300 ##############################

                return JsonResponse({'response': result, 'DslamType': 'fiberhomeAN3300'})

            elif dslam_type == 4:  ############################## fiberhomeAN2200 ##############################

                return JsonResponse({'response': result})

            elif dslam_type == 5:  ############################## fiberhomeAN5006 ##############################

                return JsonResponse({'response': result, 'DslamType': 'fiberhomeAN5006'})

            elif dslam_type == 7:  ########################### zyxel1248 ##########################
                return JsonResponse({'response': dslam_type})
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})


class GetUserPortInfoFromPartakAPIView(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def get(self, request, format=None):
        try:
            username = request.query_params.get('username', None)
            url = 'https://my.pishgaman.net/api/pte/getCustomerInfo?Adsltel={0}'.format(username)
            url_response = requests.get(url, headers={"Content-Type": "application/json"}, verify=False)
            response = url_response.json()
            return Response(response, status=status.HTTP_200_OK)
        except Exception as ex:
            print(ex)
            return Response(str(ex), status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GetDslamIdByFqdnAPIView(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def get(self, request, format=None):
        try:
            fqdn = request.query_params.get('fqdn', None)
            dslam_id = DSLAM.objects.get(Q(fqdn__iexact=str(fqdn).lower())).id
            return JsonResponse({'dslam_id': dslam_id}, status=status.HTTP_200_OK)
        except Exception as ex:
            print(ex)
            return JsonResponse({'response': str(ex)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





class GetDSLAMIdByIPAPIView(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def get(self, request, format=None):
        try:
            ip = request.query_params.get('ip', None)
            query = "select id from dslam_dslam where ip='{}'".format(ip)
            cursor = connection.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            if cursor.rowcount > 0:
                return JsonResponse({'dslam_id': rows[0]}, status=status.HTTP_200_OK)
            else:
                return JsonResponse({'response': 'No matching dslam with this IP were found.'},
                                    status=status.HTTP_404_NOT_FOUND)
        except Exception as ex:
            print(ex)
            return JsonResponse({'response': str(ex)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class SearchFqdnsAPIView(views.APIView):
    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def get(self, request, format=None):
        try:
            fqdn = request.GET.get('fqdn', None)
            portman_fqdns = DSLAM.objects.filter(fqdn__icontains=fqdn).values('fqdn')
            return Response({"portman_fqdn": portman_fqdns},
                            status=status.HTTP_200_OK)
        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return JsonResponse({'row': str(ex) + '////' + str(exc_tb.tb_lineno)})


class NGNRegisterAPIView(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, format=None):
        device_ip = get_device_ip(request)
        data = request.data
        print('============================================================================')
        print(device_ip)
        print(data)
        print('============================================================================')
        command = data.get('command', None)
        command = command_recognise(command)
        dslam_id = request.data.get('dslam_id')
        dslamObj = DSLAM.objects.get(id=dslam_id)
        params = data.get('params', None)
        dslam_type = dslamObj.dslam_type_id
        log_port_data = f"{dslamObj.fqdn}/{params['port_conditions']['slot_number']}/{params['port_conditions']['port_number']}"
        log_username = params.get('username')
        log_date = datetime.now()
        try:
            if dslam_type != 2:
                return JsonResponse({'result': 'DSLAM Type Does Not Support This Command.'},
                                    status=status.HTTP_400_BAD_REQUEST)
            result = ngn_registaration_runCommands(dslamObj, command, params)
            return JsonResponse({'result': result}, status=status.HTTP_201_CREATED)

        except Exception as ex:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            log_params = PortmanLogging.prepare_variables(self, log_port_data, log_username, command, '', log_date,
                                                          device_ip, 'NGN Register', False,
                                                          str(ex) + '/' + str(exc_tb.tb_lineno), '')
            PortmanLogging('', log_params)
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            return JsonResponse({'result': 'Error is {0}'.format(ex), 'Line': str(exc_tb.tb_lineno)})


class DSLAMPortSnapshotSerializer(serializers.ModelSerializer):
    class Meta:
        model = DSLAMPortSnapshot
        fields = ('id', 'dslam_id', 'snp_date', 'downstream_snr_flag', 'upstream_snr_flag',
                  'upstream_attenuation_flag', 'downstream_attenuation_flag')


class DSLAMPortSnapshotViewSet(mixins.ListModelMixin,
                               mixins.RetrieveModelMixin,
                               viewsets.GenericViewSet):
    serializer_class = DSLAMPortSnapshotSerializer
    permission_classes = (IsAuthenticated,)
    queryset = DSLAMPortSnapshot.objects.all()

    def get_queryset(self):
        dslam_fqdn = self.request.query_params.get('fqdn', None)
        dslamObj = DSLAM.objects.get(fqdn=str(dslam_fqdn).lower())
        dslam_id = dslamObj.pk
        slot_number = self.request.query_params.get('slot_number', None)
        port_number = self.request.query_params.get('port_number', None)
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)
        queryset = self.queryset
        if dslam_fqdn and slot_number and port_number and start_date and end_date:
            print(self.request.query_params.get('fqdn', None))
            print(self.request.query_params.get('slot_number', None))
            print(self.request.query_params.get('port_number', None))
            print(self.request.query_params.get('start_date', None))
            print(self.request.query_params.get('end_date', None))
            queryset = queryset.filter(dslam_id=dslam_id,
                                       snp_date__range=(start_date + ' ' + '00:00:00', end_date + ' ' + '23:59:59'),
                                       slot_number=slot_number, port_number=port_number)
        return queryset


class AddToPishgamanAPIView(views.APIView):
    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request):
        data = request.data
        # if data['port'].get('old_vlan_name', None) is None and 'an2200' in  data['port'].get('fqdn', None):
        #     return Response(dict(results="The old_vlan_name field is required."), status=status.HTTP_400_BAD_REQUEST)

        ip = get_device_ip(request)
        pte = AddToPishgamanService(data, ip)
        result = pte.add_to_pishgaman()
        response = check_status(result)
        return response


def ngn_registaration_runCommands(dslamObj, command, params):
    result = utility.dslam_port_run_command(dslamObj.pk, command, params)
    if command == 'ngn_register_port':
        if 'Failure' in result['result']:
            if 'IP existed already' in result['result']:
                return dict({'result': result, 'msg': 'IP existed already', 'code': 10})
            else:
                return dict({'result': result, 'msg': 'Error', 'code': 20})
        return dict({'result': result, 'msg': result, 'code': 0})
    elif command == 'sip_configuration':
        return result
    elif command == 'assign_number_to_user':
        return result
    elif command == 'reset_sip_configuration':
        return result
    elif command == 'display_if_sip_attribute_running':
        return re.sub(r'-{2,}\sMore\s\(\s\w+\s\'\w\'\s\w+\s\w+\s\)\s\S+|\S+\W\d+\w', '', result['result'])
    elif command == 'display_sippstnuser_reg_state':
        return result
    elif command == 'display_sippstnuser_call_state':
        return result
    elif command == 'ngn_board_confirm':
        return result
    elif command == 'ngn_delete_number':
        return result
    elif command.lower() == 'show shelf':
        return result          
    else:
        return JsonResponse({'result': 'Command is not allowed for ngn section'}, status=status.HTTP_400_BAD_REQUEST)


class GetDslamPortSnapShotAPIView(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, format=None):
        try:
            return
        except Exception as ex:
            return ex


class GetSNMPPortStatusAPIView(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, format=None):
        try:
            device_ip = get_device_ip(request)
            data = request.data
            port_params_obj = GetPortStatusParamsService(data, device_ip)
            result = port_params_obj.get_port_params()
            response = check_status(result)
            return response

        except Exception as ex:
            return ex


class GetSNMPPortTrafficAPIView(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, format=None):
        try:
            device_ip = get_device_ip(request)
            data = request.data
            port_traffic_obj = GetPortTrafficService(data, device_ip)
            result = port_traffic_obj.get_port_params()
            response = check_status(result)
            return response

        except Exception as ex:
            return ex


class UserStatusAPIView(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, format=None):
        try:
            data = request.data
            username = data.get('username')
            ibs_service = GetDataFromIBS()
            user_status = ibs_service.get_user_info_from_ibs(username)
            return JsonResponse(dict(
                result=user_status.get('is_online'), 
                mac=user_status.get('mac'),
                #info=user_status.get('info'),
                error=user_status.get('error'),
                status=user_status.get('status')
            ))
        except Exception as ex:
            print(ex)
            return ex

class UserMacAddressAPIView(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def get(self, request, format=None):
        try:
            data = request.GET
            username = data.get('username')
            from_date = data.get('from_date')
            from_time = data.get('from_time')
            to_date = data.get('to_date')
            to_time = data.get('to_time')
            ibs_service = GetDataFromIBS()
            success, macId = ibs_service.get_mac_address_from_ibs(username, from_date, from_time, to_date, to_time)
            return JsonResponse(dict(mac=macId))
        except Exception as ex:
            print(ex)
            return ex

class PartakUserStatusAPIView(views.APIView):

    def get_permissions(self):
        return permissions.IsAuthenticated(),

    def post(self, request, format=None):
        try:
            data = request.data
            username = data.get('username')
            partak_user_info_service = GetDataFromPartak()
            user_status = partak_user_info_service.get_user_info_from_partak(username)
            print(user_status)
            result = user_status['AcoountInfo']['isonline']
            return JsonResponse({'result': result, 'status': status.HTTP_200_OK})
        except Exception as ex:
            print(ex)
            return ex

